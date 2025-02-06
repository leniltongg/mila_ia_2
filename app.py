#Sistema_Mila_IA

import json
import sqlite3
import io
from flask import Flask, g, render_template, request, redirect, url_for, session
from flask_login import LoginManager, login_required, login_user, logout_user, current_user, UserMixin
import openai
from flask import jsonify
from werkzeug.security import check_password_hash
from werkzeug.security import check_password_hash, generate_password_hash
from flask import make_response
from flask_login import UserMixin
from datetime import datetime
from flask import Flask, render_template, jsonify, flash, redirect, url_for, request
from werkzeug.utils import secure_filename
import os
import json
import PyPDF2  # vamos usar este para extrair texto
import pdfplumber  # se precisar deste para outra funcionalidade
from werkzeug.security import generate_password_hash
from flask import send_file
import tempfile
from routes import conteudo_bp, professores_bp, alunos_bp, simulados_bp


# Configuração básica do app
import os
from dotenv import load_dotenv

# Carrega variáveis do arquivo .env
load_dotenv()

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'  # pasta onde os PDFs serão temporariamente salvos
# Certifique-se que a pasta existe
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])
app.secret_key = os.getenv('SECRET_KEY', 'chave-secreta-temporaria')
app.register_blueprint(conteudo_bp)
app.register_blueprint(professores_bp)
app.register_blueprint(alunos_bp)
app.register_blueprint(simulados_bp)

# Adicionando `enumerate` ao contexto Jinja2
@app.context_processor
def utility_processor():
    return dict(enumerate=enumerate)

@app.template_filter('chr')
def jinja_chr(value):
    try:
        return chr(value)
    except Exception:
        return ""


# Configuração do gerenciador de login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message = "Por favor, faça login para acessar esta página."

# Caminho para o banco de dados SQLite
DATABASE = "educacional.db"

# Configure a chave de API da OpenAI
openai.api_key="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"

# Modelo de usuário para Flask-Login
from flask_login import UserMixin

class User(UserMixin):
    def __init__(self, id, nome, tipo_usuario_id, escola_id, serie_id, turma_id, email, codigo_ibge):
        self.id = id
        self.nome = nome
        self.tipo_usuario_id = tipo_usuario_id
        self.escola_id = escola_id
        self.serie_id = serie_id
        self.turma_id = turma_id
        self.email = email
        self.codigo_ibge = codigo_ibge

    def get_id(self):
        return str(self.id)



@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT id, nome, tipo_usuario_id, escola_id, serie_id, turma_id, email, codigo_ibge
        FROM usuarios
        WHERE id = ?
    """, (user_id,))
    user_data = cursor.fetchone()

    if user_data:
        app.logger.info(f"Usuário carregado: {user_data}")
        return User(
            id=user_data[0],
            nome=user_data[1],
            tipo_usuario_id=user_data[2],
            escola_id=user_data[3],
            serie_id=user_data[4],
            turma_id=user_data[5],
            email=user_data[6],
            codigo_ibge=user_data[7]
        )
    app.logger.warning(f"Usuário com ID {user_id} não encontrado.")
    return None







# Função para obter a conexão com o banco de dados
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
    return db

# Função para obter a conexão com o banco de dados
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
    return db

def get_relatorio_rede():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT 
            escolas.nome AS escola,
            COUNT(alunos.id) AS total_alunos,
            AVG(simulados.desempenho) AS media_desempenho
        FROM escolas
        LEFT JOIN alunos ON escolas.id = alunos.escola_id
        LEFT JOIN simulados ON alunos.id = simulados.aluno_id
        GROUP BY escolas.id
    """)
    return cursor.fetchall()


# Função para buscar a escola associada ao usuário
def get_escola_alocada(user_id):
    """
    Busca a escola associada ao usuário logado no banco de dados.
    Retorna um dicionário com as informações da escola.
    """
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT e.id, e.nome_da_escola
        FROM escolas e
        JOIN usuarios u ON u.escola_id = e.id
        WHERE u.id = ?
    """, (user_id,))
    escola = cursor.fetchone()

    if escola:
        return {
            "id": escola[0],
            "nome": escola[1]
        }
    else:
        return None
    
def gerar_parecer_ia(relatorio):
    # Supondo que você tenha uma função `analise_ia` que recebe os dados do relatório
    # e retorna uma análise detalhada e sugestões para melhoria
    parecer = analise_ia(relatorio)
    return parecer


# Função para buscar tipos de ensino de uma escola (para uso interno)
def get_tipos_ensino(escola_id):
    """
    Busca os tipos de ensino associados a uma escola pelo ID.
    """
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT tipo_ensino
        FROM escolas
        WHERE id = ?
    """, (escola_id,))
    result = cursor.fetchone()

    if result and result[0]:
        return result[0].split(", ")
    return []




# Fecha a conexão com o banco de dados ao final de cada requisição
@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()


# Função para gerar perguntas com OpenAI
import openai
import re

import random

def gerar_perguntas(disciplina, assunto, quantidade, nivel, alternativas):
    prompt = f"""
    Crie {quantidade} perguntas de múltipla escolha sobre o assunto "{assunto}" na disciplina "{disciplina}".
    Nível de dificuldade: {nivel}.
    Cada pergunta deve ter {alternativas} alternativas.

    Formato esperado:
    1. Pergunta
    A) Opção 1
    B) Opção 2
    C) Opção 3
    D) Opção 4
    Resposta correta: (A, B, C, ou D)
    """

    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Você é um gerador de perguntas de múltipla escolha."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1500,
            temperature=0.7,
        )

        perguntas_raw = response['choices'][0]['message']['content']
        print(f"Resposta bruta da IA: {perguntas_raw}")

        import re
        perguntas = []
        blocos = perguntas_raw.strip().split("\n\n")

        for bloco in blocos:
            try:
                linhas = bloco.strip().split("\n")
                pergunta_texto = re.search(r"^\d+\.\s*(.+)", linhas[0]).group(1).strip()
                opcoes = [linha[3:].strip() for linha in linhas[1:alternativas + 1]]
                resposta_correta = re.search(r"Resposta correta:\s*\(?([ABCD])\)?", bloco, re.IGNORECASE).group(1).strip()

                perguntas.append({
                    "pergunta": pergunta_texto,
                    "opcoes": opcoes,
                    "resposta_correta": resposta_correta  # Certifique-se de que é apenas 'A', 'B', etc.
                })
            except (IndexError, AttributeError) as e:
                print(f"Erro ao processar bloco: {bloco}\nErro: {e}")
                continue

        if not perguntas:
            raise ValueError("Nenhuma pergunta válida foi encontrada na resposta da IA.")

        print(f"Perguntas processadas: {perguntas}")
        return perguntas

    except openai.error.OpenAIError as api_error:
        print(f"Erro ao acessar a API OpenAI: {api_error}")
        return None
    except Exception as e:
        print(f"Erro inesperado ao gerar perguntas: {e}")
        return None

def enviar_questoes_automaticamente(serie_id, codigo_ibge):
    """
    Envia automaticamente as questões do banco_questoes para alunos da série especificada
    em escolas vinculadas ao código IBGE (codigo_ibge).
    """
    db = get_db()
    cursor = db.cursor()

    # Buscar as questões do banco para a série especificada
    cursor.execute(
        """
        SELECT id, questao, alternativa_a, alternativa_b, alternativa_c, alternativa_d, alternativa_e, questao_correta
        FROM banco_questoes
        WHERE serie_id = ?
        """,
        (serie_id,),
    )
    questoes = cursor.fetchall()

    if not questoes:
        print(f"Nenhuma questão encontrada para a série ID {serie_id}.")
        return

    # Buscar alunos (tipo_usuario_id = 4) da série nas escolas vinculadas ao código IBGE
    cursor.execute(
        """
        SELECT usuarios.id, usuarios.nome, usuarios.email
        FROM usuarios
        JOIN escolas ON usuarios.escola_id = escolas.id
        WHERE usuarios.serie_id = ? AND usuarios.tipo_usuario_id = 4 AND escolas.codigo_ibge = ?
        """,
        (serie_id, codigo_ibge),
    )
    alunos = cursor.fetchall()

    if not alunos:
        print(f"Nenhum aluno encontrado para a série ID {serie_id} e cidade ID {codigo_ibge}.")
        return

    # Criar o simulado no banco
    cursor.execute(
        """
        INSERT INTO simulados (titulo, serie_id)
        VALUES (?, ?)
        """,
        (f"Simulado Automático Série {serie_id}", serie_id),
    )
    simulado_id = cursor.lastrowid

    # Associar as questões ao simulado
    for questao in questoes:
        cursor.execute(
            """
            INSERT INTO simulado_questoes (simulado_id, questao_id)
            VALUES (?, ?)
            """,
            (simulado_id, questao[0]),
        )

    # Associar o simulado aos alunos
    for aluno in alunos:
        cursor.execute(
            """
            INSERT INTO aluno_simulado (aluno_id, simulado_id)
            VALUES (?, ?)
            """,
            (aluno[0], simulado_id),
        )

    # Finalizar a operação
    db.commit()
    print(f"Simulado ID {simulado_id} enviado para {len(alunos)} alunos da série ID {serie_id}.")


# Inicializa o banco de dados
def init_db():
    with app.app_context():
        db = get_db()
        cursor = db.cursor()
        # Demais inicializações...

        # Tabela de disciplinas
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS disciplinas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL UNIQUE
            )
        """)

        # Outras tabelas existentes no sistema...
        # Tabela de assuntos
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS assunto (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                disciplina TEXT NOT NULL,
                assunto TEXT NOT NULL,
                professor_id INTEGER,
                FOREIGN KEY (professor_id) REFERENCES usuarios (id)
            )
        """)
        # Insere disciplinas básicas (caso não existam)
        disciplinas_iniciais = [
            "Português", "Matemática", "Ciências", "História", "Geografia"
        ]
        for disciplina in disciplinas_iniciais:
            cursor.execute("""
                INSERT OR IGNORE INTO disciplinas (nome) VALUES (?)
            """, (disciplina,))

        db.commit()
        print("Banco de dados inicializado!")

        # Criação da tabela `pontuacoes`
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS pontuacoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                aluno_id INTEGER NOT NULL,
                pontos INTEGER DEFAULT 0,
                FOREIGN KEY (aluno_id) REFERENCES alunos(id)
            )
        """)

        # Tabelas básicas
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                senha TEXT NOT NULL,
                tipo_usuario_id TEXT NOT NULL,
                escola_id INTEGER,
                FOREIGN KEY (escola_id) REFERENCES escolas(id)
            )
        """)
        # Demais tabelas (escolas, alunos, professores, etc.) conforme seu código original
        db.commit()
        print("Tabelas criadas/verificadas com sucesso!")

import re

# Função de validação de respostas
# Função de validação de respostas
def validar_resposta(pergunta, resposta_ia, assunto):
    import re
    numeros = re.findall(r"([\d]+(?:\.\d+)?)", pergunta)
    if numeros:
        try:
            numeros = [float(n) for n in numeros]
            if "adição" in assunto.lower():
                resultado_correto = sum(numeros)
            elif "subtração" in assunto.lower():
                resultado_correto = numeros[0] - sum(numeros[1:])
            elif "multiplicação" in assunto.lower():
                resultado_correto = 1
                for n in numeros:
                    resultado_correto *= n
            elif "divisão" in assunto.lower():
                resultado_correto = numeros[0]
                for n in numeros[1:]:
                    resultado_correto /= n
            else:
                # Para assuntos sem validação específica
                return True

            # Comparar com a resposta fornecida pela IA
            return abs(resultado_correto - float(resposta_ia)) < 0.01
        except Exception as e:
            print(f"[DEBUG] Erro ao validar resposta: {e}")
            return False
    return True


from datetime import datetime
import json
import openai
from flask import Flask, render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user

def gerar_perguntas_conhecimentos_gerais(quantidade=10, nivel="Médio"):
    prompt = f"""
    Gere {quantidade} perguntas de conhecimentos gerais no nível {nivel}.
    Cada pergunta deve incluir 4 alternativas (A, B, C, D) e uma resposta correta.
    Formato de exemplo:
    1. Qual é a capital da França?
    A) Paris
    B) Londres
    C) Roma
    D) Berlim
    Resposta correta: A
    """
    try:
        resposta = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Você é um gerador de simulados."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1000,
            temperature=0.7
        )
        perguntas_texto = resposta['choices'][0]['message']['content']

        perguntas = []
        for pergunta in perguntas_texto.split("\n\n"):
            linhas = pergunta.strip().split("\n")
            if len(linhas) < 6:  # Ignorar perguntas incompletas
                continue
            pergunta_texto = linhas[0]
            alternativas = [linha[3:] for linha in linhas[1:5]]
            resposta_correta = linhas[5].split(":")[-1].strip()
            perguntas.append({
                "pergunta": pergunta_texto,
                "opcoes": alternativas,
                "correta": resposta_correta
            })
        return perguntas
    except Exception as e:
        print(f"Erro ao gerar perguntas: {e}")
        return []


def criar_simulado_diario():
    db = get_db()
    cursor = db.cursor()

    # Verificar se já existe um simulado diário para hoje
    cursor.execute("SELECT id FROM simulado_diario WHERE data = date('now')")
    simulado_existente = cursor.fetchone()

    if simulado_existente:
        print("Simulado diário já existe para hoje.")
        return  # Não criar novamente

    # Criar o novo simulado diário
    cursor.execute("""
        INSERT INTO simulado_diario (data, perguntas)
        VALUES (date('now'), NULL)
    """)
    simulado_id = cursor.lastrowid

    # Selecionar 10 perguntas aleatórias para o novo simulado
    cursor.execute("""
        SELECT id, texto_pergunta, alternativa_a, alternativa_b, alternativa_c, alternativa_d, correta
        FROM perguntas_simulado
        ORDER BY RANDOM()
        LIMIT 10
    """)
    perguntas = cursor.fetchall()

    # Inserir perguntas associadas ao simulado
    for pergunta in perguntas:
        cursor.execute("""
            INSERT INTO perguntas_simulado (simulado_id, texto_pergunta, alternativa_a, alternativa_b, alternativa_c, alternativa_d, correta)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (simulado_id, pergunta[1], pergunta[2], pergunta[3], pergunta[4], pergunta[5], pergunta[6]))

    db.commit()
    print("Novo simulado diário criado com sucesso.")


# Rota protegida: home

@app.route("/")
def index():
    print("Função 'index' chamada!")
    return redirect(url_for("login"))

@app.route("/home")
@login_required
def home():
    # Verificar o tipo de usuário
    tipo_usuario_id = current_user.tipo_usuario_id

    # Redirecionar com base no tipo de usuário
    if tipo_usuario_id == "Professor":
        return redirect(url_for("portal_professores"))
    elif tipo_usuario_id == "Aluno":
        return redirect(url_for("alunos_bp.portal_alunos"))
    elif tipo_usuario_id == "Administrador":
        return redirect(url_for("portal_administrador"))
    elif tipo_usuario_id == "Administração da Escola":
        return redirect(url_for("portal_administracao"))
    elif tipo_usuario_id == "Secretaria de Educação":
        return redirect(url_for("portal_secretaria_educacao"))
    else:
        return redirect(url_for("login"))



from flask_login import login_user, current_user
from flask import flash, session

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        login_identifier = request.form['email']  # Pode ser email ou CPF
        senha = request.form['senha']

        db = get_db()
        cursor = db.cursor()
        # Modifica a consulta para buscar por email OU CPF
        cursor.execute("""
            SELECT id, nome, tipo_usuario_id, escola_id, serie_id, turma_id, email, senha, codigo_ibge
            FROM usuarios
            WHERE email = ? OR cpf = ?
        """, (login_identifier, login_identifier))
        user_data = cursor.fetchone()

        if user_data:
            user_id, nome, tipo_usuario_id, escola_id, serie_id, turma_id, email_db, senha_hash, codigo_ibge = user_data

            if check_password_hash(senha_hash, senha):
                user = User(
                    id=user_id,
                    nome=nome,
                    tipo_usuario_id=tipo_usuario_id,
                    escola_id=escola_id,
                    serie_id=serie_id,
                    turma_id=turma_id,
                    email=email_db,
                    codigo_ibge=codigo_ibge
                )
                login_user(user)

                next_page = request.args.get('next') or {
                    1: 'portal_administrador',
                    2: 'portal_administracao',
                    3: 'portal_professores',
                    4: 'alunos_bp.portal_alunos',  # Alterado para usar o endpoint do blueprint
                    5: 'portal_secretaria_educacao'
                }.get(tipo_usuario_id, 'login')

                return redirect(url_for(next_page))
            else:
                error = "Senha incorreta. Por favor, tente novamente."
        else:
            error = "Usuário não encontrado. Por favor, verifique seu email ou CPF."

    return render_template('login.html', error=error)

# Rota de logout

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))

import base64

@app.route('/redefinir-senha', defaults={'token': None}, methods=['GET', 'POST'])
@app.route('/redefinir-senha/<token>', methods=['GET', 'POST'])
def redefinir_senha(token):
    if not token:
        flash("Token inválido ou ausente.", "danger")
        return redirect(url_for('recuperar_senha'))

    try:
        # Decodificar o token
        token_data = base64.urlsafe_b64decode(token.encode()).decode()
        user_id, email = token_data.split(":")
        user_id = int(user_id)

        if request.method == 'POST':
            nova_senha = request.form.get('nova_senha')
            confirmar_senha = request.form.get('confirmar_senha')

            if nova_senha != confirmar_senha:
                flash("As senhas não coincidem.", "danger")
                return redirect(url_for('redefinir_senha', token=token))

            db = get_db()
            cursor = db.cursor()

            # Verificar se a nova senha é igual à atual
            cursor.execute("SELECT senha FROM usuarios WHERE id = ?", (user_id,))
            current_password_hash = cursor.fetchone()

            if not current_password_hash:
                flash("Usuário não encontrado ou token inválido.", "danger")
                return redirect(url_for('recuperar_senha'))

            current_password_hash = current_password_hash[0]

            if check_password_hash(current_password_hash, nova_senha):
                flash("Essa já é a sua senha.", "warning")
                return redirect(url_for('redefinir_senha', token=token))

            # Atualizar a senha no banco
            nova_senha_hash = generate_password_hash(nova_senha)
            cursor.execute("UPDATE usuarios SET senha = ? WHERE id = ?", (nova_senha_hash, user_id))
            db.commit()

            # Mensagem de sucesso
            flash("Senha redefinida com sucesso!", "success")
            return redirect(url_for('redefinir_senha', token=token))

    except Exception as e:
        flash(f"Erro ao processar o token: {e}", "danger")
        return redirect(url_for('recuperar_senha'))

    return render_template('redefinir_senha.html', token=token)


import base64
import json

@app.route('/recuperar-senha', methods=['GET', 'POST'])
def recuperar_senha():
    if request.method == 'POST':
        email = request.form['email']
        
        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT id, nome FROM usuarios WHERE email = ?", (email,))
        user = cursor.fetchone()

        if user:
            user_id, nome = user

            # Gerar um token simples baseado no ID e email
            token_data = f"{user_id}:{email}"
            token = base64.urlsafe_b64encode(token_data.encode()).decode()

            reset_link = url_for('redefinir_senha', token=token, _external=True)
            print(f"Link de redefinição de senha para {email}: {reset_link}")

            flash("Link de redefinição de senha gerado e visível no console.", "success")
        else:
            flash("Email não encontrado.", "danger")

    return render_template('recuperar_senha.html')



# Portais (rotas protegidas)
@app.route('/portal_administrador')
@login_required
def portal_administrador():
    # Verifica se o tipo_usuario_id é 1 (Administrador)
    if current_user.tipo_usuario_id != 1:
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for('login'))

    db = get_db()
    cursor = db.cursor()

    # Buscar informações gerais, caso necessário
    # Exemplo de lógica para futuras implementações
    # cursor.execute("SELECT COUNT(*) FROM escolas")
    # total_escolas = cursor.fetchone()[0]


    return render_template(
        'portal_administrador.html',
        title="Portal do Administrador"
    )

def extrair_texto_pdf(pdf_path):
    """
    Extrai texto de um arquivo PDF
    """
    import PyPDF2
    texto = ""
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                texto += page.extract_text() + "\n"
        return texto
    except Exception as e:
        print(f"Erro ao extrair texto do PDF: {str(e)}")
        return None

@app.route("/upload", methods=["POST"])
def upload_file():
    try:
        if "file" not in request.files:
            return jsonify({"error": "Nenhum arquivo enviado"}), 400

        file = request.files["file"]
        usar_ia = request.form.get("usar_ia") == "true"

        if file.filename == "":
            return jsonify({"error": "Nenhum arquivo selecionado"}), 400

        if not file.filename.endswith(".pdf"):
            return jsonify({"error": "Apenas arquivos PDF são permitidos"}), 400

        # Criar pasta temporária se não existir
        if not os.path.exists(app.config["UPLOAD_FOLDER"]):
            os.makedirs(app.config["UPLOAD_FOLDER"])

        # Salvar arquivo temporariamente
        filename = secure_filename(file.filename)
        temp_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(temp_path)

        # Extrair texto do PDF
        text = extrair_texto_pdf(temp_path)
        
        # Remover arquivo temporário
        os.remove(temp_path)

        if not text:
            return jsonify({"error": "Não foi possível extrair texto do PDF"}), 400

        # Processar o texto para extrair questões
        if usar_ia:
            questoes = processar_com_ia(text)
        else:
            questoes = processar_sem_ia(text)

        if not questoes:
            return jsonify({"error": "Nenhuma questão encontrada no arquivo"}), 400

        # Verificar se todas as questões têm disciplina_id
        for questao in questoes:
            if not questao.get("disciplina_id"):
                return jsonify({"error": "ID da disciplina não encontrado em uma ou mais questões"}), 400

        # Salvar questões no banco
        salvar_questoes_no_banco(questoes)

        return jsonify({"message": "Arquivo processado com sucesso", "questoes": len(questoes)}), 200

    except Exception as e:
        print(f"Erro ao processar arquivo: {str(e)}")
        return jsonify({"error": str(e)}), 500
        
import json
import pdfplumber
from flask import flash, redirect, url_for, request
from flask_login import login_required, current_user

@app.route("/upload_questoes_ia", methods=["POST"])
@login_required
def upload_questoes_ia():
    if current_user.tipo_usuario_id != 5:  # Apenas para Secretaria de Educação
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))

    file = request.files.get("file")
    if not file:
        flash("Nenhum arquivo enviado!", "danger")
        return redirect(url_for("portal_secretaria_educacao"))

    if not file.filename.endswith(".pdf"):
        flash("Formato de arquivo inválido! Apenas arquivos PDF são permitidos.", "danger")
        return redirect(url_for("portal_secretaria_educacao"))

    try:
        # Extração de texto do PDF
        text = extract_text_from_pdf(file)
        if not text:
            flash("Não foi possível extrair texto do PDF.", "danger")
            return redirect(url_for("portal_secretaria_educacao"))

        # Dividir o texto em chunks
        chunks = split_text_into_chunks(text)
        print(f"[DEBUG] Texto dividido em {len(chunks)} chunks")

        # Processar cada chunk e consolidar resultados
        all_questions = []
        for i, chunk in enumerate(chunks):
            print(f"[DEBUG] Processando chunk {i+1}/{len(chunks)}")
            questions = processar_com_ia(chunk)
            if questions:
                all_questions.extend(questions)

        if not all_questions:
            flash("Não foi possível gerar questões do PDF. Verifique o conteúdo do arquivo.", "danger")
            return redirect(url_for("portal_secretaria_educacao"))

        # Salvar questões no banco de dados
        salvar_questoes_no_banco(all_questions)
        flash(f"Processamento concluído! {len(all_questions)} questões foram geradas.", "success")

    except Exception as e:
        print(f"[DEBUG] Erro ao processar PDF: {e}")
        flash(f"Erro ao processar o arquivo: {e}", "danger")

    return redirect(url_for("portal_secretaria_educacao"))

from openai import OpenAI

client = openai.Client(api_key="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")

def processar_com_ia(text):
    """
    Processa um chunk de texto usando GPT-3.5 para extrair e estruturar questões existentes.
    """
    try:
        prompt = f"""
        IMPORTANTE: NÃO CRIE novas questões. Apenas identifique e estruture as questões que já existem no texto fornecido.
        
        Analise o texto e encontre as questões de múltipla escolha que já existem nele.
        Para cada questão encontrada, retorne um array JSON com a seguinte estrutura:
        [
            {{
                "disciplina": "nome da disciplina",
                "assunto": "assunto específico",
                "questao": "texto da questão",
                "alternativa_a": "texto da alternativa a",
                "alternativa_b": "texto da alternativa b",
                "alternativa_c": "texto da alternativa c",
                "alternativa_d": "texto da alternativa d",
                "alternativa_e": "texto da alternativa e",
                "questao_correta": "letra da alternativa correta (A-E)",
                "serie_id": "número da série (1-9 fundamental, 1-3 médio)",
                "mes_id": "número do mês (1-12)",
                "simulado_id": "número sequencial do simulado no mês (1, 2, 3, etc)"
            }}
        ]

        IMPORTANTE:
        1. Retorne todas as questões contidas no pdf, nem mais nem menos.
        2. Retorne APENAS o array JSON puro, sem marcadores de código.
        3. Mantenha as questões e alternativas concisas.
        4. Garanta que todas as questões sigam exatamente o formato JSON especificado.
    
        Texto para análise:
        {text}
        """
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Você é um especialista em educação que SEMPRE retorna respostas em formato JSON puro, sem marcadores de código ou formatação adicional. Você SEMPRE gera exatamente 5 questões, nem mais nem menos."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.5,  # Reduzindo a temperatura para maior consistência
            max_tokens=2000
        )

        # Extrair e validar as questões do JSON retornado
        content = response.choices[0].message.content.strip()
        
        # Remover qualquer marcador de código ou formatação adicional
        content = content.replace('```json', '').replace('```', '').strip()
        
        # Garantir que o conteúdo comece com [ e termine com ]
        if not content.startswith('[') or not content.endswith(']'):
            print(f"[DEBUG] Resposta mal formatada: {content}")
            return None

        try:
            # Tenta carregar o JSON diretamente
            questoes = json.loads(content)
            if not isinstance(questoes, list):
                print(f"[DEBUG] Resposta não é uma lista: {content}")
                return None
        except json.JSONDecodeError:
            print(f"[DEBUG] Erro ao decodificar JSON: {content}")
            return None

        # Validar estrutura das questões
        validated_questions = []
        required_fields = [
            'disciplina', 'assunto', 'questao', 
            'alternativa_a', 'alternativa_b', 'alternativa_c', 
            'alternativa_d', 'alternativa_e', 'questao_correta',
            'serie_id', 'mes_id', 'simulado_id'
        ]
        
        for q in questoes:
            if all(key in q for key in required_fields):
                # Validar série_id e mes_id
                try:
                    q['serie_id'] = int(str(q['serie_id']))
                    q['mes_id'] = int(str(q['mes_id']))
                    q['simulado_id'] = int(str(q['simulado_id']))
                    if not (1 <= q['mes_id'] <= 12):
                        print(f"[DEBUG] mes_id inválido: {q['mes_id']}")
                        continue
                    if not ((1 <= q['serie_id'] <= 9) or (1 <= q['serie_id'] <= 3)):
                        print(f"[DEBUG] serie_id inválido: {q['serie_id']}")
                        continue
                    if q['simulado_id'] < 1:
                        print(f"[DEBUG] simulado_id inválido: {q['simulado_id']}")
                        continue
                    validated_questions.append(q)
                except (ValueError, TypeError):
                    print(f"[DEBUG] Erro ao converter serie_id ou mes_id para números: {q}")
                    continue
            else:
                missing = [f for f in required_fields if f not in q]
                print(f"[DEBUG] Questão ignorada por falta de campos obrigatórios: {missing}")

        # Se nenhuma questão foi validada, tenta novamente com um novo prompt
        if not validated_questions:
            print("[DEBUG] Nenhuma questão válida gerada, tentando novamente...")
            return processar_com_ia(text)  # Recursão para tentar novamente

        return validated_questions

    except Exception as e:
        print(f"[DEBUG] Erro ao processar texto com IA: {str(e)}")
        return None

def processar_sem_ia(text):
    """
    Processa um texto para extrair questões de múltipla escolha usando padrões simples.
    Não depende de IA, tornando o processo mais econômico.
    """
    if not text or not text.strip():
        print("Texto vazio recebido")
        return None
        
    try:
        import re
        questoes = []
        
        # Log do texto recebido
        print(f"Texto recebido (primeiros 200 caracteres): {text[:200]}")
        
        # Dividir o texto em possíveis questões
        partes = text.split("Questão ")
        print(f"Número de partes encontradas: {len(partes)}")
        
        for i, parte in enumerate(partes[1:], 1):  # Ignorar primeira parte que pode ser cabeçalho
            print(f"\nProcessando questão {i}:")
            # Remove números no início da parte
            parte = re.sub(r'^\d+\s*', '', parte.strip())
            linhas = parte.split('\n')
            print(f"Número de linhas na questão: {len(linhas)}")
            
            texto_questao = []
            alternativas = {'A': '', 'B': '', 'C': '', 'D': '', 'E': ''}
            questao_correta = ""
            disciplina_id = None
            nome_disciplina = None
            assunto = ""
            serie_id = 1
            mes_id = 1
            
            # Processar linha por linha
            alternativa_atual = None
            primeira_linha = True
            for linha in linhas:
                linha = linha.strip()
                if not linha:
                    continue
                
                # Remove números no início da primeira linha do texto da questão
                if primeira_linha:
                    linha = re.sub(r'^\d+\s*', '', linha)
                    primeira_linha = False
                
                # Verificar metadados primeiro
                if linha.startswith("Disciplina:"):
                    nome_disciplina = linha.split(":")[-1].strip()
                    print(f"Disciplina encontrada: {nome_disciplina}")
                    disciplina_id = get_disciplina_id(nome_disciplina)
                    if disciplina_id is None:
                        print(f"Erro: Não foi possível obter ID para disciplina: {nome_disciplina}")
                        continue
                elif linha.startswith("Assunto:"):
                    assunto = linha.split(":")[-1].strip()
                elif linha.startswith("Resposta Correta:"):
                    questao_correta = linha.split(":")[-1].strip()
                elif linha.startswith("Serie_id:"):
                    serie_id = int(linha.split(":")[-1].strip())
                elif linha.startswith("Mes_id:"):
                    mes_id = int(linha.split(":")[-1].strip())
                # Verificar se é uma alternativa apenas se não for metadado
                elif match := re.match(r'^\(?([A-E])\)?\s+(.+)$', linha):
                    alternativa_atual = match.group(1)
                    alternativas[alternativa_atual] = match.group(2)
                    print(f"Alternativa {alternativa_atual} encontrada: {alternativas[alternativa_atual][:50]}...")
                elif alternativa_atual:
                    # Continua texto da alternativa
                    alternativas[alternativa_atual] += ' ' + linha
                else:
                    # É parte do enunciado
                    texto_questao.append(linha)
                    print(f"Linha de enunciado encontrada: {linha[:50]}...")
            
            # Se encontrou alternativas e tem disciplina_id, é uma questão válida
            if any(alternativas.values()) and disciplina_id is not None:
                print(f"Questão {i} válida encontrada com {len([a for a in alternativas.values() if a])} alternativas")
                questao = {
                    "disciplina_id": disciplina_id,
                    "assunto": assunto or "Não especificado",
                    "questao": ' '.join(texto_questao),
                    "alternativa_a": alternativas['A'],
                    "alternativa_b": alternativas['B'],
                    "alternativa_c": alternativas['C'],
                    "alternativa_d": alternativas['D'],
                    "alternativa_e": alternativas['E'],
                    "questao_correta": questao_correta,
                    "serie_id": serie_id,
                    "mes_id": mes_id,
                    "simulado_id": 1
                }
                questoes.append(questao)
                print(f"Questão {i} adicionada com disciplina_id: {disciplina_id}")
            else:
                print(f"Questão {i} ignorada - sem alternativas ou sem disciplina_id")
        
        print(f"\nTotal de questões processadas: {len(questoes)}")
        return questoes if questoes else None

    except Exception as e:
        print(f"Erro ao processar texto: {str(e)}")
        import traceback
        print(f"Traceback completo: {traceback.format_exc()}")
        return None

def get_disciplina_id(nome_disciplina):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id FROM disciplinas WHERE nome = ?", (nome_disciplina,))
    result = cursor.fetchone()
    if result:
        return result[0]
    else:
        # Se a disciplina não existe, criar
        cursor.execute("INSERT INTO disciplinas (nome) VALUES (?)", (nome_disciplina,))
        db.commit()
        return cursor.lastrowid

def salvar_questoes_no_banco(questoes):
    db = get_db()
    cursor = db.cursor()
    try:
        for questao in questoes:
            cursor.execute("""
                INSERT INTO banco_questoes (
                    questao, alternativa_a, alternativa_b, alternativa_c, alternativa_d, alternativa_e,
                    questao_correta, disciplina_id, assunto, serie_id, mes_id
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                questao["questao"],
                questao["alternativa_a"],
                questao["alternativa_b"],
                questao["alternativa_c"],
                questao["alternativa_d"],
                questao.get("alternativa_e"),
                questao["questao_correta"],
                questao["disciplina_id"],
                questao["assunto"],
                questao["serie_id"],
                questao["mes_id"]
            ))
        db.commit()
    except Exception as e:
        db.rollback()
        raise e

import json
import openai


def analise_ia(text):
    # Exemplo de implementação de análise de texto
    questoes = []
    # Separar o texto em questões
    questoes_texto = text.split("QUESTÃO")
    for q_text in questoes_texto[1:]:
        questao = {}
        linhas = q_text.split("\n")
        questao['questao'] = linhas[1].strip()
        questao['alternativa_a'] = linhas[3].strip()
        questao['alternativa_b'] = linhas[4].strip()
        questao['alternativa_c'] = linhas[5].strip()
        questao['alternativa_d'] = linhas[6].strip()
        questao['alternativa_e'] = linhas[7].strip() if linhas[7].startswith("(E)") else None
        questao['questao_correta'] = linhas[-6].split(":")[-1].strip()
        questao['disciplina'] = linhas[-5].split(":")[-1].strip()
        questao['assunto'] = linhas[-4].split(":")[-1].strip()
        questao['serie_id'] = int(linhas[-3].split(":")[-1].strip())
        questao['mes_id'] = int(linhas[-2].split(":")[-1].strip())
        questoes.append(questao)
    return

@app.route("/visualizar_banco_questoes")
@login_required
def visualizar_banco_questoes():
    if not current_user.tipo_usuario_id == 5:
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))
    
    db = get_db()
    cursor = db.cursor()
    
    # Recuperar questões do banco
    cursor.execute("""
        SELECT id, questao, alternativa_a, alternativa_b, alternativa_c, alternativa_d, alternativa_e, questao_correta, disciplina, assunto, serie_id
        FROM banco_questoes
    """)
    questoes = cursor.fetchall()
    
    return render_template("visualizar_banco_questoes.html", questoes=questoes)

def split_text_into_chunks(text, max_chunk_size=2000):
    """Divide o texto em chunks menores para processamento."""
    chunks = []
    current_chunk = ""
    for p in text.split('\n\n'):
        if len(current_chunk) + len(p) <= max_chunk_size:
            current_chunk += p + "\n\n"
        else:
            chunks.append(current_chunk.strip())
            current_chunk = p + "\n\n"
    if current_chunk:
        chunks.append(current_chunk.strip())
    return chunks

def generate_questions_with_ai(text):
    """
    Função para gerar questões usando IA e identificar disciplina e assunto.
    """
    openai.api_key = "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"

    # Prompt para a IA
    prompt = f"""
    Leia o seguinte texto e crie 5 questões de múltipla escolha com 4 alternativas cada.
    Para cada questão, identifique a disciplina e o assunto abordado.
    Retorne no seguinte formato:
    
    [
      {{
        "questao": "Enunciado da questão",
        "alternativa_a": "Texto da alternativa A",
        "alternativa_b": "Texto da alternativa B",
        "alternativa_c": "Texto da alternativa C",
        "alternativa_d": "Texto da alternativa D",
        "alternativa_e": "Texto da alternativa E (se aplicável)",
        "questao_correta": "Letra da alternativa correta",
        "disciplina": "Disciplina identificada",
        "assunto": "Assunto identificado"
        "serie_id": "Série identificada"
      }},
      ...
    ]
    
    Texto:
    {text[:1000]}  # Limita a entrada para evitar excesso de dados
    """

    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",  # Troquei o modelo
            messages=[
                {"role": "system", "content": "Você é um gerador de perguntas de múltipla escolha."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=2000,
            temperature=0.7,
        )
        
        print(f"Resposta da IA: {response['choices'][0]['message']['content']}")
        return response['choices'][0]['message']['content']

    except openai.error.OpenAIError as e:
        print(f"Erro ao acessar a API OpenAI: {e}")
        return None



def extract_text_from_pdf(file):
    """
    Extrai texto de um arquivo PDF usando pdfplumber.
    """
    try:
        with pdfplumber.open(file) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() + "\n"
            return text
    except Exception as e:
        print(f"[DEBUG] Erro ao extrair texto do PDF: {e}")
        return None


@app.route("/portal_secretaria_educacao", methods=["GET", "POST"])
@login_required
def portal_secretaria_educacao():
    if current_user.tipo_usuario_id != 5:  # Verifica se é uma Secretaria de Educação
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))

    db = get_db()
    cursor = db.cursor()

    # Buscar todas as séries disponíveis
    print("[DEBUG] Buscando séries disponíveis...")
    cursor.execute("SELECT id, nome FROM series")
    series = cursor.fetchall()
    print(f"[DEBUG] Séries encontradas: {series}")

    # Buscar os meses disponíveis
    meses = [
        (1, "Janeiro"), (2, "Fevereiro"), (3, "Março"),
        (4, "Abril"), (5, "Maio"), (6, "Junho"),
        (7, "Julho"), (8, "Agosto"), (9, "Setembro"),
        (10, "Outubro"), (11, "Novembro"), (12, "Dezembro")
    ]
    print("[DEBUG] Meses carregados.")

    # Buscar simulados já gerados
    print("[DEBUG] Buscando simulados gerados...")
    cursor.execute("""
        SELECT sg.id, s.nome AS serie_nome, sg.mes_id, d.nome AS disciplina_nome, 
           sg.data_envio, sg.status
        FROM simulados_gerados sg
        JOIN series s ON sg.serie_id = s.id
        JOIN disciplinas d ON sg.disciplina_id = d.id
        ORDER BY sg.data_envio DESC
    """)
    simulados_gerados = cursor.fetchall()
    print(f"[DEBUG] Simulados já gerados: {simulados_gerados}")

    # Buscar o codigo_ibge do usuário atual (Secretaria de Educação)
    cursor.execute("SELECT codigo_ibge FROM usuarios WHERE id = ?", (current_user.id,))
    codigo_ibge = cursor.fetchone()[0]

    # Buscar escolas que têm o mesmo codigo_ibge
    cursor.execute("""
        SELECT id, nome_da_escola
        FROM escolas
        WHERE codigo_ibge = ?
    """, (codigo_ibge,))
    escolas = cursor.fetchall()

    # Buscar o número total de escolas
    total_escolas = len(escolas)

    # Buscar o número de alunos na mesma codigo_ibge
    cursor.execute("SELECT COUNT(*) FROM usuarios WHERE tipo_usuario_id = 4 AND codigo_ibge = ?", (codigo_ibge,))
    numero_alunos = cursor.fetchone()[0]

    # Buscar o número de simulados gerados na mesma codigo_ibge
    cursor.execute("""
        SELECT COUNT(*)
        FROM simulados_gerados sg
        JOIN usuarios u ON sg.serie_id = u.serie_id
        WHERE u.codigo_ibge = ?
    """, (codigo_ibge,))
    numero_simulados_gerados = cursor.fetchone()[0]

    # Calcular a média geral de desempenho dos simulados respondidos na mesma codigo_ibge
    cursor.execute("""
        SELECT AVG(d.desempenho)
        FROM desempenho_simulado d
        JOIN usuarios u ON d.aluno_id = u.id
        WHERE u.codigo_ibge = ?
    """, (codigo_ibge,))
    media_geral = cursor.fetchone()[0] or 0

    # Buscar disciplinas disponíveis
    print("[DEBUG] Buscando disciplinas disponíveis...")
    cursor.execute("""
        SELECT DISTINCT d.id, d.nome
        FROM disciplinas d
        JOIN banco_questoes bq ON bq.disciplina_id = d.id
    """)
    disciplinas = cursor.fetchall()
    print(f"[DEBUG] Disciplinas encontradas: {disciplinas}")

    if request.method == "POST":
        print("[DEBUG] Processando requisição POST...")
        serie_id = request.form.get("serie_id")
        mes_id = request.form.get("mes_id")
        disciplina_id = request.form.get("disciplina_id")
        print(f"[DEBUG] Dados recebidos do formulário - Série: {serie_id}, Mês: {mes_id}, Disciplina ID: {disciplina_id}")

        if not serie_id or not mes_id or not disciplina_id:
            flash("Ano escolar, Mês e Componente Curricular são obrigatórios para gerar o simulado.", "danger")
            print("[DEBUG] Ano escolar, Mês ou Componente curricular não fornecidos.")
            return render_template(
                "portal_secretaria_educacao.html",
                series=series,
                meses=meses,
                simulados_gerados=simulados_gerados,
                total_escolas=total_escolas,
                total_alunos=numero_alunos,
                total_simulados=numero_simulados_gerados,
                media_geral=media_geral,
                disciplinas=disciplinas
            )

        try:
            # Verificar se o simulado já foi gerado para a série, mês e disciplina selecionados
            cursor.execute("""
                SELECT 1 FROM simulados_gerados 
                WHERE serie_id = ? AND mes_id = ? AND disciplina_id = ?
            """, (serie_id, mes_id, disciplina_id))
            if cursor.fetchone():
                flash("Simulado já gerado para esta série, mês e disciplina.", "danger")
                print(f"[DEBUG] Simulado já existente para série {serie_id}, mês {mes_id} e disciplina {disciplina_id}.")
            else:
                # Buscar questões do banco para a série, mês e disciplina selecionados
                print("[DEBUG] Buscando questões para o simulado...")
                cursor.execute("""
                    SELECT id FROM banco_questoes 
                    WHERE serie_id = ? AND mes_id = ? AND disciplina_id = ?
                """, (serie_id, mes_id, disciplina_id))
                questoes = cursor.fetchall()
                print(f"[DEBUG] Questões encontradas: {questoes}")

                if not questoes:
                    flash("Não há questões disponíveis para esta série, mês e disciplina.", "warning")
                    print("[DEBUG] Nenhuma questão encontrada para a série, mês e disciplina fornecidos.")
                else:
                    # Inserir o simulado no banco
                    print("[DEBUG] Inserindo simulado no banco...")
                    cursor.execute("""
                        INSERT INTO simulados_gerados (serie_id, mes_id, disciplina_id, status, data_envio) 
                        VALUES (?, ?, ?, 'gerado', datetime('now'))
                    """, (serie_id, mes_id, disciplina_id))
                    simulado_id = cursor.lastrowid
                    print(f"[DEBUG] Simulado gerado com ID: {simulado_id}")

                    # Associar as questões ao simulado gerado
                    print("[DEBUG] Associando questões ao simulado...")
                    for questao in questoes:
                        cursor.execute("INSERT INTO simulado_questoes (simulado_id, questao_id) VALUES (?, ?)", (simulado_id, questao[0]))
                    db.commit()
                    flash("Simulado gerado com sucesso!", "success")
                    print("[DEBUG] Simulado gerado com sucesso e questões associadas.")

        except Exception as e:
            db.rollback()
            print(f"[DEBUG] Erro ao gerar simulado: {e}")
            flash(f"Erro ao gerar simulado: {e}", "danger")

    return render_template(
        "portal_secretaria_educacao.html",
        series=series,
        meses=meses,
        simulados_gerados=simulados_gerados,
        total_escolas=total_escolas,
        total_alunos=numero_alunos,
        total_simulados=numero_simulados_gerados,
        media_geral=media_geral,
        disciplinas=disciplinas
    )


@app.route("/visualizar_simulado/<int:simulado_id>", methods=["GET", "POST"])
@login_required
def visualizar_simulado(simulado_id):
    if current_user.tipo_usuario_id != 5:  # Apenas para Secretaria de Educação
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))

    db = get_db()
    cursor = db.cursor()

    # Buscar detalhes do simulado e suas questões
    print(f"[DEBUG] Buscando questões para o simulado ID {simulado_id}")
    cursor.execute("""
        SELECT 
            bq.questao, 
            bq.alternativa_a, 
            bq.alternativa_b, 
            bq.alternativa_c, 
            bq.alternativa_d, 
            bq.alternativa_e,
            bq.questao_correta
        FROM simulado_questoes sq
        JOIN banco_questoes bq ON sq.questao_id = bq.id
        WHERE sq.simulado_id = ?
    """, (simulado_id,))
    questoes = cursor.fetchall()
    print(f"[DEBUG] Questões encontradas: {questoes}")

    if not questoes:
        flash("Nenhuma questão encontrada para este simulado.", "danger")
        return redirect(url_for("portal_secretaria_educacao"))

    # Formatar questões para o template
    questoes_formatadas = [
        {
            "pergunta": questao[0],
            "alternativas": {
                "A": questao[1],
                "B": questao[2],
                "C": questao[3],
                "D": questao[4],
                "E": questao[5],
            },
            "correta": questao[6]
        }
        for questao in questoes
    ]
    print(f"[DEBUG] Questões formatadas: {questoes_formatadas}")

    if request.method == "POST":
        # Atualizar status do simulado para "enviado" e associar alunos
        cursor.execute("""
            UPDATE simulados_gerados
            SET status = 'enviado' 
            WHERE id = ?
        """, (simulado_id,))

        # Associar simulado aos alunos da série correspondente
        cursor.execute("""
            SELECT serie_id FROM simulados_gerados WHERE id = ?
        """, (simulado_id,))
        serie_id = cursor.fetchone()[0]

        cursor.execute("""
            SELECT id FROM usuarios WHERE tipo_usuario_id = 4 AND serie_id = ?
        """, (serie_id,))
        alunos = cursor.fetchall()

        for aluno in alunos:
            cursor.execute("""
                INSERT INTO aluno_simulado (aluno_id, simulado_id, status)
                VALUES (?, ?, 'pendente')
            """, (aluno[0], simulado_id))

        db.commit()
        flash("Simulado enviado com sucesso para os alunos!", "success")
        return redirect(url_for("portal_secretaria_educacao"))

    return render_template("simulado_gerado.html", questoes=questoes_formatadas, simulado_id=simulado_id)

@app.route("/cancelar_simulado/<int:simulado_id>")
@login_required
def cancelar_simulado(simulado_id):
    if current_user.tipo_usuario_id != 5:  # Apenas Secretaria de Educação
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))
    
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Atualiza o status do simulado para cancelado
        cursor.execute("""
            UPDATE simulados_gerados 
            SET status = 'pendente' 
            WHERE id = ?
        """, (simulado_id,))
        
        # Remove as atribuições do simulado aos alunos
        cursor.execute("""
            DELETE FROM aluno_simulado 
            WHERE simulado_id = ?
        """, (simulado_id,))
        
        db.commit()
        flash("Simulado cancelado com sucesso!", "success")
        
    except Exception as e:
        db.rollback()
        flash(f"Erro ao cancelar simulado: {str(e)}", "danger")
    
    return redirect(url_for("portal_secretaria_educacao"))

@app.route('/visualizar_simulado/<int:simulado_id>/editar-campo', methods=['POST'])
@login_required
def editar_campo_simulado(simulado_id):
    if current_user.tipo_usuario_id != 5:  # Verifica se é professor
        return jsonify({'success': False, 'message': 'Acesso não autorizado'}), 403
    
    try:
        tipo = request.form.get('tipo')
        index = int(request.form.get('index'))
        valor = request.form.get('valor')
        
        db = get_db()
        cursor = db.cursor()
        
        # Primeiro, buscar o ID da questão no banco_questoes usando a tabela correta
        cursor.execute('''
            SELECT bq.id, bq.questao, bq.alternativa_a, bq.alternativa_b, 
                   bq.alternativa_c, bq.alternativa_d, bq.alternativa_e, bq.questao_correta
            FROM simulado_questoes sq
            JOIN banco_questoes bq ON sq.questao_id = bq.id
            WHERE sq.simulado_id = ?
            LIMIT 1 OFFSET ?
        ''', (simulado_id, index))
        
        questao = cursor.fetchone()
        if not questao:
            return jsonify({'success': False, 'message': 'Questão não encontrada'}), 404
            
        pergunta_id = questao[0]
        
        # Atualizar o campo específico no banco_questoes
        if tipo == 'pergunta':
            cursor.execute('UPDATE banco_questoes SET questao = ? WHERE id = ?',
                         (valor, pergunta_id))
        elif tipo.startswith('alternativa-'):
            letra = tipo.split('-')[1].lower()
            coluna = f'alternativa_{letra}'
            cursor.execute(f'UPDATE banco_questoes SET {coluna} = ? WHERE id = ?',
                         (valor, pergunta_id))
        elif tipo == 'correta':
            cursor.execute('UPDATE banco_questoes SET questao_correta = ? WHERE id = ?',
                         (valor, pergunta_id))
        
        db.commit()
        return jsonify({'success': True, 'message': 'Campo atualizado com sucesso'})
        
    except Exception as e:
        db.rollback()
        print(f"Erro ao salvar: {str(e)}")  # Para debug
        return jsonify({'success': False, 'message': str(e)}), 500



@app.route('/salvar_edicao_simulado/<int:simulado_id>', methods=['POST'])
@login_required
def salvar_edicao_simulado(simulado_id):
    if current_user.tipo_usuario_id != 5:
        return jsonify({'success': False, 'message': 'Acesso não autorizado'})

    try:
        db = get_db()
        cursor = db.cursor()

        # Log inicial
        print("\n=== DADOS RECEBIDOS DO FORMULÁRIO ===")
        for key, value in request.form.items():
            print(f"{key}: {value}")
        print("=====================================\n")

        # Coletar os índices das questões enviadas
        perguntas = []
        for key in request.form:
            if key.startswith('perguntas['):
                idx = key.split('[')[1].split(']')[0]
                if idx not in perguntas:
                    perguntas.append(idx)

        for idx in perguntas:
            # Recuperar dados do formulário
            pergunta = request.form.get(f'perguntas[{idx}][pergunta]', '').strip()
            alternativa_a = request.form.get(f'perguntas[{idx}][alternativas][A]', '').strip()
            alternativa_b = request.form.get(f'perguntas[{idx}][alternativas][B]', '').strip()
            alternativa_c = request.form.get(f'perguntas[{idx}][alternativas][C]', '').strip()
            alternativa_d = request.form.get(f'perguntas[{idx}][alternativas][D]', '').strip()
            alternativa_e = request.form.get(f'perguntas[{idx}][alternativas][E]', '').strip()
            resposta_correta = request.form.get(f'perguntas[{idx}][correta]', '').strip()

            # Debug: Mostrar os dados processados
            print(f"\n--- Questão {idx} ---")
            print(f"Pergunta: {pergunta}")
            print(f"A: {alternativa_a}")
            print(f"B: {alternativa_b}")
            print(f"C: {alternativa_c}")
            print(f"D: {alternativa_d}")
            print(f"E: {alternativa_e}")
            print(f"Resposta Correta: {resposta_correta}")
            print("----------------------")

            # Validar campos obrigatórios
            campos_obrigatorios = {
                'pergunta': pergunta,
                'alternativa A': alternativa_a,
                'alternativa B': alternativa_b,
                'alternativa C': alternativa_c,
                'alternativa D': alternativa_d,
                'resposta correta': resposta_correta
            }

            campos_vazios = [nome for nome, valor in campos_obrigatorios.items() if not valor]

            if campos_vazios:
                return jsonify({
                    'success': False,
                    'message': f'Os seguintes campos estão vazios: {", ".join(campos_vazios)}'
                })

            # Validar resposta correta
            alternativas_validas = ['A', 'B', 'C', 'D']
            if alternativa_e:
                alternativas_validas.append('E')

            if resposta_correta not in alternativas_validas:
                return jsonify({
                    'success': False,
                    'message': f'A resposta correta deve ser uma das alternativas disponíveis: {", ".join(alternativas_validas)}'
                })

            if resposta_correta == 'E' and not alternativa_e:
                return jsonify({
                    'success': False,
                    'message': 'A alternativa E foi marcada como correta mas não foi preenchida'
                })

            # Buscar ID da questão no banco de dados
            cursor.execute("""
                SELECT bq.id
                FROM simulado_questoes sq
                JOIN banco_questoes bq ON sq.questao_id = bq.id
                WHERE sq.simulado_id = ?
                LIMIT 1 OFFSET ?
            """, (simulado_id, int(idx)))

            result = cursor.fetchone()
            if not result:
                print(f"Erro: Questão {int(idx) + 1} não encontrada no banco de dados.")
                return jsonify({'success': False, 'message': f'Questão {int(idx) + 1} não encontrada'})

            questao_id = result[0]
            print(f"ID da questão no banco: {questao_id}")

            # Atualizar a questão no banco de dados
            cursor.execute("""
                UPDATE banco_questoes 
                SET questao = ?, alternativa_a = ?, alternativa_b = ?, alternativa_c = ?, alternativa_d = ?, alternativa_e = ?, questao_correta = ?
                WHERE id = ?
            """, (pergunta, alternativa_a, alternativa_b, alternativa_c, alternativa_d, alternativa_e, resposta_correta, questao_id))

        # Salvar alterações no banco
        db.commit()
        print("Atualização concluída com sucesso!")
        return jsonify({'success': True, 'message': 'Simulado atualizado com sucesso!'})

    except Exception as e:
        db.rollback()
        print(f"Erro ao atualizar simulado: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro ao atualizar simulado: {str(e)}'})


@app.route('/portal_professores', methods=['GET'])
@login_required
def portal_professores():
    db = get_db()
    cursor = db.cursor()

    # Recuperar as turmas vinculadas ao professor na tabela professor_turma_escola
    cursor.execute("""
        SELECT DISTINCT t.id, se.nome AS serie, t.turma, e.nome_da_escola AS escola
        FROM professor_turma_escola pte
        JOIN turmas t ON pte.turma_id = t.id
        JOIN series se ON t.serie_id = se.id
        JOIN escolas e ON t.escola_id = e.id
        WHERE pte.professor_id = ?
    """, (current_user.id,))
    turmas = cursor.fetchall()

    print("Turmas vinculadas:", turmas)

    # Recuperar filtros
    assunto_filtro = request.args.get('assunto', '').strip()
    avaliacao_filtro = request.args.get('avaliacao', '').strip()
    page = int(request.args.get('page', 1))
    per_page = 10
    offset = (page - 1) * per_page

    # Base da consulta para simulados gerados
    query = """
        SELECT 
            s.id, 
            a.nome AS assunto,
            d.nome AS disciplina, 
            se.nome AS serie,
            t.id AS turma_id,
            t.turma AS letra_turma,
            s.data_envio,
            s.status
        FROM simulados_gerados s
        JOIN series se ON s.serie_id = se.id
        JOIN disciplinas d ON s.disciplina_id = d.id
        JOIN professor_turma_escola pte ON pte.professor_id = ?
        JOIN turmas t ON t.serie_id = s.serie_id AND t.id = pte.turma_id
        LEFT JOIN assuntos a ON a.disciplina_id = s.disciplina_id AND a.professor_id = pte.professor_id
    """
    filters = []
    params = [current_user.id]

    # Aplicar filtros, se disponíveis
    if assunto_filtro:
        filters.append("a.nome LIKE ?")
        params.append(f"%{assunto_filtro}%")
    if avaliacao_filtro:
        filters.append("s.status = ?")
        params.append(avaliacao_filtro)

    if filters:
        query += " AND " + " AND ".join(filters)

    # Adicionar paginação
    query += " ORDER BY s.id DESC LIMIT ? OFFSET ?"
    params.extend([per_page, offset])

    print("Consulta SQL final:", query)
    print("Parâmetros da consulta:", params)

    # Executar a consulta para simulados
    cursor.execute(query, params)
    simulados = cursor.fetchall()
    print("Simulados retornados antes do tratamento:", simulados)

    # Tratar valores retornados e ajustar a saída
    simulados_tratados = [
        (
            simulado[0],  # s.id
            simulado[1],  # a.nome (assunto)
            simulado[2],  # d.nome (disciplina)
            simulado[3],  # se.nome (serie)
            simulado[4],  # t.id (turma_id)
            simulado[5],  # t.turma (letra da turma)
            simulado[6],  # s.data_envio
            simulado[7]   # s.status
        )
        for simulado in simulados
    ]
    print("Simulados tratados:", simulados_tratados)

    # Contar total de simulados para paginação
    count_query = """
        SELECT COUNT(*)
        FROM simulados_gerados s
        JOIN series se ON s.serie_id = se.id
        JOIN disciplinas d ON s.disciplina_id = d.id
        JOIN professor_turma_escola pte ON pte.professor_id = ?
        JOIN turmas t ON t.serie_id = s.serie_id AND t.id = pte.turma_id
        LEFT JOIN assuntos a ON a.disciplina_id = s.disciplina_id AND a.professor_id = pte.professor_id
    """
    if filters:
        count_query += " AND " + " AND ".join(filters)

    print("Consulta SQL de contagem:", count_query)
    print("Parâmetros da contagem:", params[:-2])

    cursor.execute(count_query, params[:-2])
    total_simulados = cursor.fetchone()[0]
    total_pages = (total_simulados + per_page - 1) // per_page

    return render_template(
        "portal_professores.html",
        title="Portal dos Professores",
        turmas=turmas,
        simulados=simulados_tratados,
        total_pages=total_pages,
        current_page=page,
        assunto_filtro=assunto_filtro,
        avaliacao_filtro=avaliacao_filtro
    )

from flask import render_template, redirect, url_for, flash
from flask_login import login_required, current_user

@app.route("/portal_administracao")
@login_required
def portal_administracao():
    if not current_user.is_authenticated:
        print("Usuário não autenticado.")
        return redirect(url_for("login"))

    print(f"Usuário atual: {current_user.tipo_usuario_id}")
    if current_user.tipo_usuario_id != 2:
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))

    return render_template("portal_administracao.html")

@app.route("/get_alunos", methods=["GET"])
@login_required
def get_alunos():
    turma_id = request.args.get("turma_id")
    if not turma_id:
        return jsonify([])  # Retorna lista vazia se turma_id não for fornecida.

    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT id, nome 
        FROM usuarios 
        WHERE tipo_usuario_id = 4 AND turma_id = ?
    """, (turma_id,))
    alunos = cursor.fetchall()

    return jsonify([{"id": aluno[0], "nome": aluno[1]} for aluno in alunos])


@app.route("/get_alunos_por_turma", methods=["GET"])
@login_required
def get_alunos_por_turma():
    turma_id = request.args.get("turma_id")
    if not turma_id:
        return jsonify([])

    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT id, nome
        FROM usuarios
        WHERE turma_id = ? AND tipo_usuario_id = 4
    """, (turma_id,))
    alunos = cursor.fetchall()

    return jsonify([{"id": aluno[0], "nome": aluno[1]} for aluno in alunos])


# @app.route("/portal_alunos")
# @login_required
# def portal_alunos():
#     if current_user.tipo_usuario_id != 4:  # Apenas alunos podem acessar
#         flash("Acesso não autorizado.", "danger")
#         return redirect(url_for("index"))

#     db = get_db()
#     cursor = db.cursor()

#     # Buscar simulados pendentes do aluno, ordenando por status (Pendente primeiro) e data
#     cursor.execute("""
#         SELECT sg.id, d.nome AS disciplina_nome, sg.mes_id, 
#                strftime('%d-%m-%Y', date(sg.data_envio)) as data_envio,
#                CASE 
#                    WHEN EXISTS (
#                        SELECT 1 
#                        FROM desempenho_simulado ds 
#                        WHERE ds.simulado_id = sg.id 
#                        AND ds.aluno_id = am.aluno_id
#                    ) THEN 'Respondido'
#                    ELSE 'Pendente'
#                END as status
#         FROM aluno_simulado am
#         JOIN simulados_gerados sg ON am.simulado_id = sg.id
#         JOIN disciplinas d ON sg.disciplina_id = d.id
#         WHERE am.aluno_id = ?
#         ORDER BY 
#             CASE 
#                 WHEN EXISTS (
#                     SELECT 1 
#                     FROM desempenho_simulado ds 
#                     WHERE ds.simulado_id = sg.id 
#                     AND ds.aluno_id = am.aluno_id
#                 ) THEN 1 
#                 ELSE 0 
#             END,
#             sg.data_envio DESC
#     """, (current_user.id,))
#     simulados = cursor.fetchall()

#     # Lista de meses para exibição
#     meses = [
#         (1, "Janeiro"), (2, "Fevereiro"), (3, "Março"),
#         (4, "Abril"), (5, "Maio"), (6, "Junho"),
#         (7, "Julho"), (8, "Agosto"), (9, "Setembro"),
#         (10, "Outubro"), (11, "Novembro"), (12, "Dezembro")
#     ]

#     return render_template("alunos/portal_alunos.html", simulados=simulados, meses=meses)





from flask import flash

@app.route("/simulado_diario", methods=["GET", "POST"])
@login_required
def responder_simulado_diario():
    db = get_db()
    cursor = db.cursor()

    # Verificar se há simulado diário para hoje
    cursor.execute("SELECT perguntas FROM simulado_diario WHERE data = date('now')")
    simulado_diario = cursor.fetchone()

    if not simulado_diario:
        # Gerar 10 perguntas usando a IA
        perguntas_geradas = gerar_perguntas_conhecimentos_gerais(quantidade=10)
        print(f"[DEBUG] Perguntas Geradas: {perguntas_geradas}")
        if perguntas_geradas and len(perguntas_geradas) == 10:
            # Serializar perguntas e salvar no banco
            perguntas_serializadas = json.dumps(perguntas_geradas)
            cursor.execute("INSERT INTO simulado_diario (data, perguntas) VALUES (date('now'), ?)", (perguntas_serializadas,))
            db.commit()
            simulado_diario = (perguntas_serializadas,)
        else:
            flash("Erro ao gerar o simulado diário ou número insuficiente de perguntas.", "danger")
            return redirect(url_for("alunos_bp.portal_alunos"))

    perguntas = json.loads(simulado_diario[0])

    # Verificar se o aluno já respondeu ao simulado diário hoje
    cursor.execute("""
        SELECT COUNT(*) 
        FROM respostas_simulado_diario 
        WHERE aluno_id = ? AND date(data_envio) = date('now')
    """, (current_user.id,))
    ja_respondeu = cursor.fetchone()[0] > 0

    if ja_respondeu:
        flash("Você já respondeu o simulado diário hoje. Tente novamente amanhã!", "info")
        return redirect(url_for("alunos_bp.portal_alunos"))

    if request.method == "POST":
        # Coletar as respostas do aluno e convertê-las para letras
        respostas_dadas = [
            chr(64 + int(request.form.get(f"resposta_{idx + 1}")))  # Converte índices para letras (1 -> A, 2 -> B, ...)
            for idx in range(len(perguntas))
        ]

        # Obter as respostas corretas
        respostas_certas = [pergunta["correta"] for pergunta in perguntas]

        # Calcular acertos
        acertos = sum(1 for aluno_resp, correta_resp in zip(respostas_dadas, respostas_certas) if aluno_resp == correta_resp)
        total_questoes = len(perguntas)

        try:
            # Salvar o simulado completo em uma única linha no formato JSON
            cursor.execute("""
                INSERT INTO respostas_simulado_diario (aluno_id, respostas, correta, acertos, total_questoes, data_envio)
                VALUES (?, ?, ?, ?, ?, datetime('now'))
            """, (
                current_user.id,
                json.dumps(respostas_dadas),  # Armazena as respostas como JSON
                json.dumps(respostas_certas),  # Armazena as respostas corretas como JSON
                acertos,
                total_questoes
            ))

            db.commit()
            flash(f"Você respondeu o simulado diário! Acertos: {acertos}/{total_questoes}", "success")
        except Exception as e:
            db.rollback()
            flash(f"Erro ao salvar suas respostas: {e}", "danger")

        return redirect(url_for("alunos_bp.portal_alunos"))

    perguntas_formatadas = [
        {
            "id": str(idx + 1),
            "pergunta": pergunta["pergunta"],
            "opcoes": pergunta["opcoes"]
        }
        for idx, pergunta in enumerate(perguntas)
    ]

    return render_template("simulado_diario.html", perguntas=perguntas_formatadas)


@app.route('/ranking', methods=['GET'])
@login_required
def ranking():
    db = get_db()
    cursor = db.cursor()

    # Recuperar os filtros da requisição
    escola_id = request.args.get('escola_id')
    turma_id = request.args.get('turma_id')

    # Base da consulta
    query = """
        SELECT 
            u.nome, 
            r.pontuacao, 
            r.data_participacao, 
            e.nome AS escola, 
            s.nome || ' ' || t.turma AS turma
        FROM ranking r
        JOIN usuarios u ON r.aluno_id = u.id
        JOIN escolas e ON u.escola_id = e.id
        JOIN turmas t ON u.turma_id = t.id
        JOIN series s ON t.serie_id = s.id
    """
    filters = []
    params = []

    # Adicionar filtros com base nos parâmetros fornecidos
    if escola_id:
        filters.append("e.id = ?")
        params.append(escola_id)
    if turma_id:
        filters.append("t.id = ?")
        params.append(turma_id)

    # Adicionar cláusulas WHERE se necessário
    if filters:
        query += " WHERE " + " AND ".join(filters)

    # Ordenar o ranking
    query += " ORDER BY r.pontuacao DESC, r.data_participacao ASC"

    try:
        # Executar a consulta principal
        cursor.execute(query, params)
        ranking_data = cursor.fetchall()

        # Recuperar dados para os filtros
        cursor.execute("SELECT id, nome FROM escolas")
        escolas = cursor.fetchall()

        cursor.execute("""
            SELECT t.id, s.nome || ' ' || t.turma 
            FROM turmas t
            JOIN series s ON t.serie_id = s.id
        """)
        turmas = cursor.fetchall()

        return render_template(
            'ranking.html',
            ranking_data=ranking_data,
            escolas=escolas,
            turmas=turmas,
            selected_escola=escola_id,
            selected_turma=turma_id
        )
    except Exception as e:
        flash(f"Erro ao carregar o ranking: {e}", "danger")
        return redirect(url_for("home"))

@app.route('/atualizar_ranking', methods=['GET'])
@login_required
def atualizar_ranking():
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute("""
            INSERT INTO ranking (aluno_id, pontuacao, data_participacao)
            SELECT 
                aluno_id, 
                SUM(acertos) AS pontuacao, 
                MAX(DATE(data_envio)) AS data_participacao
            FROM respostas_simulado_diario
            GROUP BY aluno_id
            ON CONFLICT(aluno_id) DO UPDATE SET
                pontuacao = excluded.pontuacao,
                data_participacao = excluded.data_participacao;
        """)
        db.commit()
        flash("Ranking atualizado com sucesso!", "success")
    except Exception as e:
        db.rollback()
        flash(f"Erro ao atualizar o ranking: {e}", "danger")

    return redirect(url_for("ranking"))



from flask import session

@app.route('/simulados_disponiveis')
@login_required
def simulados_disponiveis():
    if current_user.tipo_usuario_id != 1:  # Apenas alunos
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT * FROM simulados WHERE status = 'aprovado'
    """)
    simulados = cursor.fetchall()
    return render_template("simulados_disponiveis.html", simulados=simulados)

@app.route('/revisar_simulado/<int:simulado_id>', methods=['GET', 'POST'])
@login_required
def revisar_simulado(simulado_id):
    if current_user.tipo_usuario_id != 3:  # Apenas professores
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Buscar simulado e questões associadas
    cursor.execute("SELECT * FROM simulados WHERE id = ? AND professor_id = ?", (simulado_id, current_user.id))
    simulado = cursor.fetchone()

    if not simulado:
        flash("Simulado não encontrado ou você não tem permissão para acessá-lo.", "error")
        return redirect(url_for("portal_professores"))

    cursor.execute("SELECT * FROM questoes WHERE simulado_id = ?", (simulado_id,))
    questoes = cursor.fetchall()

    if request.method == 'POST':
        acao = request.form.get('acao')  # 'aprovar' ou 'rejeitar'
        if acao == 'aprovar':
            cursor.execute("UPDATE simulados SET status = 'aprovado' WHERE id = ?", (simulado_id,))
            flash("Simulado aprovado com sucesso!", "success")
        elif acao == 'rejeitar':
            cursor.execute("UPDATE simulados SET status = 'rejeitado' WHERE id = ?", (simulado_id,))
            flash("Simulado rejeitado.", "info")
        else:
            flash("Ação inválida.", "error")
        db.commit()
        return redirect(url_for("portal_professores"))

    return render_template("revisar_simulado.html", simulado=simulado, questoes=questoes)

@app.route("/enviar_simulado", methods=["POST"])
@login_required
def enviar_simulado():
    if current_user.tipo_usuario_id != 5:  # Garantir que apenas a Secretaria de Educação acesse
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("portal_secretaria_educacao"))

    serie_id = request.form.get("serie_id")
    if not serie_id:
        flash("Por favor, selecione uma série.", "warning")
        return redirect(url_for("portal_secretaria_educacao"))

    # Lógica para gerar e enviar simulados para os alunos da série selecionada
    try:
        db = get_db()
        cursor = db.cursor()

        # Verificar alunos na série e enviar simulado
        cursor.execute("""
            SELECT id FROM usuarios
            WHERE tipo_usuario_id = 4 AND serie_id = ?
        """, (serie_id,))
        alunos = cursor.fetchall()

        if not alunos:
            flash("Nenhum aluno encontrado para a série selecionada.", "info")
            return redirect(url_for("portal_secretaria_educacao"))

        for aluno in alunos:
            # Lógica para vincular o simulado aos alunos
            cursor.execute("""
                INSERT INTO simulados_gerados (aluno_id, serie_id, status)
                VALUES (?, ?, ?)
            """, (aluno[0], serie_id, 'enviado'))

        db.commit()
        flash("Simulado enviado com sucesso!", "success")
    except Exception as e:
        db.rollback()
        flash(f"Erro ao enviar simulado: {e}", "danger")

    return redirect(url_for("portal_secretaria_educacao"))

@app.route('/enviar_respostas', methods=['POST'])
@login_required
def enviar_respostas():
    if current_user.tipo_usuario_id != 4:
        return redirect(url_for("home"))

    simulado_id = request.form.get('simulado_id')

    if not simulado_id:
        return "Simulado inválido.", 400

    db = get_db()
    cursor = db.cursor()

    try:
        # Obter informações do simulado
        cursor.execute("""
            SELECT respostas_certas, disciplina_id, assunto, turma_id
            FROM simulados
            WHERE id = ?
        """, (simulado_id,))
        simulado = cursor.fetchone()

        if not simulado:
            return "Simulado inválido.", 400

        respostas_certas, disciplina_id, assunto_id, turma_id = simulado
        respostas_certas = json.loads(respostas_certas)

        # Processar respostas do aluno
        respostas_dadas = []
        for key, value in request.form.items():
            if key.startswith("resposta_"):
                respostas_dadas.append(int(value))

        # Calcular acertos e salvar respostas do aluno
        acertos = 0
        total_questoes = len(respostas_certas)

        for i, (resposta, correta) in enumerate(zip(respostas_dadas, respostas_certas)):
            correta_flag = resposta == correta
            if correta_flag:
                acertos += 1

            # Salvar cada resposta na tabela respostas_aluno
            cursor.execute("""
                INSERT INTO respostas_aluno (aluno_id, pergunta_id, resposta, correta)
                VALUES (?, ?, ?, ?)
            """, (current_user.id, i + 1, resposta, correta_flag))

        # Calcular a porcentagem de acertos
        porcentagem = round((acertos / total_questoes) * 100, 2)

        # Inserir resumo do simulado na tabela resultados_simulados
        cursor.execute("""
            INSERT INTO resultados_simulados (simulado_id, aluno_id, respostas, acertos, total_questoes, data_envio, respostas_certas)
            VALUES (?, ?, ?, ?, ?, datetime('now'), ?)
        """, (simulado_id, current_user.id, json.dumps(respostas_dadas), acertos, total_questoes, json.dumps(respostas_certas)))

        # Inserir ou atualizar na tabela desempenho
        # Inserir ou atualizar na tabela desempenho
        cursor.execute("""
            INSERT INTO desempenho (
                aluno_id, disciplina_id, assunto_id, acertos, total_questoes, porcentagem, turma_id, simulado_id, data_envio
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            ON CONFLICT(aluno_id, disciplina_id, assunto_id) DO UPDATE SET
                acertos = acertos + excluded.acertos,
                total_questoes = total_questoes + excluded.total_questoes,
                porcentagem = (acertos * 100.0 / total_questoes)
        """, (current_user.id, disciplina_id, assunto_id, acertos, total_questoes, porcentagem, turma_id, simulado_id))

        db.commit()

        return f"Você acertou {acertos} de {total_questoes} perguntas! ({porcentagem}%)"
    except Exception as e:
        db.rollback()
        return f"Erro ao processar respostas: {e}", 500

@app.route('/atualizar_desempenho', methods=['POST'])
@login_required
def atualizar_desempenho():
    if current_user.tipo_usuario_id != 1:
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Obter os dados da tabela resultados_simulados
    cursor.execute("""
        SELECT r.simulado_id, r.aluno_id, r.acertos, r.total_questoes, r.data_envio, s.disciplina_id, s.assunto
        FROM resultados_simulados r
        JOIN simulados s ON r.simulado_id = s.id
    """)
    resultados = cursor.fetchall()

    for resultado in resultados:
        simulado_id, aluno_id, acertos, total_questoes, data_envio, disciplina_id, assunto_id = resultado

        # Calcular a porcentagem
        porcentagem = (acertos / total_questoes) * 100

        # Verificar se já existe registro na tabela desempenho para evitar duplicação
        cursor.execute("""
            SELECT id FROM desempenho
            WHERE aluno_id = ? AND disciplina_id = ? AND assunto_id = ?
        """, (aluno_id, disciplina_id, assunto_id))
        registro_existente = cursor.fetchone()

        if registro_existente:
            # Atualizar registro existente
            cursor.execute("""
                UPDATE desempenho
                SET acertos = ?, total_questoes = ?, porcentagem = ?, data_envio = ?
                WHERE id = ?
            """, (acertos, total_questoes, porcentagem, data_envio, registro_existente[0]))
        else:
            # Inserir novo registro
            cursor.execute("""
                INSERT INTO desempenho (aluno_id, disciplina_id, assunto_id, acertos, total_questoes, porcentagem, turma_id, simulado_id, data_envio)
                ON CONFLICT(aluno_id, disciplina_id, assunto_id) DO UPDATE SET
                    acertos = acertos + excluded.acertos,
                    total_questoes = total_questoes + excluded.total_questoes,
                    porcentagem = (acertos * 100.0 / total_questoes)
            """, (current_user.id, disciplina_id, assunto_id, acertos, total_questoes, porcentagem, turma_id, simulado_id))
    db.commit()

    # Calcular a média da turma
    cursor.execute("""
        SELECT AVG(porcentagem) AS media_turma
        FROM desempenho
    """)
    media_turma = cursor.fetchone()[0]

    return f"A média da turma é: {media_turma:.2f}%"

@app.route('/turma/<int:turma_id>/alunos')
@login_required
def ver_alunos_turma(turma_id):
    db = get_db()
    cursor = db.cursor()

    # Buscar a turma pelo ID
    cursor.execute("""
        SELECT se.nome AS serie, t.turma, e.nome AS escola, t.id AS turma_id
        FROM turmas t
        JOIN series se ON t.serie_id = se.id  -- Junção para obter o nome da série
        JOIN escolas e ON t.escola_id = e.id
        WHERE t.id = ?
    """, (turma_id,))
    turma = cursor.fetchone()

    if not turma:
        return "Turma não encontrada", 404

    # Buscar alunos vinculados à turma
    cursor.execute("""
        SELECT id, nome
        FROM usuarios
        WHERE turma_id = ? AND tipo_usuario_id = ?
    """, (turma_id, 4))  # 4 é o ID correspondente ao tipo 'Aluno'
    alunos = cursor.fetchall()

    # Buscar desempenho dos alunos na turma
    cursor.execute("""
        SELECT 
            d.aluno_id, 
            u.nome AS aluno_nome, 
            d.disciplina_id, 
            d.assunto_id, 
            d.acertos, 
            d.total_questoes, 
            d.porcentagem
        FROM desempenho d
        JOIN usuarios u ON d.aluno_id = u.id
        WHERE u.turma_id = ?
    """, (turma_id,))
    desempenho = cursor.fetchall()

    return render_template(
        'alunos_turma.html',
        turma=turma,
        alunos=alunos,
        desempenho=desempenho
    )

def enviar_simulado_para_alunos(simulado_id, serie_id):
    db = get_db()
    cursor = db.cursor()

    # Selecionar alunos da série
    cursor.execute(
        """
        SELECT id FROM usuarios
        WHERE tipo_usuario_id = 4 AND serie_id = ?
        """,
        (serie_id,),
    )
    alunos = cursor.fetchall()

    # Associar o simulado aos alunos (se necessário, criar uma tabela aluno_simulado)
    for aluno in alunos:
        cursor.execute(
            """
            INSERT INTO aluno_simulado (aluno_id, simulado_id)
            VALUES (?, ?)
            """,
            (aluno["id"], simulado_id),
        )
    db.commit()


# @app.route('/turma/<int:turma_id>/simulado/<int:simulado_id>/enviar', methods=['POST'])
# @login_required
# def enviar_simulado(turma_id, simulado_id):
#     db = get_db()
#     cursor = db.cursor()

#     try:
#         # Garantir que o simulado pertence à turma e ao professor logado
#         cursor.execute("""
#             SELECT id FROM simulados_gerados
#             WHERE id = ? AND turma_id = ? AND professor_id = ? AND status = 'pendente'
#         """, (simulado_id, turma_id, current_user.id))
#         simulado = cursor.fetchone()

#         if not simulado:
#             flash("Simulado inválido ou já enviado.", "error")
#             return redirect(url_for("visualizar_simulado", turma_id=turma_id, simulado_id=simulado_id))

#         # Atualizar o status para 'enviado'
#         cursor.execute("""
#             UPDATE simulados_gerados
#             SET status = 'enviado'
#             WHERE id = ?
#         """, (simulado_id,))
#         db.commit()

#         flash("Simulado enviado com sucesso!", "success")
#         return redirect(url_for("portal_professores"))

#     except Exception as e:
#         db.rollback()
#         print(f"[DEBUG] Erro ao enviar simulado: {e}")
#         flash("Erro ao enviar simulado. Tente novamente.", "error")
#         return redirect(url_for("visualizar_simulado", turma_id=turma_id, simulado_id=simulado_id))

# @app.route('/turma/<int:turma_id>/simulado/<int:simulado_id>/visualizar', methods=['GET'])
# @login_required
# def visualizar_simulado(turma_id, simulado_id):
#     if current_user.tipo_usuario_id != 3:  # Apenas professores
#         return redirect(url_for("home"))

#     db = get_db()
#     cursor = db.cursor()

#     # Recuperar informações do simulado
#     cursor.execute("""
#         SELECT s.assunto_id, s.nivel, s.quantidade_questoes, s.quantidade_alternativas, s.respostas_certas, a.nome AS assunto_nome
#         FROM simulados_gerados s
#         JOIN assuntos a ON s.assunto_id = a.id
#         WHERE s.id = ? AND s.turma_id = ?
#     """, (simulado_id, turma_id))
#     simulado_info = cursor.fetchone()

#     if not simulado_info:
#         flash("Simulado não encontrado.", "error")
#         return redirect(url_for("portal_professores"))

#     assunto_id, nivel, quantidade_questoes, quantidade_alternativas, respostas_certas, assunto_nome = simulado_info

#     # Recuperar perguntas do simulado
#     # As perguntas devem ser armazenadas no banco ou reconstruídas aqui, conforme o seu sistema
#     perguntas = json.loads(respostas_certas)  # Exemplo, ajuste conforme necessário

#     return render_template(
#         "visualizar_simulado.html",
#         simulado=perguntas,
#         turma_id=turma_id,
#         simulado_id=simulado_id,
#         assunto_nome=assunto_nome,
#         nivel=nivel,
#         quantidade_questoes=quantidade_questoes,
#         quantidade_alternativas=quantidade_alternativas
#     )


# def destacar_palavra(pergunta, palavras_chave):
#     """
#     Função para destacar palavras-chave em uma pergunta.
#     :param pergunta: Texto da pergunta.
#     :param palavras_chave: Lista de palavras-chave para destacar.
#     :return: Pergunta com palavras-chave destacadas.
#     """
#     for palavra in palavras_chave:
#         pergunta = pergunta.replace(
#             palavra, f"<strong>{palavra}</strong>"
#         )
#     return pergunta

    

# @app.route("/gerar_simulado", methods=["POST"])
# @login_required
# def gerar_simulado():
#     if not current_user.tipo_usuario_id == 5:  # Apenas Secretaria de Educação
#         flash("Acesso não autorizado.", "danger")
#         return redirect(url_for("portal_secretaria_educacao"))

#     titulo = request.form.get("titulo")
#     escola_id = request.form.get("escola_id")
#     tipo_ensino_id = request.form.get("tipo_ensino_id")
#     serie_id = request.form.get("serie_id")
#     quantidade = int(request.form.get("quantidade"))

#     db = get_db()
#     cursor = db.cursor()

#     # Verificar se a escola pertence ao código IBGE do usuário
#     cursor.execute(
#         """
#         SELECT id
#         FROM escolas
#         WHERE id = ? AND codigo_ibge = ?
#         """,
#         (escola_id, current_user.codigo_ibge),
#     )
#     escola = cursor.fetchone()

#     if not escola:
#         flash("Acesso não autorizado à escola selecionada.", "danger")
#         return redirect(url_for("portal_secretaria_educacao"))

#     # Selecionar questões da série e escola
#     cursor.execute(
#         """
#         SELECT id FROM banco_questoes
#         WHERE serie_id = ? AND EXISTS (
#             SELECT 1 FROM escolas WHERE id = ? AND codigo_ibge = ?
#         )
#         ORDER BY RANDOM()
#         LIMIT ?
#         """,
#         (serie_id, escola_id, current_user.codigo_ibge, quantidade),
#     )
#     questoes = cursor.fetchall()

#     if not questoes:
#         flash("Nenhuma questão encontrada para a série selecionada.", "danger")
#         return redirect(url_for("portal_secretaria_educacao"))

#     try:
#         # Inserir o simulado na tabela simulados
#         cursor.execute(
#             """
#             INSERT INTO simulados (titulo, serie_id)
#             VALUES (?, ?)
#             """,
#             (titulo, serie_id),
#         )
#         simulado_id = cursor.lastrowid

#         # Inserir as questões no simulado
#         for questao in questoes:
#             cursor.execute(
#                 """
#                 INSERT INTO simulado_questoes (simulado_id, questao_id)
#                 VALUES (?, ?)
#                 """,
#                 (simulado_id, questao["id"]),
#             )

#         db.commit()

#         # Enviar o simulado aos alunos da série
#         enviar_simulado_para_alunos(simulado_id, serie_id)

#         flash("Simulado gerado e enviado com sucesso!", "success")
#     except Exception as e:
#         db.rollback()
#         flash(f"Erro ao gerar o simulado: {e}", "danger")

#     return redirect(url_for("portal_secretaria_educacao"))


# @app.route('/turma/<int:turma_id>/gerar_simulado', methods=['GET', 'POST'])
# @login_required
# def gerar_simulado_turma(turma_id):
#     if current_user.tipo_usuario_id != 3:  # Apenas professores
#         return redirect(url_for("home"))

#     db = get_db()
#     cursor = db.cursor()

#     # Buscar informações da turma
#     cursor.execute("""
#         SELECT t.serie_id, t.id AS turma_id, e.id AS escola_id
#         FROM turmas t
#         JOIN escolas e ON t.escola_id = e.id
#         WHERE t.id = ?
#     """, (turma_id,))
#     turma_info = cursor.fetchone()

#     if not turma_info:
#         flash("Turma inválida ou não encontrada.", "error")
#         return redirect(url_for("portal_professores"))

#     serie_id, turma_id, escola_id = turma_info
#     disciplina_id = request.form.get("disciplina_id") or request.args.get("disciplina_id")
#     assuntos = []
#     assunto_nome = None

#     # Buscar assuntos existentes
#     if disciplina_id:
#         try:
#             disciplina_id = int(disciplina_id)
#             cursor.execute("""
#                 SELECT id, nome
#                 FROM assuntos
#                 WHERE disciplina_id = ? AND serie_id = ?
#                 ORDER BY nome ASC
#             """, (disciplina_id, serie_id))
#             assuntos = cursor.fetchall()
#         except Exception as e:
#             print(f"[DEBUG] Erro ao buscar assuntos: {e}")

#     # Processar o formulário POST
#     if request.method == 'POST' and disciplina_id:
#         try:
#             assunto_id = request.form.get("assunto_id")
#             novo_assunto = request.form.get("novo_assunto", "").strip()
#             nivel = request.form.get("nivel", "").strip()
#             quantidade_questoes = int(request.form.get("quantidade_questoes", "").strip())
#             quantidade_alternativas = int(request.form.get("quantidade_alternativas", "").strip())

#             if not nivel or not quantidade_questoes or not quantidade_alternativas:
#                 raise ValueError("Campos obrigatórios estão ausentes!")

#             # Inserir ou selecionar assunto
#             if novo_assunto:
#                 cursor.execute("""
#                     SELECT id FROM assuntos
#                     WHERE nome = ? AND disciplina_id = ? AND serie_id = ? AND professor_id = ?
#                 """, (novo_assunto, disciplina_id, serie_id, current_user.id))
#                 resultado = cursor.fetchone()

#                 if resultado:
#                     assunto_id = resultado[0]
#                 else:
#                     cursor.execute("""
#                         INSERT INTO assuntos (nome, disciplina_id, serie_id, professor_id)
#                         VALUES (?, ?, ?, ?)
#                     """, (novo_assunto, disciplina_id, serie_id, current_user.id))
#                     assunto_id = cursor.lastrowid

#             elif not assunto_id:
#                 raise ValueError("Nenhum assunto selecionado ou criado.")

#             cursor.execute("SELECT nome FROM assuntos WHERE id = ?", (assunto_id,))
#             assunto_nome = cursor.fetchone()[0]

#             # Salvar simulado no banco de dados
#             cursor.execute("""
#                 INSERT INTO simulados_gerados (assunto_id, turma_id, serie_id, escola_id, disciplina_id, professor_id, nivel, quantidade_questoes, quantidade_alternativas, respostas_certas, status)
#                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
#             """, (assunto_id, turma_id, serie_id, escola_id, disciplina_id, current_user.id, nivel, quantidade_questoes, quantidade_alternativas, "[]", "pendente"))
#             simulado_id = cursor.lastrowid

#             # Gerar perguntas diretamente com a IA
#             simulado = gerar_perguntas(
#                 disciplina=disciplina_id,
#                 assunto=assunto_nome,
#                 quantidade=quantidade_questoes,
#                 nivel=nivel,
#                 alternativas=quantidade_alternativas
#             )

#             if not simulado:
#                 raise ValueError("Erro ao gerar perguntas. Nenhuma pergunta retornada.")

#             # Adicionar letras às alternativas manualmente no Python
#             for pergunta in simulado:
#                 pergunta["opcoes"] = [f"{letter}) {opcao}" for letter, opcao in zip(["A", "B", "C", "D"], pergunta["opcoes"])]

#             # Processar respostas corretas e salvar perguntas na tabela perguntas_simulado
#             respostas_certas = []
#             for pergunta in simulado:
#                 resposta_index = ord(pergunta["resposta_correta"].strip().upper()) - ord('A')
#                 if 0 <= resposta_index < len(pergunta["opcoes"]):
#                     resposta_letra = ["A", "B", "C", "D"][resposta_index]  # Mapear índice para letra
#                     resposta_texto = pergunta["opcoes"][resposta_index]  # Obter o texto da resposta correta
#                     respostas_certas.append(resposta_letra)

#                      # Adicionar o texto da resposta correta no dicionário de perguntas
#                     pergunta["resposta_correta_texto"] = resposta_texto
                
#                     # Inserir pergunta na tabela perguntas_simulado
#                     cursor.execute("""
#                         INSERT INTO perguntas_simulado (simulado_id, texto_pergunta, alternativa_a, alternativa_b, alternativa_c, alternativa_d, correta)
#                         VALUES (?, ?, ?, ?, ?, ?, ?)
#                     """, (
#                     simulado_id,
#                     pergunta["pergunta"],
#                     pergunta["opcoes"][0],
#                     pergunta["opcoes"][1],
#                     pergunta["opcoes"][2],
#                     pergunta["opcoes"][3],
#                     resposta_letra
#                 ))

#             # Atualizar respostas_certas no banco
#             respostas_certas_serializadas = json.dumps(respostas_certas)
#             cursor.execute("""
#                 UPDATE simulados_gerados
#                 SET respostas_certas = ?
#                 WHERE id = ?
#             """, (respostas_certas_serializadas, simulado_id))

#             db.commit()

#             flash("Simulado gerado com sucesso!", "success")

#             return render_template(
#                 "visualizar_simulado.html",
#                 simulado=simulado,
#                 turma_id=turma_id,
#                 disciplina_id=disciplina_id,
#                 assunto_id=assunto_id,
#                 assunto_nome=assunto_nome,
#                 nivel=nivel,
#                 quantidade_questoes=quantidade_questoes,
#                 quantidade_alternativas=quantidade_alternativas,
#                 simulado_id=simulado_id
#             )
#         except Exception as e:
#             db.rollback()
#             print(f"[DEBUG] Erro ao processar formulário: {e}")
#             flash(f"Erro ao gerar simulado: {str(e)}", "error")

#     cursor.execute("SELECT id, nome FROM disciplinas ORDER BY nome ASC")
#     disciplinas = cursor.fetchall()
#     return render_template(
#         "gerar_simulado_turma.html",
#         turma_id=turma_id,
#         disciplinas=disciplinas,
#         assuntos=assuntos,
#         disciplina_id=disciplina_id,
#         assunto_nome=assunto_nome
#     )

# @app.route('/turma/<int:turma_id>/simulado/<int:simulado_id>/avaliar', methods=['POST'])
# @login_required
# def avaliar_simulado(turma_id, simulado_id):
#     if current_user.tipo_usuario_id != 3:  # Apenas professores podem avaliar
#         return redirect(url_for("home"))

#     avaliacao = request.form.get("avaliacao")
#     if not avaliacao or not avaliacao.isdigit() or int(avaliacao) not in range(1, 6):
#         flash("Selecione uma avaliação válida entre 1 e 5 estrelas.", "error")
#         return redirect(url_for("visualizar_simulado", turma_id=turma_id, simulado_id=simulado_id))

#     db = get_db()
#     cursor = db.cursor()

#     cursor.execute("""
#         UPDATE simulados_gerados
#         SET avaliacao = ?
#         WHERE id = ? AND turma_id = ?
#     """, (int(avaliacao), simulado_id, turma_id))
#     db.commit()

#     flash("Avaliação salva com sucesso!", "success")
#     return redirect(url_for("visualizar_simulado", turma_id=turma_id, simulado_id=simulado_id))

@app.route('/aluno/simulados', methods=['GET'])
@login_required
def listar_simulados_aluno():
    if current_user.tipo_usuario_id != 4:  # Apenas alunos podem acessar
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Buscar simulados disponíveis para o aluno
    cursor.execute("""
        SELECT sg.id, a.nome AS assunto, sg.nivel, sg.quantidade_questoes
        FROM simulados_gerados sg
        JOIN assuntos a ON sg.assunto_id = a.id
        JOIN turmas t ON sg.turma_id = t.id
        JOIN usuarios u ON u.turma_id = t.id
        WHERE u.id = ?
    """, (current_user.id,))
    simulados = cursor.fetchall()

    return render_template("portal_alunos.html", simulados=simulados)



@app.template_filter('chr')
def chr_filter(value):
    try:
        return chr(value)
    except (TypeError, ValueError):
        return value

app.jinja_env.filters['chr'] = chr_filter

@app.route("/gerar_simulado/<int:serie_id>/<int:mes_id>/<string:disciplina>", methods=["POST"])
@login_required
def gerar_simulado(serie_id, mes_id, disciplina):
    if current_user.tipo_usuario_id != 5:  # Verifica se é um usuário da secretaria
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Primeiro, buscar o ID da disciplina pelo nome
    cursor.execute("SELECT id FROM disciplinas WHERE nome = ?", (disciplina,))
    disciplina_result = cursor.fetchone()
    
    if not disciplina_result:
        flash(f"Disciplina '{disciplina}' não encontrada.", "danger")
        return redirect(url_for("portal_secretaria_educacao"))
        
    disciplina_id = disciplina_result[0]

    # Verificar se já existe um simulado para o mês, série e disciplina
    cursor.execute("""
        SELECT id FROM simulados_gerados 
        WHERE serie_id = ? AND mes_id = ? AND disciplina_id = ?
    """, (serie_id, mes_id, disciplina_id))
    simulado_existente = cursor.fetchone()

    if simulado_existente:
        flash("Já existe um simulado gerado para esta série, mês e disciplina.", "warning")
        return redirect(url_for("portal_secretaria_educacao"))

    # Buscar questões disponíveis para a série, mês e disciplina
    cursor.execute("""
        SELECT id FROM banco_questoes 
        WHERE serie_id = ? AND mes_id = ? AND disciplina_id = ?
    """, (serie_id, mes_id, disciplina_id))
    questoes = cursor.fetchall()

    if not questoes:
        flash("Nenhuma questão disponível para esta série, mês e disciplina.", "warning")
        return redirect(url_for("portal_secretaria_educacao"))

    # Criar um novo simulado
    cursor.execute("""
        INSERT INTO simulados_gerados (serie_id, mes_id, disciplina_id, status, data_envio)
        VALUES (?, ?, ?, 'enviado', datetime('now'))
    """, (serie_id, mes_id, disciplina_id))
    simulado_id = cursor.lastrowid

    # Associar questões ao simulado
    cursor.executemany("""
        INSERT INTO simulado_questoes (simulado_id, questao_id)
        VALUES (?, ?)
    """, [(simulado_id, questao[0]) for questao in questoes])

    db.commit()

    flash(f"Simulado para o ano escolar {serie_id}, mês {mes_id} e componente curricular '{disciplina}' gerado com sucesso!", "success")
    return redirect(url_for("portal_secretaria_educacao"))


@app.route('/listar_disciplinas/<int:serie_id>/<int:mes_id>')
def listar_disciplinas(serie_id, mes_id):
    db = get_db()
    cursor = db.cursor()
    
    # Buscar disciplinas que têm questões disponíveis
    cursor.execute("""
        SELECT DISTINCT d.nome 
        FROM disciplinas d
        INNER JOIN banco_questoes bq ON d.id = bq.disciplina_id
        WHERE bq.serie_id = ? AND bq.mes_id = ?
    """, (serie_id, mes_id))
    
    disciplinas = cursor.fetchall()
    return jsonify([{"disciplina": d[0]} for d in disciplinas])

@app.route("/aluno/simulado/<int:simulado_id>", methods=["GET", "POST"])
@login_required
def responder_simulado(simulado_id):
    if current_user.tipo_usuario_id != 4:  # Apenas alunos podem acessar
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))

    db = get_db()
    cursor = db.cursor()

    # Verificar se o aluno já respondeu este simulado
    cursor.execute("""
        SELECT respostas_aluno, respostas_corretas, desempenho
        FROM desempenho_simulado
        WHERE aluno_id = ? AND simulado_id = ?
    """, (current_user.id, simulado_id))
    desempenho = cursor.fetchone()

    # Buscar questões do simulado com nome da disciplina
    cursor.execute("""
        SELECT bq.id, bq.disciplina_id, d.nome as disciplina_nome, bq.questao, 
               bq.alternativa_a, bq.alternativa_b, bq.alternativa_c, 
               bq.alternativa_d, bq.alternativa_e, bq.questao_correta
        FROM simulado_questoes sq
        JOIN banco_questoes bq ON sq.questao_id = bq.id
        JOIN disciplinas d ON bq.disciplina_id = d.id
        WHERE sq.simulado_id = ?
        ORDER BY bq.disciplina_id
    """, (simulado_id,))
    questoes = cursor.fetchall()

    if not questoes:
        flash("Nenhuma questão encontrada para este simulado.", "danger")
        return redirect(url_for("alunos_bp.portal_alunos"))

    questoes_por_disciplina = {}
    for q in questoes:
        disciplina_nome = q[2]
        if disciplina_nome not in questoes_por_disciplina:
            questoes_por_disciplina[disciplina_nome] = []
        questoes_por_disciplina[disciplina_nome].append(q)

    if desempenho:  # Se já respondeu, mostrar resultados
        respostas_aluno = json.loads(desempenho[0])
        respostas_corretas = json.loads(desempenho[1])
        
        # Preparar dados para visualização
        resultados = []
        total_questoes = len(questoes)
        acertos = 0
        
        for q in questoes:
            questao_id = str(q[0])
            resposta_aluno = respostas_aluno.get(questao_id, '')
            resposta_correta = q[9]  # Já está em letra
            
            resultado = {
                'questao': q[3],
                'disciplina': q[2],
                'alternativas': [q[4], q[5], q[6], q[7], q[8]],
                'resposta_aluno': resposta_aluno,
                'resposta_correta': resposta_correta,
                'acertou': resposta_aluno == resposta_correta,
                'numero_questao': questao_id  # Adicionando o número da questão
            }
            if resultado['acertou']:
                acertos += 1
            resultados.append(resultado)
        
        percentual_acertos = (acertos / total_questoes) * 100 if total_questoes > 0 else 0
        
        return render_template(
            "responder_simulado.html",
            simulado_id=simulado_id,
            questoes_por_disciplina=questoes_por_disciplina,
            modo_visualizacao=True,
            resultados=resultados,
            total_questoes=total_questoes,
            acertos=acertos,
            percentual_acertos=percentual_acertos
        )

    if request.method == "POST":
        # Converte as respostas do aluno de números para letras (1->A, 2->B, etc)
        respostas = {}
        for key, value in request.form.items():
            if key.startswith("resposta_"):
                questao_id = key.split("_")[1]
                try:
                    respostas[questao_id] = chr(64 + int(value))  # Converte número para letra
                except (ValueError, TypeError):
                    continue
        
        # Pega as respostas corretas do banco (já estão em letra)
        respostas_corretas = {
            str(q[0]):q[9]
            for q in questoes
        }
        
        # Calcula apenas para as questões que existem em respostas_corretas
        total_questoes = len(respostas_corretas)
        if total_questoes > 0:
            respostas_certas = sum(
                1 for key, value in respostas.items() 
                if key in respostas_corretas and respostas_corretas[key] == value
            )
            desempenho = (respostas_certas / total_questoes) * 100
        else:
            respostas_certas = 0
            desempenho = 0

        # Inserir dados na tabela desempenho_simulado
        cursor.execute("""
            INSERT INTO desempenho_simulado (aluno_id, simulado_id, escola_id, serie_id, codigo_ibge, respostas_aluno, respostas_corretas, desempenho)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            current_user.id,
            simulado_id,
            current_user.escola_id,
            current_user.serie_id,
            current_user.codigo_ibge,
            json.dumps(respostas),
            json.dumps(respostas_corretas),
            desempenho
        ))
        db.commit()

        flash(f"Simulado respondido com sucesso! Você acertou {respostas_certas} de {total_questoes} questões.", "success")
        return redirect(url_for("alunos_bp.portal_alunos"))

    return render_template(
        "responder_simulado.html",
        simulado_id=simulado_id,
        questoes_por_disciplina=questoes_por_disciplina,
        modo_visualizacao=False
    )
    
# @app.route('/aluno/<int:aluno_id>/relatorio_pdf')
# @login_required
# def gerar_relatorio_pdf(aluno_id):
#     # Lógica para gerar o PDF
#     pass  # Atualize com os dados e o conteúdo do PDF




@app.route('/gerar_parecer/<int:aluno_id>', methods=['GET'])
def gerar_parecer(aluno_id):
    desempenho = session.get('desempenho', {}).get(aluno_id, None)

    if not desempenho:
        return "Nenhum desempenho registrado para este aluno.", 404

    try:
        # Gera o parecer com a IA
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Você é um gerador de pareceres escolares baseado em desempenho."},
                {
                    "role": "user",
                    "content": (
                        f"Um aluno respondeu {desempenho['total']} perguntas, acertando {desempenho['acertos']}. "
                        "Gere um parecer detalhado sobre o desempenho do aluno, destacando pontos fortes, fracos e sugestões de melhoria."
                    )
                }
            ],
            max_tokens=500,
            temperature=0.7
        )
        parecer = response['choices'][0]['message']['content']
        return render_template('parecer.html', parecer=parecer, aluno_id=aluno_id)

    except openai.error.OpenAIError as e:
        return f"Erro ao gerar parecer: {e}", 500
    
@app.route('/ver_relatorio_escola', methods=['GET'])
@login_required
def ver_relatorio_escola():
    if current_user.tipo_usuario_id != 2:
        return redirect(url_for("home"))

    # Obter a escola alocada ao administrador
    escola_alocada = get_escola_alocada(current_user.id)
    if not escola_alocada:
        flash("Erro ao localizar a escola associada ao usuário.", "danger")
        return redirect(url_for("portal_administracao"))

    # Buscar os tipos de ensino disponíveis para a escola
    tipos_ensino = get_tipos_ensino(escola_alocada["id"])

    return render_template(
        'ver_relatorio_escola.html',
        escola_alocada=escola_alocada,
        tipos_ensino=tipos_ensino
    )



@app.route('/ver_relatorio_secretaria', methods=['GET'])
@login_required
def ver_relatorio_secretaria():
    if current_user.tipo_usuario_id != 5:  # Secretaria de Educação
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("home"))

    tipo_relatorio = request.args.get('tipo_relatorio', 'rede')
    escola_id = request.args.get('escola_id')

    db = get_db()
    cursor = db.cursor()

    try:
        print(f"[DEBUG] Tipo de relatório solicitado: {tipo_relatorio}")
        print(f"[DEBUG] Escola ID (se fornecido): {escola_id}")

        if tipo_relatorio == 'rede':
            print("[DEBUG] Gerando relatório geral da rede municipal...")
            # Relatório Geral da Rede Municipal
            cursor.execute("""
                SELECT escolas.nome_da_escola AS escola, 
                       COUNT(usuarios.id) AS total_alunos, 
                       AVG(desempenho_simulado.desempenho) AS media_desempenho
                FROM escolas
                LEFT JOIN usuarios ON escolas.id = usuarios.escola_id AND usuarios.tipo_usuario_id = 4
                LEFT JOIN desempenho_simulado ON usuarios.id = desempenho_simulado.aluno_id
                WHERE escolas.codigo_ibge = ?
                GROUP BY escolas.id
            """, (current_user.codigo_ibge,))
            relatorio_geral = cursor.fetchall()
            print(f"[DEBUG] Relatório Geral da Rede: {relatorio_geral}")

            # Geração de parecer pela IA
            parecer_ia = gerar_parecer_ia(relatorio_geral) if relatorio_geral else "Sem dados para gerar parecer."

            return render_template(
                'relatorio_rede_municipal.html',
                tipo_relatorio=tipo_relatorio,
                relatorio_geral=relatorio_geral,
                parecer_ia=parecer_ia
            )

        elif tipo_relatorio == 'escola' and escola_id:
            print("[DEBUG] Gerando relatório detalhado por escola...")
            # Relatório Detalhado por Escola
            cursor.execute("""
                SELECT serie_ids.nome AS serie, 
                       turmas.nome AS turma, 
                       COUNT(usuarios.id) AS total_alunos, 
                       AVG(desempenho_simulado.desempenho) AS media_desempenho
                FROM turmas
                LEFT JOIN serie_ids ON turmas.serie_id = serie_ids.id
                LEFT JOIN usuarios ON turmas.id = usuarios.turma_id AND usuarios.tipo_usuario_id = 4
                LEFT JOIN desempenho_simulado ON usuarios.id = desempenho_simulado.aluno_id
                WHERE turmas.escola_id = ?
                GROUP BY turmas.id
            """, (escola_id,))
            relatorio_por_escola = cursor.fetchall()
            print(f"[DEBUG] Relatório por Escola: {relatorio_por_escola}")

            # Geração de parecer pela IA
            parecer_ia = gerar_parecer_ia(relatorio_por_escola) if relatorio_por_escola else "Sem dados para gerar parecer."

            return render_template(
                'relatorio_por_escola.html',
                tipo_relatorio=tipo_relatorio,
                relatorio_por_escola=relatorio_por_escola,
                parecer_ia=parecer_ia
            )

        print("[DEBUG] Buscando todas as escolas cadastradas...")
        # Buscar todas as escolas cadastradas
        cursor.execute("SELECT id, nome_da_escola FROM escolas WHERE codigo_ibge = ?", (current_user.codigo_ibge,))
        escolas = cursor.fetchall()
        print(f"[DEBUG] Escolas encontradas: {escolas}")

        return render_template(
            'ver_relatorio_secretaria.html',
            tipo_relatorio=tipo_relatorio,
            escolas=escolas
        )

    except Exception as e:
        print(f"[ERRO] Erro ao gerar relatório: {e}")
        flash("Erro ao gerar relatório.", "danger")
        return redirect(url_for("portal_secretaria_educacao"))

@app.route('/relatorios_dashboard', methods=['GET'])
@login_required
def relatorios_dashboard():
    if current_user.tipo_usuario_id != 5:  # Apenas para Secretaria de Educação
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))

    db = get_db()
    cursor = db.cursor()

    # Buscar o `codigo_ibge` do usuário
    cursor.execute("SELECT codigo_ibge FROM usuarios WHERE id = ?", (current_user.id,))
    codigo_ibge= cursor.fetchone()[0]

    # Desempenho geral
    cursor.execute("""
        WITH UltimoDesempenho AS (
            SELECT 
                d.aluno_id,
                AVG(d.desempenho) as desempenho,
                ROW_NUMBER() OVER (PARTITION BY d.aluno_id ORDER BY d.data_resposta DESC) as rn
            FROM desempenho_simulado d
            JOIN usuarios u ON d.aluno_id = u.id
            WHERE u.codigo_ibge = ? AND u.tipo_usuario_id = 4
            GROUP BY d.aluno_id
        )
        SELECT AVG(desempenho) AS desempenho_geral
        FROM UltimoDesempenho
        WHERE rn = 1
    """, (codigo_ibge,))
    desempenho_geral = cursor.fetchone()[0] or 0  # Define 0 se não houver dados

    # Melhor escola
    cursor.execute("""
        WITH DesempenhoEscola AS (
            SELECT 
                u.escola_id,
                AVG(d.desempenho) as media_escola
            FROM desempenho_simulado d
            JOIN usuarios u ON d.aluno_id = u.id
            WHERE u.codigo_ibge = ? AND u.tipo_usuario_id = 4
            GROUP BY u.escola_id
        )
        SELECT e.nome_da_escola, de.media_escola
        FROM escolas e
        JOIN DesempenhoEscola de ON e.id = de.escola_id
        WHERE e.codigo_ibge = ?
        ORDER BY de.media_escola DESC
        LIMIT 1
    """, (codigo_ibge, codigo_ibge))
    melhor_escola = cursor.fetchone() or ("Nenhuma escola", 0)  # Define valores padrão se não houver dados

    # Desempenho geral por mês
    cursor.execute("""
        WITH DesempenhoMensal AS (
            SELECT 
                d.aluno_id,
                AVG(d.desempenho) as media_desempenho,
                CASE strftime('%m', d.data_resposta)
                    WHEN '01' THEN 'Janeiro'
                    WHEN '02' THEN 'Fevereiro'
                    WHEN '03' THEN 'Março'
                    WHEN '04' THEN 'Abril'
                    WHEN '05' THEN 'Maio'
                    WHEN '06' THEN 'Junho'
                    WHEN '07' THEN 'Julho'
                    WHEN '08' THEN 'Agosto'
                    WHEN '09' THEN 'Setembro'
                    WHEN '10' THEN 'Outubro'
                    WHEN '11' THEN 'Novembro'
                    WHEN '12' THEN 'Dezembro'
                END as mes_nome,
                strftime('%m', d.data_resposta) as mes_numero
            FROM desempenho_simulado d
            JOIN usuarios u ON d.aluno_id = u.id
            WHERE u.codigo_ibge = ? AND u.tipo_usuario_id = 4
            GROUP BY d.aluno_id, strftime('%m', d.data_resposta)
        )
        SELECT 
            mes_nome, 
            AVG(media_desempenho) AS media_desempenho
        FROM DesempenhoMensal
        GROUP BY mes_nome, mes_numero
        ORDER BY mes_numero
    """, (codigo_ibge,))
    desempenho_mensal = cursor.fetchall()

    # Ranking de escolas
    cursor.execute("""
        WITH DesempenhoEscola AS (
            SELECT 
                u.escola_id,
                AVG(d.desempenho) as media_escola
            FROM desempenho_simulado d
            JOIN usuarios u ON d.aluno_id = u.id
            WHERE u.codigo_ibge = ? AND u.tipo_usuario_id = 4
            GROUP BY u.escola_id
        )
        SELECT e.nome_da_escola, de.media_escola
        FROM escolas e
        JOIN DesempenhoEscola de ON e.id = de.escola_id
        WHERE e.codigo_ibge = ?
        ORDER BY de.media_escola DESC
    """, (codigo_ibge, codigo_ibge))
    ranking_escolas = cursor.fetchall()

    # Ranking dos 5 melhores alunos por ano escolar
    cursor.execute("""
        WITH DesempenhoAluno AS (
            SELECT 
                d.aluno_id,
                AVG(d.desempenho) as media_desempenho
            FROM desempenho_simulado d
            JOIN usuarios u ON d.aluno_id = u.id
            WHERE u.codigo_ibge = ? AND u.tipo_usuario_id = 4
            GROUP BY d.aluno_id
        )
        SELECT s.nome AS ano_escolar, u.nome AS aluno, da.media_desempenho AS media_aluno
        FROM series s
        JOIN usuarios u ON s.id = u.serie_id
        JOIN DesempenhoAluno da ON u.id = da.aluno_id
        ORDER BY s.id, media_aluno DESC
        LIMIT 5
    """, (codigo_ibge,))
    ranking_alunos = cursor.fetchall()

    # Contagem de alunos por faixa de desempenho
    cursor.execute("""
        WITH DesempenhoMedioAluno AS (
            SELECT 
                d.aluno_id,
                AVG(d.desempenho) as media_desempenho
            FROM desempenho_simulado d
            JOIN usuarios u ON d.aluno_id = u.id
            WHERE u.codigo_ibge = ? AND u.tipo_usuario_id = 4
            GROUP BY d.aluno_id
        )
        SELECT 
            COUNT(CASE WHEN media_desempenho BETWEEN 0 AND 20 THEN 1 END) as faixa_0_20,
            COUNT(CASE WHEN media_desempenho BETWEEN 21 AND 40 THEN 1 END) as faixa_21_40,
            COUNT(CASE WHEN media_desempenho BETWEEN 41 AND 60 THEN 1 END) as faixa_41_60,
            COUNT(CASE WHEN media_desempenho BETWEEN 61 AND 80 THEN 1 END) as faixa_61_80,
            COUNT(CASE WHEN media_desempenho BETWEEN 81 AND 100 THEN 1 END) as faixa_81_100
        FROM DesempenhoMedioAluno
    """, (codigo_ibge,))
    faixas = cursor.fetchone()

    return render_template(
        'relatorios_dashboard.html',
        desempenho_geral=desempenho_geral,
        melhor_escola=melhor_escola,
        desempenho_mensal=desempenho_mensal,
        ranking_escolas=ranking_escolas,
        ranking_alunos=ranking_alunos,
        faixa_0_20=faixas[0],
        faixa_21_40=faixas[1],
        faixa_41_60=faixas[2],
        faixa_61_80=faixas[3],
        faixa_81_100=faixas[4]
    )

@app.route('/detalhar_desempenho', methods=['GET'])
@login_required
def detalhar_desempenho():
    if current_user.tipo_usuario_id != 5:
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))

    db = get_db()
    cursor = db.cursor()

    # Desempenho por Ano Escolar
    cursor.execute("""
        SELECT s.nome AS ano_escolar, AVG(d.desempenho) AS media_ano
        FROM desempenho_simulado d
        JOIN usuarios u ON d.aluno_id = u.id
        JOIN series s ON u.serie_id = s.id
        WHERE u.codigo_ibge = ?
        GROUP BY s.id
        ORDER BY s.id
    """, (current_user.codigo_ibge,))
    desempenho_por_ano = cursor.fetchall()

    # Desempenho por Disciplina - Atualizado para usar disciplina_id
    cursor.execute("""
        SELECT d.nome AS componente_curricular, d.id AS disciplina_id, AVG(ds.desempenho) AS media_disciplina
        FROM desempenho_simulado ds
        JOIN simulado_questoes sq ON ds.simulado_id = sq.simulado_id
        JOIN banco_questoes bq ON sq.questao_id = bq.id
        JOIN disciplinas d ON bq.disciplina_id = d.id
        WHERE ds.codigo_ibge = ?
        GROUP BY bq.disciplina_id, d.nome, d.id
        ORDER BY media_disciplina DESC;
    """, (current_user.codigo_ibge,))
    desempenho_por_disciplina = cursor.fetchall()

    return render_template(
        "detalhar_desempenho.html",
        desempenho_por_ano=desempenho_por_ano,
        desempenho_por_disciplina=desempenho_por_disciplina
    )


@app.route('/detalhar_desempenho/disciplina/<int:disciplina_id>', methods=['GET'])
@login_required
def detalhar_assuntos_por_disciplina(disciplina_id):
    if current_user.tipo_usuario_id != 5:
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))

    db = get_db()
    cursor = db.cursor()

    # Verificar se a disciplina existe e obter seu nome
    cursor.execute("""
        SELECT id, nome FROM disciplinas WHERE id = ?
    """, (disciplina_id,))
    disciplina = cursor.fetchone()
    if not disciplina:
        flash("Disciplina não encontrada.", "danger")
        return redirect(url_for("detalhar_desempenho"))

    # Buscar questões e suas estatísticas de acerto
    cursor.execute("""
        SELECT 
            bq.id,
            bq.questao,
            bq.assunto,
            COUNT(DISTINCT ds.aluno_id) as total_respostas,
            SUM(CASE 
                WHEN json_extract(ds.respostas_aluno, '$.' || bq.id) = bq.questao_correta 
                THEN 1 
                ELSE 0 
            END) as total_acertos,
            (SUM(CASE 
                WHEN json_extract(ds.respostas_aluno, '$.' || bq.id) = bq.questao_correta 
                THEN 1 
                ELSE 0 
            END) * 100.0 / COUNT(DISTINCT ds.aluno_id)) as porcentagem_acertos
        FROM banco_questoes bq
        JOIN simulado_questoes sq ON bq.id = sq.questao_id
        JOIN desempenho_simulado ds ON sq.simulado_id = ds.simulado_id
        WHERE bq.disciplina_id = ? AND ds.codigo_ibge = ?
        GROUP BY bq.id, bq.questao, bq.assunto
        ORDER BY porcentagem_acertos DESC;
    """, (disciplina_id, current_user.codigo_ibge))
    estatisticas_questoes = cursor.fetchall()

    return render_template(
        "detalhar_assunto.html",
        disciplina=disciplina[1],  # Nome da disciplina
        disciplina_id=disciplina_id,
        estatisticas_questoes=estatisticas_questoes
    )

@app.route('/relatorio_escolas')
@login_required
def relatorio_escolas():
    if current_user.tipo_usuario_id != 5:  # Verifica se é secretaria de educação
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))

    db = get_db()
    cursor = db.cursor()

    # Buscar desempenho geral das escolas
    cursor.execute("""
        SELECT 
            e.id as id,
            e.nome_da_escola as nome,
            COUNT(DISTINCT ds.aluno_id) as total_alunos,
            COALESCE(ROUND(AVG(ds.desempenho), 2), 0) as media_geral
        FROM escolas e
        LEFT JOIN usuarios u ON e.id = u.escola_id
        LEFT JOIN desempenho_simulado ds ON u.id = ds.aluno_id
        WHERE e.codigo_ibge = ? AND u.tipo_usuario_id = 4
        GROUP BY e.id, e.nome_da_escola
        ORDER BY media_geral DESC
    """, (current_user.codigo_ibge,))
    
    # Converter resultados em lista de dicionários
    colunas_escolas = ['id', 'nome', 'total_alunos', 'media_geral']
    escolas = [dict(zip(colunas_escolas, escola)) for escola in cursor.fetchall()]

    # Buscar desempenho por série
    cursor.execute("""
        SELECT 
            s.id as serie_id,
            s.nome as serie_nome,
            e.id as escola_id,
            e.nome_da_escola as escola_nome,
            COUNT(DISTINCT ds.aluno_id) as total_alunos,
            COALESCE(ROUND(AVG(ds.desempenho), 2), 0) as media
        FROM series s
        JOIN usuarios u ON s.id = u.serie_id
        JOIN escolas e ON u.escola_id = e.id
        LEFT JOIN desempenho_simulado ds ON u.id = ds.aluno_id
        WHERE e.codigo_ibge = ? AND u.tipo_usuario_id = 4
        GROUP BY s.id, s.nome, e.id, e.nome_da_escola
        ORDER BY s.id, media DESC
    """, (current_user.codigo_ibge,))
    
    # Converter resultados em lista de dicionários
    colunas_series = ['serie_id', 'serie_nome', 'escola_id', 'escola_nome', 'total_alunos', 'media']
    resultados_series = [dict(zip(colunas_series, serie)) for serie in cursor.fetchall()]

    # Buscar melhores alunos por série
    cursor.execute("""
        SELECT 
            s.id as serie_id,
            s.nome as serie_nome,
            u.nome as aluno_nome,
            COALESCE(ROUND(AVG(ds.desempenho), 2), 0) as media
        FROM series s
        JOIN usuarios u ON s.id = u.serie_id
        LEFT JOIN desempenho_simulado ds ON u.id = ds.aluno_id
        WHERE u.codigo_ibge = ? AND u.tipo_usuario_id = 4
        GROUP BY s.id, s.nome, u.id, u.nome
        HAVING media > 0
        ORDER BY s.id, media DESC
    """, (current_user.codigo_ibge,))

    # Processar melhores alunos
    melhores_alunos = {}
    for row in cursor.fetchall():
        if row[0] not in melhores_alunos:  # Se ainda não temos um melhor aluno para esta série
            melhores_alunos[row[0]] = {
                'ano_escolar': row[1].replace('Série', 'Ano'),
                'nome': row[2],
                'media': row[3]
            }

    # Buscar contagem de alunos ativos por série
    cursor.execute("""
        SELECT 
            s.id as serie_id,
            COUNT(DISTINCT u.id) as total_alunos,
            COUNT(DISTINCT ds.aluno_id) as alunos_ativos
        FROM series s
        JOIN usuarios u ON s.id = u.serie_id
        LEFT JOIN desempenho_simulado ds ON u.id = ds.aluno_id
        WHERE u.codigo_ibge = ? AND u.tipo_usuario_id = 4
        GROUP BY s.id
    """, (current_user.codigo_ibge,))

    alunos_por_serie = {row[0]: {'total': row[1], 'ativos': row[2]} for row in cursor.fetchall()}

    # Organizar dados por série
    series = {}
    for resultado in resultados_series:
        serie_id = resultado['serie_id']
        if serie_id not in series:
            series[serie_id] = {
                'nome': resultado['serie_nome'],
                'escolas': [],
                'total_alunos': alunos_por_serie.get(serie_id, {'total': 0, 'ativos': 0})['total'],
                'alunos_ativos': alunos_por_serie.get(serie_id, {'total': 0, 'ativos': 0})['ativos']
            }
        series[serie_id]['escolas'].append({
            'id': resultado['escola_id'],
            'nome': resultado['escola_nome'],
            'total_alunos': resultado['total_alunos'],
            'media': resultado['media']
        })

    return render_template('relatorio_escolas.html', 
                         escolas=escolas,
                         series=series.values(),
                         melhores_alunos=list(melhores_alunos.values()))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

@app.route('/relatorio_escolas/excel')
@login_required
def relatorio_escolas_excel():
    if current_user.tipo_usuario_id != 5:
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))

    db = get_db()
    cursor = db.cursor()

    # Buscar desempenho geral das escolas
    cursor.execute("""
        SELECT 
            e.id as id,
            e.nome_da_escola as nome,
            COUNT(DISTINCT ds.aluno_id) as total_alunos,
            COALESCE(ROUND(AVG(ds.desempenho), 2), 0) as media_geral
        FROM escolas e
        LEFT JOIN usuarios u ON e.id = u.escola_id
        LEFT JOIN desempenho_simulado ds ON u.id = ds.aluno_id
        WHERE e.codigo_ibge = ? AND u.tipo_usuario_id = 4
        GROUP BY e.id, e.nome_da_escola
        ORDER BY media_geral DESC
    """, (current_user.codigo_ibge,))
    
    # Converter resultados em lista de dicionários
    colunas_escolas = ['id', 'nome', 'total_alunos', 'media_geral']
    escolas = [dict(zip(colunas_escolas, escola)) for escola in cursor.fetchall()]

    # Buscar desempenho por série
    cursor.execute("""
        SELECT 
            s.id as serie_id,
            s.nome as serie_nome,
            e.id as escola_id,
            e.nome_da_escola as escola_nome,
            COUNT(DISTINCT ds.aluno_id) as total_alunos,
            COALESCE(ROUND(AVG(ds.desempenho), 2), 0) as media
        FROM series s
        JOIN usuarios u ON s.id = u.serie_id
        JOIN escolas e ON u.escola_id = e.id
        LEFT JOIN desempenho_simulado ds ON u.id = ds.aluno_id
        WHERE e.codigo_ibge = ? AND u.tipo_usuario_id = 4
        GROUP BY s.id, s.nome, e.id, e.nome_da_escola
        ORDER BY s.id, media DESC
    """, (current_user.codigo_ibge,))
    
    # Converter resultados em lista de dicionários
    colunas_series = ['serie_id', 'serie_nome', 'escola_id', 'escola_nome', 'total_alunos', 'media']
    resultados_series = [dict(zip(colunas_series, serie)) for serie in cursor.fetchall()]

    # Organizar dados por série
    series = {}
    for resultado in resultados_series:
        serie_id = resultado['serie_id']
        if serie_id not in series:
            series[serie_id] = {
                'nome': resultado['serie_nome'],
                'escolas': []
            }
        series[serie_id]['escolas'].append({
            'id': resultado['escola_id'],
            'nome': resultado['escola_nome'],
            'total_alunos': resultado['total_alunos'],
            'media': resultado['media']
        })

    # Criar novo workbook
    wb = Workbook()
    
    # Estilo para cabeçalhos
    header_style = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Planilha de Ranking Geral
    ws = wb.active
    ws.title = "Ranking Geral"
    
    # Cabeçalhos
    headers = ["Posição", "Escola", "Total de Alunos", "Média Geral"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Dados do ranking geral
    for row, escola in enumerate(escolas, 2):
        ws.cell(row=row, column=1, value=f"{row-1}º")
        ws.cell(row=row, column=2, value=escola['nome'])
        ws.cell(row=row, column=3, value=escola['total_alunos'])
        ws.cell(row=row, column=4, value=f"{escola['media_geral']}%")
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Criar planilhas para cada série
    for serie in series.values():
        ws = wb.create_sheet(title=serie['nome'])
        
        # Cabeçalhos
        headers = ["Posição", "Escola", "Total de Alunos", "Média"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_style
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Dados da série
        for row, escola in enumerate(serie['escolas'], 2):
            ws.cell(row=row, column=1, value=f"{row-1}º")
            ws.cell(row=row, column=2, value=escola['nome'])
            ws.cell(row=row, column=3, value=escola['total_alunos'])
            ws.cell(row=row, column=4, value=f"{escola['media']}%")
        
        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Salvar o arquivo em memória
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='relatorio_escolas.xlsx'
    )


from datetime import datetime

from flask import render_template, make_response, flash, redirect, url_for
from flask_login import login_required, current_user
from datetime import datetime
from weasyprint import HTML
import tempfile
import os

from flask import make_response
import matplotlib.pyplot as plt
import numpy as np
import tempfile
import os
from weasyprint import HTML
from flask import make_response

def gerar_grafico_distribuicao(distribuicao):
    """Gera um gráfico de distribuição de desempenho e salva como imagem."""

    categorias = list(distribuicao.keys())
    valores = list(distribuicao.values())

    # Cores para cada faixa
    cores = ['#dc3545', '#fd7e14', '#ffc107', '#28a745']

    fig, ax = plt.subplots(figsize=(5, 3))
    ax.barh(categorias, valores, color=cores)

    ax.set_xlabel('Número de Alunos')
    ax.set_title('Distribuição de Desempenho')

    # Salvar o gráfico como imagem temporária
    img_temp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    plt.savefig(img_temp.name, format='png', bbox_inches='tight')
    plt.close(fig)
    
    return img_temp.name  # Retorna o caminho da imagem temporária

@app.route('/relatorio_escolas/pdf')
@login_required
def relatorio_escolas_pdf():
    if current_user.tipo_usuario_id != 5:
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("index"))

    db = get_db()
    cursor = db.cursor()

    # Buscar dados por série
    cursor.execute("""
        SELECT 
            s.id as serie_id,
            s.nome as serie_nome,
            COUNT(DISTINCT u.id) as total_alunos,
            COUNT(DISTINCT e.id) as total_escolas,
            COALESCE(ROUND(AVG(ds.desempenho), 2), 0) as media_geral
        FROM series s
        JOIN usuarios u ON s.id = u.serie_id
        JOIN escolas e ON u.escola_id = e.id
        LEFT JOIN desempenho_simulado ds ON u.id = ds.aluno_id
        WHERE u.tipo_usuario_id = 4 AND e.codigo_ibge = ?
        GROUP BY s.id, s.nome
        ORDER BY s.nome
    """, (current_user.codigo_ibge,))
    
    series_dados = cursor.fetchall()
    
    # Buscar dados das escolas por série
    cursor.execute("""
        SELECT 
            s.id as serie_id,
            e.id as escola_id,
            e.nome_da_escola as escola_nome,
            COUNT(DISTINCT u.id) as total_alunos,
            COALESCE(ROUND(AVG(ds.desempenho), 2), 0) as media
        FROM series s
        JOIN usuarios u ON s.id = u.serie_id
        JOIN escolas e ON u.escola_id = e.id
        LEFT JOIN desempenho_simulado ds ON u.id = ds.aluno_id
        WHERE u.tipo_usuario_id = 4 AND e.codigo_ibge = ?
        GROUP BY s.id, e.id, e.nome_da_escola
        ORDER BY s.id, media DESC
    """, (current_user.codigo_ibge,))
    
    escolas_dados = cursor.fetchall()

    # Buscar dados do município
    cursor.execute("""
        SELECT 
            c.nome as municipio,
            COUNT(DISTINCT u.id) as total_alunos,
            COUNT(DISTINCT e.id) as total_escolas,
            COALESCE(ROUND(AVG(ds.desempenho), 2), 0) as media_geral
        FROM cidades c
        LEFT JOIN escolas e ON c.codigo_ibge = e.codigo_ibge
        LEFT JOIN usuarios u ON e.id = u.escola_id AND u.tipo_usuario_id = 4
        LEFT JOIN desempenho_simulado ds ON u.id = ds.aluno_id
        WHERE c.codigo_ibge = ?
        GROUP BY c.codigo_ibge, c.nome
    """, (current_user.codigo_ibge,))
    
    resultado_municipio = cursor.fetchone()
    if resultado_municipio:
        dados_municipio = dict(zip(['municipio', 'total_alunos', 'total_escolas', 'media_geral'], resultado_municipio))
    else:
        dados_municipio = {'municipio': 'Desconhecido', 'total_alunos': 0, 'total_escolas': 0, 'media_geral': 0}

    # Organizar dados por série
    series = {}
    for serie in series_dados:
        serie_id = serie[0]  # serie_id
        series[serie_id] = {
            'id': serie_id,
            'nome': serie[1],  # serie_nome
            'total_alunos': serie[2],  # total_alunos
            'total_escolas': serie[3],  # total_escolas
            'media_geral': serie[4],  # media_geral
            'escolas': []
        }

    # Adicionar escolas para cada série
    for escola in escolas_dados:
        serie_id = escola[0]  # serie_id
        if serie_id in series:
            series[serie_id]['escolas'].append({
                'id': escola[1],  # escola_id
                'nome': escola[2],  # escola_nome
                'total_alunos': escola[3],  # total_alunos
                'media': escola[4]  # media
            })

    # Criar lista de todas as escolas ordenadas por média
    todas_escolas = []
    for escola in escolas_dados:
        todas_escolas.append({
            'nome': escola[2],  # escola_nome
            'media_geral': escola[4]  # media
        })
    todas_escolas.sort(key=lambda x: x['media_geral'], reverse=True)

    # Renderizar HTML do relatório
    html = render_template(
        'relatorio_escolas_view.html',
        dados_municipio=dados_municipio,
        series=series.values(),
        escolas=todas_escolas,
        data_atual=datetime.now()
    )

    # Criar arquivo temporário para o HTML
    with tempfile.NamedTemporaryFile(suffix='.html', delete=False) as f:
        f.write(html.encode('utf-8'))
        html_path = f.name

    try:
        # Gerar PDF a partir do HTML
        pdf = HTML(filename=html_path).write_pdf()

        # Criar resposta para exibição direta no navegador
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'inline; filename=relatorio_escolas.pdf'

        return response
    finally:
        # Remover arquivos temporários
        if os.path.exists(html_path):
            os.unlink(html_path)


@app.route('/gerar_pdf_relatorio', methods=['GET'])
@login_required
def gerar_pdf_relatorio():
    if current_user.tipo_usuario_id != 5:  # Secretaria de Educação
        flash("Acesso não autorizado.", "danger")
        return redirect(url_for("home"))

    tipo_relatorio = request.args.get('tipo_relatorio', 'rede')
    escola_id = request.args.get('escola_id')

    db = get_db()
    cursor = db.cursor()
    
    if tipo_relatorio == 'rede':
        # Relatório Geral da Rede Municipal
        cursor.execute("""
            SELECT escolas.nome_da_escola AS escola, 
                   COUNT(alunos.id) AS total_alunos, 
                   AVG(desempenho_simulado.desempenho) AS media_desempenho
            FROM escolas
            LEFT JOIN alunos ON escolas.id = alunos.escola_id
            LEFT JOIN desempenho_simulado ON alunos.id = desempenho_simulado.aluno_id
            WHERE escolas.codigo_ibge = ?
            GROUP BY escolas.id
        """, (current_user.codigo_ibge,))
        relatorio_geral = cursor.fetchall()
        parecer_ia = gerar_parecer_ia(relatorio_geral)
        
        pdf = gerar_pdf(relatorio_geral, parecer_ia)

    
def gerar_pdf(relatorio, parecer_ia, tipo_relatorio):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt=f"Relatório de Desempenho - {tipo_relatorio.capitalize()}", ln=True, align='C')

    for row in relatorio:
        for key, value in row.items():
            pdf.cell(200, 10, txt=f"{key}: {value}", ln=True)
        pdf.cell(200, 10, txt="", ln=True)  # Linha em branco

    pdf.cell(200, 10, txt="Parecer da IA", ln=True, align='C')
    pdf.multi_cell(200, 10, txt=parecer_ia)

    return pdf


@app.route("/relatorio_geral", methods=["GET"])
@login_required
def relatorio_geral():
    if current_user.tipo_usuario_id != [3, 1, 2]:
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT assunto.assunto, COUNT(respostas.id) AS total_respostas, 
               COALESCE(SUM(respostas.correta), 0) AS total_corretas
        FROM respostas
        JOIN simulados ON respostas.simulado_id = simulados.id
        JOIN assunto ON simulados.assunto = assunto.id
        WHERE simulados.professor_id = ?
        GROUP BY assunto.assunto
    """, (current_user.id,))
    relatorio = cursor.fetchall()

    if not relatorio:
        return render_template("relatorio_geral.html", relatorio=None, error="Nenhum dado disponível.")

    return render_template("relatorio_geral.html", relatorio=relatorio)

from weasyprint import HTML, CSS

@app.route('/aluno/<int:aluno_id>/relatorio_pdf', methods=['GET'])
@login_required
def gerar_relatorio_pdf_aluno(aluno_id):
    db = get_db()
    cursor = db.cursor()

    # Recuperar informações do aluno
    cursor.execute("""
        SELECT nome, turma_id FROM usuarios WHERE id = ? AND tipo_usuario_id = 4
    """, (aluno_id,))
    aluno = cursor.fetchone()
    if not aluno:
        return "Aluno não encontrado.", 404

    nome_aluno, turma_id = aluno

    # Consolidar desempenho do aluno por disciplina e assunto
    cursor.execute("""
        SELECT d.nome AS disciplina, a.nome AS assunto, 
               SUM(r.acertos) AS total_acertos, 
               SUM(r.total_questoes) AS total_questoes
        FROM desempenho r
        JOIN disciplinas d ON r.disciplina_id = d.id
        JOIN assuntos a ON r.assunto_id = a.id
        WHERE r.aluno_id = ?
        GROUP BY d.nome, a.nome
    """, (aluno_id,))
    desempenho = cursor.fetchall()

    print(f"[DEBUG] Dados de desempenho para aluno {aluno_id}: {desempenho}")

    if not desempenho:
        desempenho_por_disciplina = {}
    else:
        # Estruturar desempenho por disciplina
        desempenho_por_disciplina = {}
        for disciplina, assunto, total_acertos, total_questoes in desempenho:
            porcentagem = round((total_acertos / total_questoes) * 100, 2)
            if disciplina not in desempenho_por_disciplina:
                desempenho_por_disciplina[disciplina] = []
            desempenho_por_disciplina[disciplina].append({
                'assunto': assunto,
                'total': total_questoes,
                'acertos': total_acertos,
                'porcentagem': porcentagem
            })

    # Gerar parecer com a IA por disciplina
    pareceres = {}
    if desempenho_por_disciplina:
        for disciplina, resultados in desempenho_por_disciplina.items():
            mensagem = f"""
            O aluno {nome_aluno} apresentou os seguintes resultados na disciplina {disciplina}:
            """
            for resultado in resultados:
                mensagem += f"""
                Assunto: {resultado['assunto']}, Acertos: {resultado['acertos']} de {resultado['total']} ({resultado['porcentagem']}%).
                """
            mensagem += """
            Baseado nesses dados, forneça um parecer breve destacando os principais pontos fortes 
            e uma ou duas sugestões de melhoria específicas para esta disciplina.
            """

            try:
                response = openai.ChatCompletion.create(
                    model="gpt-3.5-turbo",
                    messages=[
                        {"role": "system", "content": "Você é um gerador de pareceres escolares humanizados."},
                        {"role": "user", "content": mensagem}
                    ],
                    max_tokens=250,
                    temperature=0.7
                )
                pareceres[disciplina] = response['choices'][0]['message']['content']
            except Exception as e:
                pareceres[disciplina] = f"Erro ao gerar parecer: {e}"

    # Gerar gráfico de desempenho usando matplotlib
    import matplotlib.pyplot as plt
    from io import BytesIO
    import base64

    disciplinas = list(desempenho_por_disciplina.keys())
    porcentagens = [
        sum([resultado['porcentagem'] for resultado in resultados]) / len(resultados)
        for resultados in desempenho_por_disciplina.values()
    ]

    plt.figure(figsize=(10, 6))
    plt.bar(disciplinas, porcentagens, color="skyblue")
    plt.xlabel("Disciplinas")
    plt.ylabel("Porcentagem de Acertos")
    plt.title(f"Desempenho Geral do Aluno {nome_aluno}")

    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    grafico_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    buffer.close()

    # Gerar HTML para o PDF
    professor_nome = current_user.nome
    rendered_html = render_template(
        'relatorio_pdf.html',
        nome_aluno=nome_aluno,
        desempenho_por_disciplina=desempenho_por_disciplina,
        pareceres=pareceres,
        professor_nome=professor_nome,
        grafico_base64=grafico_base64
    )

    # Gerar o PDF usando WeasyPrint
    pdf = HTML(string=rendered_html).write_pdf(stylesheets=[CSS(string="""
        body { font-family: Arial, sans-serif; margin: 20px; }
        h1, h2, h3 { color: #4CAF50; }
        .table { width: 100%; border-collapse: collapse; }
        .table th, .table td { border: 1px solid #ddd; padding: 8px; text-align: center; }
        .table th { background-color: #4CAF50; color: white; }
        .footer { text-align: center; margin-top: 20px; font-size: 0.8rem; color: #555; }
    """)])

    # Retornar o PDF como resposta para download
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=relatorio_{nome_aluno}.pdf'
    return response

from datetime import datetime

@app.route('/aluno/<int:aluno_id>/relatorio')
@login_required
def selecionar_relatorio(aluno_id):
    db = get_db()
    cursor = db.cursor()

    # Recuperar informações do aluno
    cursor.execute("""
        SELECT nome FROM usuarios WHERE id = ? AND tipo_usuario_id = 4
    """, (aluno_id,))
    aluno = cursor.fetchone()

    if not aluno:
        return "Aluno não encontrado.", 404

    nome_aluno = aluno[0]

    # Buscar simulados respondidos pelo aluno na tabela `resultados_simulados`
    cursor.execute("""
        SELECT r.simulado_id, a.nome AS assunto
        FROM resultados_simulados r
        JOIN simulados_gerados s ON r.simulado_id = s.id
        JOIN assuntos a ON s.assunto_id = a.id
        WHERE r.aluno_id = ?
        ORDER BY r.data_envio DESC
    """, (aluno_id,))
    simulados = cursor.fetchall()

    print(f"[DEBUG] Simulados respondidos para o aluno {aluno_id}: {simulados}")

    return render_template(
        'selecionar_relatorio.html',
        aluno_id=aluno_id,
        nome_aluno=nome_aluno,
        simulados=simulados,
        year=datetime.now().year  # Adicione o ano para o rodapé
    )

@app.route('/aluno/<int:aluno_id>/simulado/<int:simulado_id>/detalhes', methods=['GET'])
@login_required
def detalhar_simulado(aluno_id, simulado_id):
    # Verificar se o usuário é um aluno ou professor
    if current_user.tipo_usuario_id not in [3, 4]:  # Professores (3) ou Alunos (4)
        flash("Você não tem permissão para acessar esta página.", "error")
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Verificar se o aluno existe
    cursor.execute("""
        SELECT nome
        FROM usuarios
        WHERE id = ? AND tipo_usuario_id = 4
    """, (aluno_id,))
    aluno = cursor.fetchone()
    if not aluno:
        flash("Aluno não encontrado.", "error")
        return redirect(url_for("alunos_bp.portal_alunos"))

    nome_aluno = aluno[0]

    # Recuperar informações do simulado respondido
    cursor.execute("""
        SELECT r.simulado_id, r.respostas, r.respostas_certas, r.acertos, r.total_questoes, r.data_envio, 
               s.nivel, a.nome AS assunto, d.nome AS disciplina
        FROM resultados_simulados r
        JOIN simulados_gerados s ON r.simulado_id = s.id
        JOIN assuntos a ON s.assunto_id = a.id
        JOIN disciplinas d ON s.disciplina_id = d.id
        WHERE r.aluno_id = ? AND r.simulado_id = ?
    """, (aluno_id, simulado_id))
    resultado = cursor.fetchone()

    if not resultado:
        flash("Simulado não encontrado ou não respondido por este aluno.", "error")
        return redirect(url_for("alunos_bp.portal_alunos"))

    # Estruturar os dados para o template
    simulado_detalhes = {
        "simulado_id": resultado[0],
        "respostas": json.loads(resultado[1]),
        "respostas_certas": json.loads(resultado[2]),
        "acertos": resultado[3],
        "total_questoes": resultado[4],
        "data_envio": resultado[5],
        "nivel": resultado[6],
        "assunto": resultado[7],
        "disciplina": resultado[8]
    }

    # Recuperar perguntas do simulado
    cursor.execute("""
        SELECT id, texto_pergunta, alternativa_a, alternativa_b, alternativa_c, alternativa_d
        FROM perguntas_simulado
        WHERE simulado_id = ?
    """, (simulado_id,))
    perguntas = cursor.fetchall()

    perguntas_formatadas = [
        {
            "id": str(pergunta[0]),
            "pergunta": pergunta[1],
            "opcoes": [pergunta[2], pergunta[3], pergunta[4], pergunta[5]],
            "resposta_certa": simulado_detalhes["respostas_certas"].get(str(pergunta[0])),
            "resposta_aluno": simulado_detalhes["respostas"].get(str(pergunta[0])),
        }
        for pergunta in perguntas
    ]

    return render_template(
        "detalhar_simulado.html",
        aluno_id=aluno_id,
        nome_aluno=nome_aluno,
        simulado=simulado_detalhes,
        perguntas=perguntas_formatadas
    )


# @app.route('/aluno/<int:aluno_id>/relatorio/unidade/<int:unidade>', methods=['GET'])
# @login_required
# def gerar_relatorio_por_unidade(aluno_id, unidade):
#     db = get_db()
#     cursor = db.cursor()

#     # Mapear as unidades para os intervalos de datas
#     intervalos_unidade = {
#         1: ("2024-02-01", "2024-04-30"),
#         2: ("2024-05-01", "2024-06-30"),
#         3: ("2024-07-01", "2024-09-30"),
#         4: ("2024-10-01", "2024-12-31"),
#     }

#     if unidade not in intervalos_unidade:
#         return "Unidade inválida.", 400

#     inicio, fim = intervalos_unidade[unidade]

#     # Recuperar informações do aluno
#     cursor.execute("""
#         SELECT nome FROM usuarios WHERE id = ? AND tipo_usuario_id = 4
#     """, (aluno_id,))
#     aluno = cursor.fetchone()

#     if not aluno:
#         return "Aluno não encontrado.", 404

#     nome_aluno = aluno[0]

#     # Recuperar dados de desempenho do aluno para a unidade
#     cursor.execute("""
#         SELECT d.nome AS disciplina, a.assunto, r.acertos, r.total_questoes
#         FROM resultados_simulados r
#         JOIN simulados s ON r.simulado_id = s.id
#         JOIN disciplinas d ON s.disciplina_id = d.id
#         JOIN assunto a ON s.assunto = a.id
#         WHERE r.aluno_id = ? AND r.data_envio BETWEEN ? AND ?
#     """, (aluno_id, inicio, fim))
#     resultados = cursor.fetchall()

#     if not resultados:
#         return f"Nenhum dado disponível para a Unidade {unidade}.", 404

#     # Consolidar desempenho por assunto
#     desempenho_por_assunto = {}
#     for disciplina, assunto, acertos, total_questoes in resultados:
#         if assunto not in desempenho_por_assunto:
#             desempenho_por_assunto[assunto] = {
#                 'disciplina': disciplina,
#                 'total_acertos': 0,
#                 'total_questoes': 0
#             }
#         desempenho_por_assunto[assunto]['total_acertos'] += acertos
#         desempenho_por_assunto[assunto]['total_questoes'] += total_questoes

#     # Calcular médias e identificar pontos fortes/fracos
#     pontos_fortes = []
#     sugestoes_melhoria = []
#     desempenho_tabela = []

#     for assunto, dados in desempenho_por_assunto.items():
#         media_porcentagem = round(
#             (dados['total_acertos'] / dados['total_questoes']) * 100, 2
#         )
#         desempenho_tabela.append({
#             'disciplina': dados['disciplina'],
#             'assunto': assunto,
#             'total': dados['total_questoes'],
#             'acertos': dados['total_acertos'],
#             'porcentagem': media_porcentagem
#         })
#         if media_porcentagem >= 80:
#             pontos_fortes.append(f"{dados['disciplina']} - {assunto}")
#         elif media_porcentagem < 50:
#             sugestoes_melhoria.append(f"{dados['disciplina']} - {assunto}")

#     # Gerar parecer com a IA
#     parecer_geral = "Nenhum dado disponível para análise."  # Valor padrão
#     try:
#         mensagem = f"""
#         O aluno {nome_aluno} obteve os seguintes resultados na Unidade {unidade}:
#         """
#         for registro in desempenho_tabela:
#             mensagem += f"""
#             Disciplina: {registro['disciplina']}, Assunto: {registro['assunto']}, 
#             Acertos: {registro['acertos']} de {registro['total']} ({registro['porcentagem']}%).
#             """
#         mensagem += """
#         Baseando-se nesses dados, forneça um parecer curto e objetivo sobre o desempenho do aluno,
#         destacando apenas os principais pontos fortes e uma ou duas sugestões de melhoria.
#         """

#         response = openai.ChatCompletion.create(
#             model="gpt-3.5-turbo",
#             messages=[
#                 {"role": "system", "content": "Você é um gerador de pareceres escolares humanizados."},
#                 {"role": "user", "content": mensagem}
#             ],
#             max_tokens=250,
#             temperature=0.7
#         )
#         parecer_geral = response['choices'][0]['message']['content']
#     except Exception as e:
#         parecer_geral = f"Erro ao gerar parecer: {e}"

#     # Gerar HTML para o PDF
#     professor_nome = current_user.nome
#     rendered_html = render_template(
#         'relatorio_pdf.html',
#         nome_aluno=nome_aluno,
#         parecer_geral=parecer_geral,
#         desempenho_tabela=desempenho_tabela,
#         pontos_fortes=pontos_fortes,
#         sugestoes_melhoria=sugestoes_melhoria,
#         professor_nome=professor_nome,
#         unidade=unidade
#     )

#     # Gerar o PDF usando WeasyPrint
#     pdf = HTML(string=rendered_html).write_pdf(stylesheets=[CSS(string="""
#         body { font-family: Arial, sans-serif; margin: 20px; }
#         h1, h2, h3 { color: #4CAF50; }
#         .table { width: 100%; border-collapse: collapse; }
#         .table th, .table td { border: 1px solid #ddd; padding: 8px; text-align: center; }
#         .table th { background-color: #4CAF50; color: white; }
#         .footer { text-align: center; margin-top: 20px; font-size: 0.8rem; color: #555; }
#     """)])

#     # Retornar o PDF como resposta para download
#     response = make_response(pdf)
#     response.headers['Content-Type'] = 'application/pdf'
#     response.headers['Content-Disposition'] = f'inline; filename=relatorio_unidade_{unidade}_{nome_aluno}.pdf'
#     return response


@app.route("/debug_usuarios")
def debug_usuarios():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id, nome, email, senha, tipo_usuario_id FROM usuarios;")
    usuarios = cursor.fetchall()
    return jsonify(usuarios)

@app.route("/atualizar_senhas", methods=["GET"])
def atualizar_senhas():
    db = get_db()
    cursor = db.cursor()

    # Selecionar todas as senhas em texto simples
    cursor.execute("SELECT id, senha FROM usuarios;")
    usuarios = cursor.fetchall()

    for usuario in usuarios:
        usuario_id, senha = usuario

        # Verificar se a senha já está em hash (hashes começam com 'pbkdf2:')
        if not senha.startswith("pbkdf2:sha256:"):
            senha_hash = generate_password_hash(senha, method="pbkdf2:sha256", salt_length=8)

            # Atualizar a senha no banco
            cursor.execute("UPDATE usuarios SET senha = ? WHERE id = ?", (senha_hash, usuario_id))
            db.commit()

    return "Todas as senhas foram convertidas para hash com sucesso!"


# Rota principal com links para relatórios
@app.route("/relatorio_individual/<int:aluno_id>", methods=["GET"])
@login_required
def relatorio_individual(aluno_id):
    if current_user.tipo_usuario_id != 3:
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Buscar informações do aluno
    cursor.execute("""
        SELECT id, nome, turma_id
        FROM usuarios
        WHERE id = ? AND tipo_usuario_id = 4
    """, (aluno_id,))
    aluno = cursor.fetchone()

    if not aluno:
        return "Aluno não encontrado.", 404

    aluno_id, aluno_nome, turma_id = aluno

    # Buscar resultados dos simulados do aluno
    cursor.execute("""
        SELECT 
            r.simulado_id, 
            d.nome AS disciplina_nome, 
            a.assunto AS assunto_nome, 
            r.acertos, 
            r.total_questoes, 
            r.data_envio
        FROM resultados_simulados r
        JOIN simulados s ON r.simulado_id = s.id
        JOIN disciplinas d ON s.disciplina_id = d.id
        JOIN assunto a ON s.assunto = a.id
        WHERE r.aluno_id = ?
        ORDER BY r.data_envio DESC
    """, (aluno_id,))
    resultados = cursor.fetchall()

    if not resultados:
        return render_template(
            "relatorio_individual.html",
            aluno=(aluno_id, aluno_nome, turma_id),
            parecer_geral="Nenhum resultado encontrado para este aluno.",
            resultados=[],
            pontos_fortes=[],
            sugestoes_melhoria=[],
            detalhamento_disciplina={},
            desempenho_tabela=[],
            grafico_labels=[],
            grafico_data=[]
        )

    # Processar os resultados para análise
    desempenho_por_assunto = {}
    desempenho_tabela = []
    grafico_labels = []
    grafico_data = []

    for r in resultados:
        disciplina = r[1]
        assunto = r[2]
        acertos = r[3]
        total = r[4]
        porcentagem = (acertos / total) * 100

        # Atualizar o desempenho por assunto
        if disciplina not in desempenho_por_assunto:
            desempenho_por_assunto[disciplina] = {}
        if assunto not in desempenho_por_assunto[disciplina]:
            desempenho_por_assunto[disciplina][assunto] = {"acertos": 0, "total": 0, "porcentagem": 0}

        desempenho_por_assunto[disciplina][assunto]["acertos"] += acertos
        desempenho_por_assunto[disciplina][assunto]["total"] += total

    # Calcular médias e preparar detalhamento
    for disciplina, assuntos in desempenho_por_assunto.items():
        for assunto, dados in assuntos.items():
            dados["porcentagem"] = round((dados["acertos"] / dados["total"]) * 100, 2)
            dados["comentario"] = (
                "Excelente desempenho. Continue assim!" if dados["porcentagem"] >= 80 else
                "Bom desempenho, mas há espaço para melhorias." if dados["porcentagem"] >= 50 else
                "Recomenda-se foco adicional neste assunto para melhorar o desempenho."
            )
            desempenho_tabela.append({
                "disciplina": disciplina,
                "assunto": assunto,
                "total": dados["total"],
                "acertos": dados["acertos"],
                "porcentagem": dados["porcentagem"]
            })

    # Preparar dados do gráfico
    for disciplina, assuntos in desempenho_por_assunto.items():
        total_acertos = sum(a["acertos"] for a in assuntos.values())
        total_questoes = sum(a["total"] for a in assuntos.values())
        grafico_labels.append(disciplina)
        grafico_data.append(round((total_acertos / total_questoes) * 100, 2))

    # Gerar parecer geral com IA
    try:
        import openai
        mensagem_ia = f"""
        Aluno: {aluno_nome}
        Resultados por Disciplina e Assunto:
        """
        for disciplina, assuntos in desempenho_por_assunto.items():
            for assunto, dados in assuntos.items():
                mensagem_ia += f"""
                - Disciplina: {disciplina}, Assunto: {assunto}, Acertos: {dados['acertos']} de {dados['total']} ({dados['porcentagem']}%)
                """
        mensagem_ia += """
        Com base nos dados acima, forneça um parecer geral do aluno, destacando pontos fortes e sugestões de melhoria.
        """

        resposta = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Você é um especialista em educação."},
                {"role": "user", "content": mensagem_ia}
            ],
            max_tokens=200,
            temperature=0.7
        )
        parecer_geral = resposta['choices'][0]['message']['content']
    except Exception as e:
        parecer_geral = f"Erro ao gerar parecer geral com a IA: {e}"

    # Renderizar o template
    return render_template(
        "relatorio_individual.html",
        aluno=(aluno_id, aluno_nome, turma_id),
        resultados=resultados,
        parecer_geral=parecer_geral,
        pontos_fortes=[
            f"{assunto} ({disciplina})"
            for disciplina, assuntos in desempenho_por_assunto.items()
            for assunto, dados in assuntos.items()
            if dados["porcentagem"] >= 80
        ],
        sugestoes_melhoria=[
            f"{assunto} ({disciplina})"
            for disciplina, assuntos in desempenho_por_assunto.items()
            for assunto, dados in assuntos.items()
            if dados["porcentagem"] < 50
        ],
        detalhamento_disciplina=desempenho_por_assunto,
        desempenho_tabela=desempenho_tabela,
        grafico_labels=grafico_labels,
        grafico_data=grafico_data
    )

      
@app.route("/listar_alunos")
@login_required
def listar_alunos():
    if current_user.tipo_usuario_id != 3:
        return redirect(url_for("portal_professores"))

    db = get_db()
    cursor = db.cursor()

    # Recuperar alunos vinculados às turmas do professor na tabela professor_turma_escola
    cursor.execute("""
        SELECT u.id, u.nome
        FROM usuarios u
        JOIN professor_turma_escola pte ON u.turma_id = pte.turma_id
        WHERE pte.professor_id = ?
    """, (current_user.id,))
    alunos = cursor.fetchall()

    return render_template("listar_alunos.html", alunos=alunos)



@app.route('/cadastrar_simulado', methods=['GET', 'POST'])
@login_required
def cadastrar_simulado():
    if current_user.tipo_usuario_id != 3:  # Apenas professores podem acessar
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Buscar assuntos cadastrados pelo professor
    cursor.execute("""
        SELECT a.id, a.nome, d.nome AS disciplina
        FROM assunto a
        JOIN disciplinas d ON a.disciplina_id = d.id
        WHERE a.professor_id = ?
    """, (current_user.id,))
    assuntos = cursor.fetchall()

    if request.method == 'POST':
        assunto_id = request.form.get('assunto_id')

        # Validar campos obrigatórios
        if not assunto_id:
            flash("Por favor, selecione um assunto.", "error")
            return redirect(url_for("cadastrar_simulado"))

        # Obter informações do assunto
        cursor.execute("""
            SELECT a.nome, d.nome
            FROM assunto a
            JOIN disciplinas d ON a.disciplina_id = d.id
            WHERE a.id = ?
        """, (assunto_id,))
        assunto_info = cursor.fetchone()

        if not assunto_info:
            flash("Assunto inválido.", "error")
            return redirect(url_for("cadastrar_simulado"))

        assunto_nome, disciplina_nome = assunto_info

        try:
            # Gerar o simulado com a IA
            simulado = gerar_perguntas(disciplina_nome, quantidade=5)
            print(f"Simulado gerado: {simulado}")

            # Salvar o simulado no banco de dados
            respostas_certas = json.dumps([ord(pergunta["resposta_correta"]) - ord('A') + 1 for pergunta in simulado])
            cursor.execute("""
                INSERT INTO simulados (assunto_id, professor_id, respostas_certas)
                VALUES (?, ?, ?)
            """, (assunto_id, current_user.id, respostas_certas))
            simulado_id = cursor.lastrowid
            db.commit()

            flash("Simulado gerado e salvo com sucesso!", "success")
            return render_template("visualizar_simulado.html", simulado=simulado)

        except Exception as e:
            print(f"Erro ao gerar simulado: {e}")
            flash("Erro ao gerar o simulado. Tente novamente mais tarde.", "error")

    return render_template(
        "cadastrar_simulado.html",
        assuntos=assuntos,
    )



import os
import csv
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/importar_assuntos', methods=['GET', 'POST'])
@login_required
def importar_assuntos():
    if current_user.tipo_usuario_id != 1:
        flash("Acesso negado!", "danger")
        return redirect(url_for('portal_administrador'))

    if request.method == 'POST':
        if 'file' not in request.files:
            flash("Nenhum arquivo foi enviado.", "danger")
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '':
            flash("Nenhum arquivo selecionado.", "danger")
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            try:
                if not os.path.exists(app.config['UPLOAD_FOLDER']):
                    os.makedirs(app.config['UPLOAD_FOLDER'])
                
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)

                # Processar o CSV
                with open(filepath, newline='', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    db = get_db()
                    cursor = db.cursor()
                    registros_importados = 0

                    for row in reader:
                        print(f"Processando linha: {row}")  # Debug
                        try:
                            cursor.execute("""
                                INSERT INTO assunto (disciplina, assunto, serie_id, turma_id, escola_id, professor_id)
                                VALUES (?, ?, ?, ?, ?, ?)
                            """, (row['disciplina'], row['assunto'], row['serie_id'], None, None, None))
                            registros_importados += 1
                        except Exception as e:
                            print(f"Erro ao inserir linha: {e}")
                            db.rollback()
                            raise

                    db.commit()
                    flash(f"Importação concluída com sucesso! {registros_importados} registros adicionados.", "success")
                    return redirect(url_for('portal_administrador'))

            except Exception as e:
                print(f"Erro ao processar o arquivo: {e}")
                flash(f"Erro ao processar o arquivo: {e}", "danger")
            finally:
                if os.path.exists(filepath):
                    os.remove(filepath)
        else:
            flash("Formato de arquivo inválido. Envie um arquivo CSV.", "danger")

    return render_template('importar_assuntos.html')


@app.route("/corrigir_simulado/<int:simulado_id>", methods=["GET", "POST"])
@login_required
def corrigir_simulado(simulado_id):
    if current_user.tipo_usuario_id != 3:
        return redirect(url_for("portal_professores"))

    db = get_db()
    cursor = db.cursor()

    # Buscar informações do simulado
    cursor.execute("""
        SELECT id, assunto, turma_id, disciplina_id
        FROM simulados
        WHERE id = ?
    """, (simulado_id,))
    simulado = cursor.fetchone()

    if not simulado:
        return "Simulado não encontrado.", 404

    print(f"Simulado encontrado: {simulado}")

    # Buscar respostas dos alunos no banco de resultados
    cursor.execute("""
        SELECT u.nome AS aluno_nome, r.respostas, r.acertos, r.data_envio
        FROM resultados_simulados r
        JOIN usuarios u ON r.aluno_id = u.id
        WHERE r.simulado_id = ?
    """, (simulado_id,))
    respostas = cursor.fetchall()

    if not respostas:
        print(f"Nenhuma resposta encontrada para o simulado ID {simulado_id}.")
        return "Nenhuma resposta foi encontrada para este simulado.", 200

    return render_template(
        "corrigir_simulado.html",
        simulado=simulado,
        respostas=respostas
    )


import matplotlib.pyplot as plt
from io import BytesIO
import base64

@app.route("/relatorio_turma", methods=["GET"])
@login_required
def relatorio_turma():
    if current_user.tipo_usuario_id != 3:
        return redirect(url_for("portal_professores"))

    db = get_db()
    cursor = db.cursor()

    # Buscar as turmas vinculadas ao professor logado
    cursor.execute("""
        SELECT t.id, se.nome AS serie, t.turma
        FROM turmas t
        JOIN professor_turma_escola pte ON t.id = pte.turma_id
        JOIN series se ON t.serie_id = se.id
        WHERE pte.professor_id = ?
    """, (current_user.id,))
    turmas = cursor.fetchall()

    if not turmas:
        return "Nenhuma turma vinculada ao professor.", 404

    turma = turmas[0]  # Considerar a primeira turma vinculada
    turma_id = turma[0]

    # Buscar desempenho da turma
    cursor.execute("""
        SELECT 
            u.nome AS aluno_nome,
            d.nome AS disciplina_nome,
            a.nome AS assunto_nome,
            dp.acertos,
            dp.total_questoes,
            dp.porcentagem
        FROM desempenho dp
        JOIN usuarios u ON dp.aluno_id = u.id
        JOIN disciplinas d ON dp.disciplina_id = d.id
        JOIN assuntos a ON dp.assunto_id = a.id
        WHERE dp.turma_id = ?
    """, (turma_id,))
    desempenho = cursor.fetchall()

    # Gerar gráfico com matplotlib
    disciplinas = [record[1] for record in desempenho]
    porcentagens = [record[5] for record in desempenho]

    plt.figure(figsize=(10, 6))
    plt.bar(disciplinas, porcentagens, color="skyblue")
    plt.xlabel("Disciplinas")
    plt.ylabel("Porcentagem de Acertos")
    plt.title("Desempenho por Disciplina")

    # Salvar o gráfico em memória como imagem
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    buffer.close()

    # Gerar HTML para o PDF
    rendered_html = render_template(
        "relatorio_turma.html",
        turma={"serie": turma[1], "turma": turma[2]},
        desempenho=desempenho,
        parecer=None,
        error=None,
        grafico=image_base64
    )

    # Gerar o PDF usando WeasyPrint
    pdf = HTML(string=rendered_html).write_pdf(stylesheets=[CSS(string="""
        body { font-family: Arial, sans-serif; margin: 20px; }
        h1, h2, h3 { color: #4CAF50; }
        .table { width: 100%; border-collapse: collapse; }
        .table th, .table td { border: 1px solid #ddd; padding: 8px; text-align: center; }
        .table th { background-color: #4CAF50; color: white; }
        .footer { text-align: center; margin-top: 20px; font-size: 0.8rem; color: #555; }
    """)])

    # Retornar o PDF como resposta para download
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=relatorio_turma_{turma[1]}_{turma[2]}.pdf'
    return response


@app.route("/cadastrar_disciplina", methods=["GET", "POST"])
@login_required
def cadastrar_disciplina():
    # Verifica se o usuário é Administrador
    if current_user.tipo_usuario_id != 1:
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Lógica para POST
    if request.method == "POST":
        nome = request.form.get("nome")

        # Validação do campo "nome"
        if not nome.strip():
            return render_template(
                "cadastrar_disciplina.html",
                error="O nome da disciplina é obrigatório!"
            )

        try:
            # Inserir no banco de dados
            cursor.execute("""
                INSERT INTO disciplinas (nome)
                VALUES (?)
            """, (nome.strip(),))
            db.commit()
            return render_template(
                "cadastrar_disciplina.html",
                success="Disciplina cadastrada com sucesso!"
            )
        except sqlite3.IntegrityError:
            # Trata duplicidade no banco
            return render_template(
                "cadastrar_disciplina.html",
                error="Essa disciplina já está cadastrada!"
            )

    # Renderiza o template para método GET
    return render_template("cadastrar_disciplina.html")


@app.route("/portal_administrador/cadastrar_turma", methods=["GET", "POST"])
@login_required
def cadastrar_turma():
    if current_user.tipo_usuario_id != 1:  # Apenas administradores podem cadastrar
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Buscar escolas cadastradas para o menu suspenso
    cursor.execute("SELECT id, nome FROM escolas")
    escolas = cursor.fetchall()

    if request.method == "POST":
        # Obter os valores do formulário
        escola_id = request.form.get("escola_id", "").strip()
        tipo_ensino_id = request.form.get("tipo_ensino_id", "").strip()
        serie_id = request.form.get("serie_id", "").strip()
        turma_nome = request.form.get("turma", "").strip()

        # Printar os valores recebidos para depuração
        print(f"escola_id: {escola_id}")
        print(f"tipo_ensino_id: {tipo_ensino_id}")
        print(f"serie_id: {serie_id}")
        print(f"turma_nome: {turma_nome}")

        # Validar campos obrigatórios
        if not all([escola_id, tipo_ensino_id, serie_id, turma_nome]):
            print("Erro: Campos obrigatórios não preenchidos!")
            return render_template(
                "cadastrar_turma.html",
                escolas=escolas,
                error="Todos os campos são obrigatórios!"
            )

        try:
            # Inserir a turma no banco de dados
            cursor.execute(
                """
                INSERT INTO turmas (escola_id, tipo_ensino_id, serie_id, turma)
                VALUES (?, ?, ?, ?)
                """,
                (int(escola_id), int(tipo_ensino_id), int(serie_id), turma_nome)
            )
            db.commit()
            print("Turma cadastrada com sucesso!")
            return render_template(
                "cadastrar_turma.html",
                escolas=escolas,
                success="Turma cadastrada com sucesso!"
            )
        except sqlite3.IntegrityError:
            print("Erro: Turma já cadastrada!")
            return render_template(
                "cadastrar_turma.html",
                escolas=escolas,
                error="Essa turma já está cadastrada!"
            )
        except Exception as e:
            print(f"Erro ao cadastrar turma: {e}")
            return render_template(
                "cadastrar_turma.html",
                escolas=escolas,
                error="Erro ao cadastrar turma. Por favor, tente novamente!"
            )

    return render_template("cadastrar_turma.html", escolas=escolas)


@app.route("/relatorio_aluno/<int:aluno_id>", methods=["GET"])
@login_required
def relatorio_aluno(aluno_id):
    if current_user.tipo_usuario_id != [3, 1]:
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Buscar informações do aluno
    cursor.execute("SELECT nome FROM alunos WHERE id = ?", (aluno_id,))
    aluno = cursor.fetchone()

    if not aluno:
        return "Aluno não encontrado.", 404

    # Buscar desempenho nos simulados
    cursor.execute("""
        SELECT s.assunto, COUNT(r.id) AS total_perguntas, SUM(r.correta) AS total_acertos
        FROM respostas r
        JOIN simulados s ON r.simulado_id = s.id
        WHERE r.aluno_id = ?
        GROUP BY s.assunto
    """, (aluno_id,))
    desempenho = cursor.fetchall()

    return render_template("relatorio_aluno.html", aluno=aluno[0], desempenho=desempenho)


@app.route("/cadastrar_escola", methods=["GET", "POST"])
@login_required
def cadastrar_escola():
    if current_user.tipo_usuario_id != 1:  # Apenas administradores podem acessar
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Buscar tipos de ensino do banco de dados
    cursor.execute("SELECT id, nome FROM tipos_ensino")
    tipos_ensino = cursor.fetchall()  # Retorna uma lista de tuplas [(1, "Ensino Infantil"), ...]

    if request.method == "POST":
        nome = request.form.get("nome")
        cep = request.form.get("cep")
        estado = request.form.get("estado")
        cidade = request.form.get("cidade")
        bairro = request.form.get("bairro")
        endereco = request.form.get("endereco")
        numero = request.form.get("numero")
        telefone = request.form.get("telefone")
        cnpj = request.form.get("cnpj")
        diretor = request.form.get("diretor")
        codigo_ibge = request.form.get("codigo_ibge")
        tipos_ensino_selecionados = request.form.getlist("tipos_ensino[]")  # Lista dos IDs dos tipos de ensino selecionados

        # Validar campos obrigatórios
        if not all([nome, cep, estado, cidade, bairro, endereco, numero, telefone, cnpj, diretor, codigo_ibge]):
            flash("Todos os campos são obrigatórios!", "danger")
            return render_template(
                "cadastrar_escola.html",
                tipos_ensino=tipos_ensino
            )

        try:
            # Inserir escola no banco de dados
            cursor.execute("""
                INSERT INTO escolas (
                    nome, cep, estado, cidade, bairro, endereco, numero, telefone, cnpj, diretor, codigo_ibge
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (nome, cep, estado, cidade, bairro, endereco, numero, telefone, cnpj, diretor, codigo_ibge))
            escola_id = cursor.lastrowid  # Pega o ID da escola recém-criada

            # Associar tipos de ensino à escola
            for tipo_id in tipos_ensino_selecionados:
                cursor.execute("""
                    INSERT INTO escola_tipos_ensino (escola_id, tipo_ensino_id)
                    VALUES (?, ?)
                """, (escola_id, tipo_id))

            db.commit()
            flash("Escola cadastrada com sucesso!", "success")
            return redirect(url_for("portal_administrador"))

        except Exception as e:
            db.rollback()
            flash(f"Erro ao cadastrar escola: {e}", "danger")
            return render_template(
                "cadastrar_escola.html",
                tipos_ensino=tipos_ensino
            )

    return render_template("cadastrar_escola.html", tipos_ensino=tipos_ensino)


@app.route("/get_tipo_ensino", methods=["GET"])
@login_required
def get_tipo_ensino():
    escola_id = request.args.get("escola_id")
    if not escola_id:
        return jsonify([])

    db = get_db()
    cursor = db.cursor()

    # Buscar os tipos de ensino associados à escola
    cursor.execute(
        """
        SELECT te.id, te.nome
        FROM escola_tipos_ensino ete
        JOIN tipos_ensino te ON ete.tipo_ensino_id = te.id
        WHERE ete.escola_id = ?
        """,
        (escola_id,)
    )
    tipos_ensino = cursor.fetchall()

    if not tipos_ensino:
        app.logger.info(f"Nenhum tipo de ensino encontrado para a escola {escola_id}.")
        return jsonify([])

    app.logger.info(f"Tipos de ensino encontrados: {get_tipos_ensino}")
    return jsonify([{"id": tipo[0], "nome": tipo[1]} for tipo in tipos_ensino])


@app.route("/get_series", methods=["GET"])
@login_required
def get_series():
    tipo_ensino_id = request.args.get("tipo_ensino_id")
    codigo_ibge = current_user.codigo_ibge

    if not tipo_ensino_id:
        return jsonify([])

    db = get_db()
    cursor = db.cursor()

    cursor.execute(
        """
        SELECT DISTINCT series.id, series.nome
        FROM series
        JOIN escolas ON series.tipo_ensino_id = escolas.tipo_ensino_id
        WHERE escolas.codigo_ibge = ?
        """,
        (codigo_ibge,),
    )
    series = cursor.fetchall()

    return jsonify([{"id": s[0], "nome": s[1]} for s in series])

import logging

@app.route("/get_turmas", methods=["GET"])
@login_required
def get_turmas():
    escola_id = request.args.get("escola_id")
    tipo_ensino_id = request.args.get("tipo_ensino_id")
    serie_id = request.args.get("serie_id")

    app.logger.info(f"Parâmetros recebidos: escola_id={escola_id}, tipo_ensino_id={tipo_ensino_id}, serie_id={serie_id}")

    if not all([escola_id, tipo_ensino_id, serie_id]):
        app.logger.warning("Parâmetros insuficientes fornecidos para /get_turmas.")
        return jsonify([])

    db = get_db()
    cursor = db.cursor()

    try:
        # Consulta para buscar turmas e nomes de série
        cursor.execute(
            """
            SELECT turmas.id, series.nome AS serie, turmas.turma
            FROM turmas
            JOIN series ON turmas.serie_id = series.id
            WHERE turmas.escola_id = ? AND turmas.tipo_ensino_id = ? AND turmas.serie_id = ?
            """,
            (escola_id, tipo_ensino_id, serie_id),
        )
        turmas = cursor.fetchall()

        app.logger.info(f"Turmas encontradas: {turmas}")

        # Retorna série e turma separadas
        return jsonify([{"id": turma[0], "serie": turma[1], "turma": turma[2]} for turma in turmas])
    except Exception as e:
        app.logger.error(f"Erro ao buscar turmas: {e}")
        return jsonify([])



@app.route('/buscar_alunos', methods=['GET'])
@login_required
def buscar_alunos():
    turma_id = request.args.get("turma_id")
    
    if not turma_id:
        return jsonify([])  # Caso não exista turma_id, retorna vazio

    db = get_db()
    cursor = db.cursor()

    # Buscar os alunos vinculados à turma
    cursor.execute("""
        SELECT id, nome 
        FROM usuarios
        WHERE turma_id = ? AND tipo_usuario_id = 4
    """, (turma_id,))
    alunos = cursor.fetchall()

    # Formatar resposta JSON
    return jsonify([{"id": aluno[0], "nome": aluno[1]} for aluno in alunos])

@app.route("/autocomplete_alunos", methods=["GET"])
@login_required
def autocomplete_alunos():
    # Verifica se o usuário tem permissão
    if current_user.tipo_usuario_id not in ["Secretaria Acadêmica"]:
        return jsonify([])

    query = request.args.get("query", "").strip()
    if len(query) < 2:
        return jsonify([])

    db = get_db()
    cursor = db.cursor()

    # Buscar nomes de alunos que correspondem à query
    cursor.execute("""
        SELECT nome 
        FROM usuarios 
        WHERE tipo_usuario_id = 4 AND nome LIKE ?
    """, (f"%{query}%",))
    alunos = [row[0] for row in cursor.fetchall()]

    return jsonify(alunos)


@app.route("/cadastrar_aluno", methods=["GET", "POST"])
@login_required
def cadastrar_aluno():
    if current_user.tipo_usuario_id not in ["Secretaria Acadêmica", "Administrador"]:
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Buscar turmas para o menu suspenso
    cursor.execute("""
        SELECT t.id, e.nome AS escola, t.tipo_ensino, t.serie_id, t.turma_id
        FROM turmas t
        JOIN escolas e ON t.escola_id = e.id
    """)
    turmas = cursor.fetchall()

    if request.method == "POST":
        nome = request.form.get("nome")
        data_nascimento = request.form.get("data_nascimento")
        turma_id = request.form.get("turma_id")

        # Verificar se todos os campos foram preenchidos
        if not all([nome, data_nascimento, turma_id]):
            return render_template(
                "cadastrar_aluno.html",
                turmas=turmas,
                error="Todos os campos são obrigatórios!"
            )

        # Buscar informações adicionais da turma
        cursor.execute("""
            SELECT e.id AS escola_id, t.tipo_ensino, t.serie_id, t.turma_id
            FROM turmas t
            JOIN escolas e ON t.escola_id = e.id
            WHERE t.id = ?
        """, (turma_id,))
        turma_info = cursor.fetchone()

        if not turma_info:
            return render_template(
                "cadastrar_aluno.html",
                turmas=turmas,
                error="A turma selecionada é inválida."
            )

        escola_id, tipo_ensino, serie_id, turma_letra = turma_info

        # Inserir aluno na tabela de alunos
        try:
            cursor.execute("""
                INSERT INTO alunos (nome, data_nascimento, escola_id, tipo_ensino, serie_id, turma_id)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (nome, data_nascimento, escola_id, tipo_ensino, serie_id, turma_letra))
            db.commit()
            return redirect(url_for("portal_secretaria_academica"))
        except sqlite3.IntegrityError as e:
            print(f"Erro ao cadastrar aluno: {e}")
            return render_template(
                "cadastrar_aluno.html",
                turmas=turmas,
                error="Erro ao cadastrar o aluno. Verifique os dados e tente novamente."
            )

    return render_template("cadastrar_aluno.html", turmas=turmas)



@app.route("/relatorios", methods=["GET"])
@login_required
def relatorios():
    # Permitir acesso apenas para o administrador
    if current_user.tipo_usuario_id != 1:
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Recuperar as escolas para o filtro
    cursor.execute("SELECT id, nome FROM escolas")
    escolas = cursor.fetchall()

    # Filtro de escola
    escola_id = request.args.get("escola_id")
    params = []
    filter_query = ""
    if escola_id:
        filter_query = "WHERE e.id = ?"
        params.append(escola_id)

    # Query para o desempenho geral e alunos de destaque
    cursor.execute(f"""
        SELECT 
            u.nome AS aluno,
            e.nome AS escola,
            d.nome AS disciplina,
            a.assunto,
            r.pontuacao,
            r.data_participacao
        FROM ranking r
        JOIN usuarios u ON r.aluno_id = u.id
        JOIN escolas e ON u.escola_id = e.id
        JOIN disciplinas d ON r.disciplina_id = d.id
        JOIN assunto a ON r.assunto_id = a.id
        {filter_query}
        ORDER BY r.pontuacao DESC, r.data_participacao ASC
        LIMIT 10
    """, params)
    alunos_destaque = cursor.fetchall()

    return render_template(
        "relatorios.html",
        escolas=escolas,
        alunos_destaque=alunos_destaque,
        selected_escola=escola_id,
    )


@app.route("/criar_administrador", methods=["GET", "POST"])
def criar_administrador():
    if request.method == "POST":
        nome = request.form.get("nome")
        email = request.form.get("email")
        senha = request.form.get("senha")

        db = get_db()
        cursor = db.cursor()
        cursor.execute("""
            INSERT INTO usuarios (nome, email, senha, tipo_usuario_id, escola_id)
            VALUES (?, ?, ?, ?, NULL)
        """, (nome, email, senha, "Administrador"))
        db.commit()

        return "Administrador criado com sucesso!"

    return render_template("criar_administrador.html")

@app.route("/cadastrar_professor", methods=["GET", "POST"])
@login_required
def cadastrar_professor():
    if current_user.tipo_usuario_id not in ["Secretaria Acadêmica", "Administrador"]:
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Buscar turmas para o menu suspenso
    cursor.execute("""
        SELECT t.id, e.nome AS escola, t.tipo_ensino, t.serie_id, t.turma_id
        FROM turmas t
        JOIN escolas e ON t.escola_id = e.id
    """)
    turmas = cursor.fetchall()

    if request.method == "POST":
        nome = request.form.get("nome")
        data_nascimento = request.form.get("data_nascimento")
        turma_id = request.form.get("turma_id")

        # Verificar se todos os campos foram preenchidos
        if not all([nome, data_nascimento, turma_id]):
            return render_template(
                "cadastrar_professor.html",
                turmas=turmas,
                error="Todos os campos são obrigatórios!"
            )

        # Buscar informações adicionais da turma
        cursor.execute("""
            SELECT e.id AS escola_id, t.tipo_ensino, t.serie_id, t.turma_id
            FROM turmas t
            JOIN escolas e ON t.escola_id = e.id
            WHERE t.id = ?
        """, (turma_id,))
        turma_info = cursor.fetchone()

        if not turma_info:
            return render_template(
                "cadastrar_professor.html",
                turmas=turmas,
                error="A turma selecionada é inválida."
            )

        escola_id, tipo_ensino, serie_id, turma_letra = turma_info

        # Inserir professor na tabela de professores
        try:
            cursor.execute("""
                INSERT INTO professores (nome, data_nascimento, escola_id, tipo_ensino, serie_id, turma_id)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (nome, data_nascimento, escola_id, tipo_ensino, serie_id, turma_letra))
            db.commit()
            return redirect(url_for("portal_secretaria_academica"))
        except sqlite3.IntegrityError as e:
            print(f"Erro ao cadastrar professor: {e}")
            return render_template(
                "cadastrar_professor.html",
                turmas=turmas,
                error="Erro ao cadastrar o professor. Verifique os dados e tente novamente."
            )

    return render_template("cadastrar_professor.html", turmas=turmas)


@app.route("/buscar_professores", methods=["GET"])
@login_required
def buscar_professores():
    nome = request.args.get("nome", "").strip()
    if len(nome) < 2:
        return jsonify([])

    db = get_db()
    cursor = db.cursor()

    # Buscar nome e data de nascimento dos professores
    cursor.execute("""
        SELECT nome, data_nascimento
        FROM usuarios
        WHERE tipo_usuario_id = 3' AND nome LIKE ?
    """, (f"%{nome}%",))
    professores = cursor.fetchall()

    # Retornar os dados no formato JSON
    return jsonify([
        {"nome": professor[0], "data_nascimento": professor[1] or ""}
        for professor in professores
    ])


@app.route('/buscar_tipo_ensino', methods=["GET"])
@login_required
def buscar_tipo_ensino():
    escola_id = request.args.get('escola_id')
    db = get_db()
    cursor = db.cursor()

    # Busca os tipos de ensino da escola pelo ID
    cursor.execute("SELECT tipo_ensino FROM escolas WHERE id = ?", (escola_id,))
    result = cursor.fetchone()
    
    if result:
        tipo_ensino = result[0].split(",")  # Divide os tipos de ensino separados por vírgula
        return jsonify([tipo.strip() for tipo in tipo_ensino])  # Remove espaços extras
    return jsonify([])


@app.route("/cadastrar_usuario", methods=["GET", "POST"])
@login_required
def cadastrar_usuario():
    print("Entrou na rota de cadastro de usuário!")  # Log para depuração
    
    if current_user.tipo_usuario_id != 1:  # Apenas administradores podem cadastrar
        print("Usuário não é administrador, redirecionando...")
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    # Buscar escolas disponíveis para exibir no formulário
    cursor.execute("SELECT id, nome FROM escolas")
    escolas = cursor.fetchall()

    if request.method == "POST":
        print("Formulário de cadastro recebido!")

        # Capturando os dados do formulário
        nome = request.form.get("nome")
        email = request.form.get("email")
        senha = request.form.get("senha")
        tipo_usuario_id = request.form.get("tipo_usuario_id")
        codigo_ibge = request.form.get("codigo_ibge")  # Código IBGE da cidade
        cep = request.form.get("cep")

        # Capturando múltiplas turmas para professores
        escolas_ids = request.form.getlist("escola_id[]")
        tipos_ensino_ids = request.form.getlist("tipo_ensino[]")
        series_ids = request.form.getlist("serie[]")
        turmas_ids = request.form.getlist("turma_id[]")

        print(f"Recebidos: Nome={nome}, Email={email}, Tipo={tipo_usuario_id}, Turmas={turmas_ids}")

        

        # Validar campos obrigatórios
        if not all([nome, email, senha, tipo_usuario_id, codigo_ibge, cep]):
            print("Erro: Campos obrigatórios não preenchidos.")
            flash("Todos os campos são obrigatórios!", "error")
            return render_template("cadastrar_usuario.html", escolas=escolas)

        # Validação específica para professores e alunos
        if tipo_usuario_id == "4":  # Aluno
            if not all([escolas_ids[0], turmas_ids[0], tipos_ensino_ids[0], series_ids[0]]):
                flash("Escola, Tipo de Ensino, Série e Turma são obrigatórios para Alunos!", "error")
                return render_template("cadastrar_usuario.html", escolas=escolas)

        if tipo_usuario_id == "3":  # Professor
            if not all([escolas_ids, turmas_ids, tipos_ensino_ids, series_ids]):
                flash("Escola, Tipo de Ensino, Série e Turma são obrigatórios para Professores!", "error")
                return render_template("cadastrar_usuario.html", escolas=escolas)

        try:
            # Inserindo o usuário na tabela `usuarios` apenas se não for professor
            senha_hash = generate_password_hash(senha)

            if tipo_usuario_id != "3":  # Não é professor
                cursor.execute(
                    """
                    INSERT INTO usuarios (nome, email, senha, tipo_usuario_id, escola_id, turma_id, tipo_ensino_id, serie_id, codigo_ibge, cep)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        nome,
                        email,
                        senha_hash,
                        tipo_usuario_id,
                        escolas_ids[0],
                        turmas_ids[0],
                        tipos_ensino_ids[0],
                        series_ids[0],
                        codigo_ibge,
                        cep,
                    ),
                )
                db.commit()
                print(f"Usuário {nome} cadastrado na tabela `usuarios` com sucesso!")
                flash("Usuário cadastrado com sucesso!", "success")
            else:  # Professor
                # Insere o professor na tabela `usuarios` com dados básicos
                cursor.execute(
                    """
                    INSERT INTO usuarios (nome, email, senha, tipo_usuario_id, codigo_ibge, cep)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        nome,
                        email,
                        senha_hash,
                        tipo_usuario_id,
                        codigo_ibge,
                        cep,
                    ),
                )
                usuario_id = cursor.lastrowid
                print(f"Professor {nome} cadastrado na tabela `usuarios` com ID {usuario_id}")

                # Insere as turmas do professor na tabela `professor_turma_escola`
                for escola_id, tipo_ensino_id, serie_id, turma_id in zip(escolas_ids, tipos_ensino_ids, series_ids, turmas_ids):
                    cursor.execute(
                        """
                        INSERT INTO professor_turma_escola (professor_id, escola_id, turma_id, tipo_ensino_id, serie_id)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (
                            usuario_id,
                            escola_id,
                            turma_id,
                            tipo_ensino_id,
                            serie_id,
                        ),
                    )
                db.commit()
                print(f"Turmas do professor {nome} cadastradas com sucesso!")
                flash("Professor cadastrado com sucesso com todas as turmas!", "success")

            return redirect(url_for("portal_administrador"))

        except Exception as e:
            db.rollback()
            print(f"Erro ao cadastrar usuário: {e}")
            flash(f"Erro ao cadastrar usuário: {e}", "error")

    return render_template("cadastrar_usuario.html", escolas=escolas)

@app.route('/alterar-senha', methods=['GET', 'POST'])
@login_required
def alterar_senha():
    error = None
    success = None

    if request.method == 'POST':
        senha_atual = request.form.get('senha_atual')
        nova_senha = request.form.get('nova_senha')
        confirmar_senha = request.form.get('confirmar_senha')

        if nova_senha != confirmar_senha:
            error = "As novas senhas não coincidem. Por favor, tente novamente."
        else:
            db = get_db()
            cursor = db.cursor()

            # Obter a senha atual do banco de dados
            cursor.execute("SELECT senha FROM usuarios WHERE id = ?", (current_user.id,))
            senha_atual_hash = cursor.fetchone()[0]

            # Validar a senha atual
            if not check_password_hash(senha_atual_hash, senha_atual):
                error = "A senha atual está incorreta. Por favor, tente novamente."
            else:
                # Gerar hash da nova senha
                nova_senha_hash = generate_password_hash(nova_senha)

                # Atualizar a senha no banco de dados
                cursor.execute("UPDATE usuarios SET senha = ? WHERE id = ?", (nova_senha_hash, current_user.id))
                db.commit()
                success = "Senha alterada com sucesso!"

    return render_template('alterar_senha.html', error=error, success=success)

@app.route('/cadastrar_usuario_escola', methods=['GET', 'POST'])
@login_required
def cadastrar_usuario_escola():
    if current_user.tipo_usuario_id != 2:
        return redirect(url_for('home'))

    # Obter a escola associada ao usuário logado
    escola_alocada = get_escola_alocada(current_user.id)
    if not escola_alocada:
        flash("Erro ao localizar a escola associada ao usuário.", "danger")
        return redirect(url_for('portal_administracao'))

    # Buscar os tipos de ensino da escola
    tipos_ensino = get_tipos_ensino(escola_alocada['id'])

    if request.method == 'POST':
        nome = request.form.get('nome')
        email = request.form.get('email')
        senha = request.form.get('senha')
        tipo_usuario_id = request.form.get('tipo_usuario_id')
        tipo_ensino = request.form.get('tipo_ensino')
        serie_id = request.form.get('serie_id')
        turma_id = request.form.get('turma_id')

        # Verificar campos obrigatórios
        if not all([nome, email, senha, tipo_usuario_id]):
            flash("Todos os campos obrigatórios devem ser preenchidos.", "danger")
            return render_template('cadastrar_usuario_escola.html', escola_alocada=escola_alocada, tipos_ensino=tipos_ensino)

        try:
            db = get_db()
            cursor = db.cursor()

            # Cadastrar usuário com base no tipo
            if tipo_usuario_id == "Professor":
                # Inserir o professor no banco
                cursor.execute("""
                    INSERT INTO usuarios (nome, email, senha, tipo_usuario_id, escola_id)
                    VALUES (?, ?, ?, ?, ?)
                """, (nome, email, senha, tipo_usuario_id, escola_alocada['id']))

                professor_id = cursor.lastrowid

                # Associar professor à turma
                if tipo_ensino and serie_id and turma_id:
                    cursor.execute("""
                        INSERT INTO professor_turma_escola (professor_id, escola_id, tipo_ensino, serie_id, turma_id)
                        VALUES (?, ?, ?, ?, ?)
                    """, (professor_id, escola_alocada['id'], tipo_ensino, serie_id, turma_id))

            elif tipo_usuario_id == "Aluno":
                # Inserir o aluno no banco
                cursor.execute("""
                    INSERT INTO usuarios (nome, email, senha, tipo_usuario_id, escola_id, tipo_ensino, serie_id, turma_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (nome, email, senha, tipo_usuario_id, escola_alocada['id'], tipo_ensino, serie_id, turma_id))

            elif tipo_usuario_id == "Administração da Escola":
                # Inserir um novo administrador da escola
                cursor.execute("""
                    INSERT INTO usuarios (nome, email, senha, tipo_usuario_id, escola_id)
                    VALUES (?, ?, ?, ?, ?)
                """, (nome, email, senha, tipo_usuario_id, escola_alocada['id']))

            db.commit()
            flash("Usuário cadastrado com sucesso!", "success")
            return redirect(url_for('portal_administracao'))
        except Exception as e:
            db.rollback()
            flash(f"Erro ao cadastrar usuário: {e}", "danger")

    return render_template('cadastrar_usuario_escola.html', escola_alocada=escola_alocada, tipos_ensino=tipos_ensino)

@app.route('/get_tipos_usuarios')
def get_tipos_usuarios():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id, descricao FROM tipos_usuarios")
    tipos = cursor.fetchall()
    return jsonify(tipos)

import csv
import pandas as pd
from flask import request, render_template, flash, redirect, url_for
from werkzeug.utils import secure_filename
from io import StringIO

# Função para validar os dados do CSV
def validate_escolas(data):
    required_columns = [
        'nome', 'cep', 'estado', 'cidade', 'bairro', 'endereco', 'numero',
        'telefone', 'cnpj', 'diretor', 'codigo_ibge', 'ensino_infantil', 
        'aee', 'fundamental_i', 'fundamental_ii', 'ensino_medio'
    ]

    missing_columns = [col for col in required_columns if col not in data.columns]
    if missing_columns:
        print(f"Faltam as seguintes colunas no arquivo CSV: {', '.join(missing_columns)}")
        return False

    for index, row in data.iterrows():
        if row.isnull().any():
            print(f"Erro: Valores ausentes na linha {index + 1}")
            return False

    return True


import pandas as pd
import sys  # Aqui está a importação do módulo sys para forçar os prints


@app.route("/cadastrar_escolas_massa", methods=["GET", "POST"])
@login_required
def cadastrar_escolas_massa():
    print("Entrou na rota de cadastro de escolas em massa!")
    sys.stdout.flush()  # Força a exibição no terminal

    if request.method == "POST":
        print("Formulário de cadastro de escolas em massa recebido!")
        sys.stdout.flush()

        # Verificar se o arquivo foi enviado
        if "file" not in request.files:
            flash("Nenhum arquivo enviado", "error")
            return redirect(request.url)

        file = request.files["file"]
        if file.filename == "":
            flash("Nenhum arquivo selecionado", "error")
            return redirect(request.url)

        try:
            # Lê o conteúdo do CSV e converte em um DataFrame, garantindo que os nomes das colunas sejam tratados corretamente
            data = pd.read_csv(file, dtype=str)
            data.columns = data.columns.str.strip()  # Remove espaços extras dos nomes das colunas

            print("Colunas detectadas no CSV:", list(data.columns))  # Mostra as colunas detectadas
            sys.stdout.flush()

            print("Primeiras linhas do CSV:\n", data.head())  # Mostra os primeiros registros do CSV
            sys.stdout.flush()

            # Definição das colunas esperadas
            required_columns = [
                'tipo_de_registro', 'codigo_inep', 'nome_da_escola', 'cep', 'codigo_ibge',
                'endereco', 'numero', 'complemento', 'bairro', 'ddd', 'telefone', 
                'telefone_2', 'email', 'ensino_fundamental'
            ]

            # Verificar se todas as colunas necessárias estão presentes
            missing_columns = [col for col in required_columns if col not in data.columns]

            if missing_columns:
                print("Faltando colunas:", missing_columns)  # Debug para ver quais colunas faltam
                sys.stdout.flush()
                flash(f"Faltam as seguintes colunas no arquivo CSV: {', '.join(missing_columns)}", "error")
                return redirect(request.url)

            # Converter os valores padrão corretamente
            data['tipo_de_registro'] = '00'  # Garante que sempre seja '00'
            data['ensino_fundamental'] = '1'  # Garante que sempre seja '1'

            # Renderizar a página de visualização com os dados do CSV
            return render_template("visualizar_escolas_massa.html", data=data.to_dict(orient='records'))

        except Exception as e:
            print(f"Erro ao processar o arquivo CSV: {e}")
            sys.stdout.flush()
            flash(f"Erro ao processar o arquivo CSV: {e}", "error")
            return redirect(request.url)

    return render_template("upload_escolas_massa.html")

@app.route("/confirmar_cadastro_escolas_massa", methods=["POST"])
@login_required
def confirmar_cadastro_escolas_massa():
    if current_user.tipo_usuario_id != 1:  # Apenas administradores podem cadastrar
        return redirect(url_for("home"))

    try:
        # Receber os dados enviados pelo AJAX
        data = request.get_json()
        escolas_data = data.get("escolas_data", [])

        if not escolas_data:
            flash("Nenhuma escola foi selecionada para cadastro.", "error")
            return redirect(request.url)

        db = get_db()
        cursor = db.cursor()

        print(f"Dados recebidos: {escolas_data}")  # Log dos dados recebidos

        # Iterar sobre as escolas para inserir os dados no banco
        for escola in escolas_data:
            escola_data = {
                "tipo_de_registro": escola.get('tipo_de_registro', '00'),
                "codigo_inep": escola.get('codigo_inep', '').strip(),
                "nome_da_escola": escola.get('nome_da_escola', '').strip(),
                "cep": escola.get('cep', '').strip(),
                "codigo_ibge": escola.get('codigo_ibge', '').strip(),
                "endereco": escola.get('endereco', '').strip(),
                "numero": escola.get('numero', '').strip(),
                "complemento": escola.get('complemento', '').strip(),
                "bairro": escola.get('bairro', '').strip(),
                "ddd": escola.get('ddd', '').strip(),
                "telefone": escola.get('telefone', '').strip(),
                "telefone_2": escola.get('telefone_2', '').strip(),
                "email": escola.get('email', '').strip(),
                "ensino_fundamental": escola.get('ensino_fundamental', 1)
            }

            print(f"Inserindo dados da escola: {escola_data}")

            # Inserir a escola no banco
            cursor.execute(""" 
                INSERT INTO escolas (
                    tipo_de_registro, codigo_inep, nome_da_escola, cep, codigo_ibge, 
                    endereco, numero, complemento, bairro, ddd, telefone, telefone_2, 
                    email, ensino_fundamental
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                escola_data["tipo_de_registro"], escola_data["codigo_inep"], escola_data["nome_da_escola"],
                escola_data["cep"], escola_data["codigo_ibge"], escola_data["endereco"],
                escola_data["numero"], escola_data["complemento"], escola_data["bairro"],
                escola_data["ddd"], escola_data["telefone"], escola_data["telefone_2"],
                escola_data["email"], escola_data["ensino_fundamental"]
            ))

            # Capturar o ID gerado pelo banco para a escola recém-inserida
            escola_id = cursor.lastrowid
            print(f"ID da escola gerado: {escola_id}")

            # Aqui você pode usar o `escola_id` para outros relacionamentos, se necessário

        # Confirmar as transações no banco
        db.commit()

        flash("Escolas cadastradas com sucesso!", "success")
        return jsonify({"status": "success"}), 200  # Resposta de sucesso para AJAX

    except Exception as e:
        print(f"Erro ao processar o cadastro das escolas: {e}")
        flash(f"Erro ao cadastrar escolas: {str(e)}", "error")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/portal_administrador/cadastrar_turmas_massa", methods=["GET", "POST"])
@login_required
def cadastrar_turmas_massa():
    if current_user.tipo_usuario_id != 1:  # Apenas administradores podem cadastrar
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    if request.method == "POST":
        if "file" not in request.files:
            flash("Nenhum arquivo enviado", "error")
            return redirect(request.url)

        file = request.files["file"]
        if file.filename == "":
            flash("Nenhum arquivo selecionado", "error")
            return redirect(request.url)

        if file and allowed_file(file.filename):
            try:
                # Ler o arquivo como UTF-8
                file_content = file.stream.read().decode("utf-8")
                data = pd.read_csv(StringIO(file_content), dtype=str)  # Certifique-se de que todos os valores são strings
                print("Dados do arquivo carregados:", data.head())

                # Processar os dados
                for index, turma in data.iterrows():
                    codigo_inep = turma['Codigo_inep'].strip()
                    tipo_ensino_nome = turma['Tipo_ensino_id'].strip()
                    serie_nome = turma['serie_id'].strip()
                    turma_nome = turma['Turma'].strip()

                    # Buscar o ID da escola com base no Código INEP
                    cursor.execute("SELECT id, codigo_inep FROM escolas WHERE codigo_inep = ?", (codigo_inep,))
                    escola = cursor.fetchone()
                    if escola:
                        data.at[index, 'escola_id'] = int(escola[0])  # Converte para inteiro
                        data.at[index, 'codigo_inep'] = escola[1]
                    else:
                        data.at[index, 'escola_id'] = None
                        data.at[index, 'codigo_inep'] = None

                    # Buscar o ID do tipo de ensino com base no nome
                    cursor.execute("SELECT id FROM tipos_ensino WHERE nome = ?", (tipo_ensino_nome,))
                    tipo_ensino_id = cursor.fetchone()
                    if tipo_ensino_id:
                        data.at[index, 'tipo_ensino_id'] = int(tipo_ensino_id[0])  # Converte para inteiro
                    else:
                        print(f"Tipo de ensino '{tipo_ensino_nome}' não encontrado no banco de dados.")
                        data.at[index, 'tipo_ensino_id'] = None

                    # Buscar o ID da série com base no nome
                    cursor.execute("SELECT id FROM series WHERE nome = ?", (serie_nome,))
                    serie_id = cursor.fetchone()
                    if serie_id:
                        data.at[index, 'serie_id'] = int(serie_id[0])  # Converte para inteiro
                    else:
                        print(f"Série '{serie_nome}' não encontrada no banco de dados.")
                        data.at[index, 'serie_id'] = None

                    # Validar a letra da turma
                    if turma_nome and turma_nome.strip():
                        data.at[index, 'turma'] = turma_nome
                    else:
                        data.at[index, 'turma'] = None

                # Converter IDs para inteiros
                data['escola_id'] = data['escola_id'].astype('Int64')  # Usa Int64 para permitir valores nulos
                data['tipo_ensino_id'] = data['tipo_ensino_id'].astype('Int64')
                data['serie_id'] = data['serie_id'].astype('Int64')

                print("Dados após conversão de IDs:", data)

                # Passar os dados para o template de visualização
                return render_template("visualizar_turmas_massa.html", data=data.to_dict(orient="records"))

            except Exception as e:
                print(f"Erro ao processar o arquivo: {e}")
                flash(f"Erro ao processar o arquivo: {e}", "error")
                return redirect(request.url)

    return render_template("upload_turmas_massa.html")



@app.route("/confirmar_cadastro_turmas", methods=["POST"])
@login_required
def confirmar_cadastro_turmas():
    if current_user.tipo_usuario_id != 1:
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    try:
        turmas_data = request.get_json()
        if not turmas_data:
            return jsonify({"status": "error", "message": "Nenhuma turma foi selecionada para cadastro."})

        for turma in turmas_data:
            escola_id = turma.get('escola_id')
            codigo_inep = turma.get('codigo_inep')  # Agora enviado corretamente
            tipo_ensino_id = turma.get('tipo_ensino_id')
            serie_id = turma.get('serie_id')
            turma_nome = turma.get('turma')

            # Verificar se todos os dados necessários estão presentes
            if not all([escola_id, codigo_inep, tipo_ensino_id, serie_id, turma_nome]):
                print(f"Dados incompletos para a turma: {turma}")
                continue

            # Inserir a turma no banco de dados
            cursor.execute("""
                INSERT INTO turmas (tipo_de_registro, codigo_inep, escola_id, tipo_ensino_id, serie_id, turma)
                VALUES ('20', ?, ?, ?, ?, ?)
            """, (codigo_inep, escola_id, tipo_ensino_id, serie_id, turma_nome))

            print(f"Turma cadastrada com sucesso: {turma_nome}")

        db.commit()
        return jsonify({"status": "success", "message": "Cadastro realizado com sucesso!"})

    except Exception as e:
        db.rollback()
        print(f"Erro ao processar o cadastro das turmas: {e}")
        return jsonify({"status": "error", "message": f"Erro ao cadastrar turmas: {str(e)}"})



@app.route("/portal_administrador/cadastrar_usuarios_massa", methods=["GET", "POST"])
@login_required
def cadastrar_usuarios_massa():
    if current_user.tipo_usuario_id != 1:  # Apenas administradores podem acessar
        return redirect(url_for("home"))

    db = get_db()
    cursor = db.cursor()

    if request.method == "POST":
        if "file" not in request.files:
            flash("Nenhum arquivo enviado.", "error")
            return redirect(request.url)

        file = request.files["file"]
        if file.filename == "":
            flash("Nenhum arquivo selecionado.", "error")
            return redirect(request.url)

        if file and allowed_file(file.filename):
            try:
                # Lê o conteúdo do arquivo CSV
                file_content = file.stream.read().decode("utf-8")
                
                # Modificação aqui: configurações específicas para leitura do CSV
                data = pd.read_csv(
                    StringIO(file_content),
                    sep=',',
                    engine='python',
                    quoting=csv.QUOTE_NONE,
                    on_bad_lines='skip',
                    escapechar='\\',
                    skipinitialspace=True
                )
                
                # Limpar nomes das colunas
                data.columns = [col.strip().strip('"') for col in data.columns]
                
                # Limpar valores das colunas de texto
                for col in data.columns:
                    if data[col].dtype == 'object':
                        data[col] = data[col].str.strip().str.strip('"')
                
                print("Dados brutos do arquivo:")
                print(data.to_string())  # Mostra todos os dados sem truncar
                print("\nColunas encontradas:", data.columns.tolist())
                
                # Verificar se as colunas necessárias estão no arquivo
                colunas_necessarias = [
                    "tipo_registro", "codigo_inep_escola", "cpf",
                    "nome", "data_nascimento", "mae", "pai", "sexo", "codigo_ibge",
                    "cep", "escola_id", "tipo_ensino_id", "serie_id", "turma_id", 
                    "codigo_ibge", "tipo_usuario_id"
                ]
                
                # Tentar corrigir o DataFrame se ele tiver apenas uma coluna
                if len(data.columns) == 1:
                    # Dividir a única coluna em várias usando o separador ','
                    data = pd.DataFrame([x.split(',') for x in data[data.columns[0]]], columns=colunas_necessarias)
                
                colunas_presentes = set(data.columns)
                
                if not set(colunas_necessarias).issubset(colunas_presentes):
                    missing_columns = set(colunas_necessarias) - colunas_presentes
                    flash(f"Colunas faltando no arquivo: {', '.join(missing_columns)}", "error")
                    return redirect(request.url)

                # Conversão e validação dos IDs e outros campos
                for index, usuario in data.iterrows():
                    try:
                        print(f"\nProcessando usuário: {usuario['nome']}")  # Log para cada usuário
                        tipo_usuario = int(usuario["tipo_usuario_id"])

                        # Para alunos e professores, buscar IDs das referências
                        if tipo_usuario in [3, 4]:
                            # Tentar buscar escola pelo nome primeiro
                            print(f"Buscando escola pelo nome: {usuario['escola_id']}")
                            cursor.execute("SELECT id FROM escolas WHERE nome_da_escola = ? OR codigo_inep = ?", (usuario["escola_id"], usuario["codigo_inep_escola"]))
                            escola_result = cursor.fetchone()
                            
                            # Se não encontrar pelo nome, tentar pelo código INEP
                            if not escola_result:
                                print(f"Escola não encontrada pelo nome, tentando código INEP: {usuario['codigo_inep_escola']}")
                                cursor.execute("SELECT id FROM escolas WHERE codigo_inep = ?", (usuario["codigo_inep_escola"],))
                                escola_result = cursor.fetchone()
                            
                            escola_id = escola_result[0] if escola_result else None
                            print(f"ID da escola encontrado: {escola_id}")

                            # Buscar tipo de ensino pela descrição
                            print(f"Buscando tipo de ensino: {usuario['tipo_ensino_id']}")
                            cursor.execute("SELECT id FROM tipos_ensino WHERE nome = ?", (usuario["tipo_ensino_id"],))
                            tipo_ensino_id = cursor.fetchone()
                            tipo_ensino_id = tipo_ensino_id[0] if tipo_ensino_id else None
                            print(f"ID do tipo de ensino encontrado: {tipo_ensino_id}")

                            # Buscar série
                            print(f"Buscando série: {usuario['serie_id']}")
                            cursor.execute("SELECT id FROM series WHERE nome = ?", (usuario["serie_id"],))
                            serie_id = cursor.fetchone()
                            serie_id = serie_id[0] if serie_id else None
                            print(f"ID da série encontrado: {serie_id}")

                            # Buscar turma
                            print(f"Buscando turma: {usuario['turma_id']}")
                            cursor.execute(
                                "SELECT id FROM turmas WHERE turma = ? AND escola_id = ? AND serie_id = ?",
                                (usuario["turma_id"], escola_id, serie_id)
                            )
                            turma_id = cursor.fetchone()
                            turma_id = turma_id[0] if turma_id else None
                            print(f"ID da turma encontrado: {turma_id}")

                            if not all([escola_id, tipo_ensino_id, serie_id, turma_id]):
                                print("AVISO: Alguns IDs não foram encontrados:")
                                print(f"- Escola ID: {escola_id}")
                                print(f"- Tipo Ensino ID: {tipo_ensino_id}")
                                print(f"- Série ID: {serie_id}")
                                print(f"- Turma ID: {turma_id}")
                        else:
                            escola_id, tipo_ensino_id, serie_id, turma_id = None, None, None, None

                        # Atualizar os valores no DataFrame
                        data.at[index, "escola_id"] = escola_id
                        data.at[index, "tipo_ensino_id"] = tipo_ensino_id
                        data.at[index, "serie_id"] = serie_id
                        data.at[index, "turma_id"] = turma_id

                        # Gerar senha padrão para cada usuário
                        data.at[index, "senha"] = "123456"  # Senha padrão para todos os usuários

                        # Gerar email padrão para cada usuário usando o CPF
                        if pd.notna(usuario["cpf"]) and str(usuario["cpf"]).strip():
                            cpf = str(usuario["cpf"]).strip()
                            data.at[index, "email"] = f"{cpf}@aluno.edu.br"

                        print(f"IDs convertidos: escola_id={escola_id}, tipo_ensino_id={tipo_ensino_id}, serie_id={serie_id}, turma_id={turma_id}")

                    except Exception as e:
                        print(f"Erro ao processar o usuário {usuario['nome']}: {e}")
                        print(f"Dados do usuário que causou erro: {usuario}")

                print("\nDados após conversão de IDs:", data.head())  # Log final

                # Renderizar pré-visualização dos dados
                return render_template("visualizar_usuarios_massa.html", data=data.to_dict(orient="records"))

            except Exception as e:
                print(f"Erro ao processar o arquivo: {e}")
                flash(f"Erro ao processar o arquivo: {e}", "error")
                return redirect(request.url)

    return render_template("upload_usuarios_massa.html")


@app.route("/confirmar_cadastro_usuarios", methods=["POST"])
@login_required
def confirmar_cadastro_usuarios():
    if current_user.tipo_usuario_id != 1:
        print("Erro: Usuário sem permissão")
        return jsonify({"status": "error", "message": "Você não tem permissão para realizar esta ação."})

    try:
        print("1. Iniciando processo de confirmação de cadastro")
        
        # Receber os dados enviados via AJAX
        data = request.get_json()
        usuarios_data = data.get("usuarios", [])

        print(f"2. Dados recebidos do AJAX: {data}")
        print(f"3. Lista de usuários para cadastro: {usuarios_data}")

        if not usuarios_data:
            print("Erro: Nenhum usuário recebido para cadastro")
            return jsonify({"status": "error", "message": "Nenhum usuário selecionado."})

        db = get_db()
        cursor = db.cursor()

        for usuario in usuarios_data:
            try:
                print(f"\n4. Iniciando cadastro do usuário: {usuario.get('nome', 'Nome não fornecido')}")
                print(f"5. Dados completos do usuário: {usuario}")

                tipo_usuario = int(usuario["tipo_usuario_id"])
                print(f"6. Tipo de usuário: {tipo_usuario}")

                # Verifica se o usuário já existe
                if "cpf" in usuario and usuario["cpf"]:
                    print(f"7a. Verificando existência por CPF: {usuario['cpf']}")
                    cursor.execute("SELECT id FROM usuarios WHERE cpf = ?", (usuario["cpf"],))
                else:
                    print(f"7b. Verificando existência por email: {usuario.get('email', 'Email não fornecido')}")
                    cursor.execute("SELECT id FROM usuarios WHERE email = ?", (usuario["email"],))
                
                usuario_existente = cursor.fetchone()
                print(f"8. Usuário existente? {usuario_existente is not None}")

                if usuario_existente:
                    usuario_id = usuario_existente[0]
                    print(f"9a. Usuário já existe com ID: {usuario_id}")
                else:
                    print("9b. Preparando para inserir novo usuário")
                    print(f"10. Valores para inserção:")
                    valores = (
                        usuario["nome"],
                        usuario.get("email"),
                        generate_password_hash(usuario["senha"]),
                        usuario["tipo_usuario_id"],
                        usuario.get("escola_id"),
                        usuario.get("turma_id"),
                        usuario.get("serie_id"),
                        usuario.get("tipo_ensino_id"),
                        usuario.get("cep"),
                        usuario.get("codigo_ibge"),
                        usuario.get("cpf"),
                        usuario.get("data_nascimento"),
                        usuario.get("mae"),
                        usuario.get("pai"),
                        usuario.get("sexo"),
                        usuario.get("codigo_ibge")
                    )
                    print(f"Valores: {valores}")

                    cursor.execute(
                        """
                        INSERT INTO usuarios (
                            nome, email, senha, tipo_usuario_id, escola_id, turma_id,
                            serie_id, tipo_ensino_id, cep, codigo_ibge, cpf, data_nascimento,
                            mae, pai, sexo, codigo_ibge
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        valores
                    )
                    db.commit()
                    usuario_id = cursor.lastrowid
                    print(f"11. Novo usuário inserido com ID: {usuario_id}")

                # Se for professor, associá-lo à tabela professor_turma_escola
                if tipo_usuario == 3:
                    print("12. Usuário é professor, vinculando à turma")
                    cursor.execute("SELECT COUNT(*) FROM professor_turma_escola WHERE professor_id = ?", (usuario_id,))
                    if cursor.fetchone()[0] == 0:
                        print("13. Inserindo vínculo professor-turma")
                        cursor.execute(
                            """
                            INSERT INTO professor_turma_escola (professor_id, escola_id, turma_id, tipo_ensino_id, serie_id)
                            VALUES (?, ?, ?, ?, ?)
                            """,
                            (usuario_id, usuario["escola_id"], usuario["turma_id"], usuario["tipo_ensino_id"], usuario["serie_id"]),
                        )
                        db.commit()
                        print("14. Vínculo professor-turma criado com sucesso")

            except Exception as e:
                db.rollback()
                print(f"ERRO ao cadastrar usuário {usuario.get('nome', 'Nome não fornecido')}")
                print(f"Detalhes do erro: {str(e)}")
                print(f"Dados que causaram erro: {usuario}")
                raise  # Re-lança a exceção para ser capturada pelo try/except externo

        print("15. Processo finalizado com sucesso")
        return jsonify({"status": "success", "message": "Usuários cadastrados com sucesso!"})

    except Exception as e:
        db.rollback()
        erro_msg = f"Erro ao processar cadastro: {str(e)}"
        print(f"ERRO FATAL: {erro_msg}")
        return jsonify({"status": "error", "message": erro_msg})
# Inicialização do Banco e Servidor

if __name__ == "__main__":
    init_db()
    app.run(debug=True)

